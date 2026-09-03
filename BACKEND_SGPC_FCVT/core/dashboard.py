# core/dashboard.py
# ============================================================
# SGPC ULEAM — Dashboard institucional + informes PDF / Excel
# ============================================================

from io import BytesIO
import unicodedata

from django.apps import apps
from django.db.models import Count, F, Q
from django.http import HttpResponse
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo

try:
    from reportlab.graphics.charts.barcharts import HorizontalBarChart, VerticalBarChart
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.shapes import Circle, Drawing, Rect, String
    from reportlab.graphics.widgets.markers import makeMarker
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        CondPageBreak,
        KeepTogether,
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

from rest_framework.exceptions import ValidationError as DRFValidationError
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.authentication import (
    JWTAuthentication,
)


# ============================================================
# CONFIGURACIÓN GENERAL
# ============================================================

TOP_DEFAULT = 10

TOP_ALLOWED = {
    5,
    10,
    15,
    20,
}


MONTH_LABELS_ES = {
    1: "Ene",
    2: "Feb",
    3: "Mar",
    4: "Abr",
    5: "May",
    6: "Jun",
    7: "Jul",
    8: "Ago",
    9: "Sep",
    10: "Oct",
    11: "Nov",
    12: "Dic",
}


MONTH_NAMES_ES = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}


# ============================================================
# TIPOS CANÓNICOS DEL DASHBOARD
#
# IMPORTANTE:
#
# Estos códigos se conservan porque forman parte del contrato
# actual entre dashboard y frontend.
#
# No corresponden necesariamente al campo TipoPublicacion.codigo.
#
# AAI -> Artículo de alto impacto
# AR  -> Artículo regional
# PON -> Ponencia
# CAP -> Capítulo de libro
# LIB -> Libro
# ============================================================

CANONICAL_TYPES = (
    {
        "codigo": "AAI",
        "nombre": "Artículo de alto impacto",
        "categoria": "articulo",
        "orden": 1,
    },
    {
        "codigo": "AR",
        "nombre": "Artículo regional",
        "categoria": "articulo",
        "orden": 2,
    },
    {
        "codigo": "PON",
        "nombre": "Ponencia",
        "categoria": "ponencia",
        "orden": 3,
    },
    {
        "codigo": "CAP",
        "nombre": "Capítulo de libro",
        "categoria": "capitulo",
        "orden": 4,
    },
    {
        "codigo": "LIB",
        "nombre": "Libro",
        "categoria": "libro",
        "orden": 5,
    },
)




# ============================================================
# ALIASES
#
# Esto permite que el dashboard acepte tanto:
#
# ?tipo_codigo=AAI
#
# como:
#
# ?tipo_codigo=articulo_alto_impacto
#
# sin romper el frontend existente.
# ============================================================

CANONICAL_ALIASES = {
    # Alto impacto
    "AAI": "AAI",
    "ALTO_IMPACTO": "AAI",
    "ARTICULO_ALTO_IMPACTO": "AAI",

    # Regional
    "AR": "AR",
    "REGIONAL": "AR",
    "ARTICULO_REGIONAL": "AR",

    # Ponencia
    "PON": "PON",
    "PONENCIA": "PON",

    # Capítulo
    "CAP": "CAP",
    "CAPITULO": "CAP",
    "CAPITULO_LIBRO": "CAP",

    # Libro
    "LIB": "LIB",
    "LIBRO": "LIB",
}


# ============================================================
# MODEL RESOLVERS
# ============================================================


def get_publicacion_model():
    return apps.get_model(
        "core",
        "Publicacion",
    )


def get_articulo_model():
    return apps.get_model(
        "core",
        "Articulo",
    )


def get_ponencia_model():
    return apps.get_model(
        "core",
        "Ponencia",
    )


def get_libro_model():
    return apps.get_model(
        "core",
        "Libro",
    )


def get_capitulo_libro_model():
    return apps.get_model(
        "core",
        "CapituloLibro",
    )


def get_publicacion_autor_model():
    return apps.get_model(
        "core",
        "PublicacionAutor",
    )


def get_facultad_model():
    return apps.get_model(
        "core",
        "Facultad",
    )


def get_carrera_model():
    return apps.get_model(
        "core",
        "Carrera",
    )


def get_sede_model():
    return apps.get_model(
        "core",
        "Sede",
    )


# ============================================================
# HELPERS GENERALES
# ============================================================


def _safe_int(value):
    try:
        return int(value)

    except (
        TypeError,
        ValueError,
    ):
        return None


def _safe_top(value):
    top = (
        _safe_int(value)
        or TOP_DEFAULT
    )

    return (
        top
        if top in TOP_ALLOWED
        else TOP_DEFAULT
    )


def _label(
    value,
    fallback="Sin dato",
):
    value = str(
        value or ""
    ).strip()

    return (
        value
        if value
        else fallback
    )


def _normalize_canonical_code(
    value,
):
    """
    Normaliza códigos y aliases enviados desde frontend.

    Ejemplos:

        AAI
        articulo_alto_impacto
        alto-impacto

    -> AAI
    """

    raw = str(
        value or ""
    ).strip()

    if not raw:
        return None

    normalized = (
        raw.upper()
        .replace("-", "_")
        .replace(" ", "_")
    )

    return (
        CANONICAL_ALIASES.get(
            normalized
        )
    )


def _canonical_meta(
    code,
):
    normalized = (
        _normalize_canonical_code(
            code
        )
    )

    if not normalized:
        return None

    return next(
        (
            item
            for item in CANONICAL_TYPES
            if item["codigo"] == normalized
        ),
        None,
    )



def _normalize_year_range(
    anio_desde,
    anio_hasta,
):
    if (
        anio_desde
        and anio_hasta
        and anio_desde > anio_hasta
    ):
        return (
            anio_hasta,
            anio_desde,
        )

    return (
        anio_desde,
        anio_hasta,
    )


def _safe_period_month(
    value,
    *,
    field_name,
):
    """
    Convierte un período mensual a una tupla (año, mes).

    Formato canónico:

        YYYY-MM

    Por compatibilidad temporal también acepta YYYY-MM-DD,
    pero el día se ignora.
    """

    raw = str(
        value or ""
    ).strip()

    if not raw:
        return None

    parts = raw.split("-")

    if len(parts) not in {
        2,
        3,
    }:
        raise DRFValidationError(
            {
                field_name: (
                    "Utilice el formato YYYY-MM."
                )
            }
        )

    try:
        year = int(
            parts[0]
        )

        month = int(
            parts[1]
        )

    except (
        TypeError,
        ValueError,
        OverflowError,
    ):
        raise DRFValidationError(
            {
                field_name: (
                    "Utilice el formato YYYY-MM."
                )
            }
        )

    if (
        year <= 0
        or month < 1
        or month > 12
    ):
        raise DRFValidationError(
            {
                field_name: (
                    "Utilice un año válido y "
                    "un mes entre 01 y 12."
                )
            }
        )

    return (
        year,
        month,
    )


def _period_month_to_string(
    value,
):
    if not value:
        return None

    year, month = value

    return (
        f"{int(year):04d}-"
        f"{int(month):02d}"
    )


def _resolve_period_filters(
    params,
):
    """
    Resuelve los filtros temporales del dashboard.

    Contrato canónico:

        mes_desde=YYYY-MM
        mes_hasta=YYYY-MM

    Los aliases históricos de período se aceptan de forma
    temporal para no romper clientes anteriores.

    Cuando existe un rango mensual, los registros sin mes se
    excluyen porque no pueden ubicarse con precisión.
    """

    raw_mes_desde = (
        params.get(
            "mes_desde"
        )
        or params.get(
            "fecha_desde"
        )
    )

    raw_mes_hasta = (
        params.get(
            "mes_hasta"
        )
        or params.get(
            "fecha_hasta"
        )
    )

    mes_desde = _safe_period_month(
        raw_mes_desde,
        field_name="mes_desde",
    )

    mes_hasta = _safe_period_month(
        raw_mes_hasta,
        field_name="mes_hasta",
    )

    if (
        mes_desde
        and mes_hasta
        and mes_desde > mes_hasta
    ):
        (
            mes_desde,
            mes_hasta,
        ) = (
            mes_hasta,
            mes_desde,
        )

    if (
        mes_desde
        or mes_hasta
    ):
        return {
            "mes_desde": mes_desde,
            "mes_hasta": mes_hasta,
            "anio_desde": None,
            "anio_hasta": None,
            "modo": "mes",
        }

    anio_desde = _safe_int(
        params.get(
            "anio_desde"
        )
    )

    anio_hasta = _safe_int(
        params.get(
            "anio_hasta"
        )
    )

    (
        anio_desde,
        anio_hasta,
    ) = _normalize_year_range(
        anio_desde,
        anio_hasta,
    )

    return {
        "mes_desde": None,
        "mes_hasta": None,
        "anio_desde": anio_desde,
        "anio_hasta": anio_hasta,
        "modo": (
            "anio"
            if (
                anio_desde
                or anio_hasta
            )
            else None
        ),
    }



# ============================================================
# QUERYSETS BASE Y FILTROS
# ============================================================


def _build_base_queryset(
    params,
    *,
    apply_period=True,
):
    Publicacion = (
        get_publicacion_model()
    )

    # El dashboard institucional representa producción científica
    # validada. Los estados operativos (borrador, en revisión,
    # observada y rechazada) pertenecen al dashboard de gestión.
    queryset = (
        Publicacion.objects
        .filter(
            estado=(
                Publicacion.ESTADO_APROBADA
            )
        )
        .select_related(
            "tipo",

            "sede",

            "carrera",
            "carrera__facultad",

            "area",
            "subarea",

            "pais",
            "ciudad",

            "proyecto",
            "proyecto__sede",
        )
    )

    sede_id = _safe_int(
        params.get(
            "sede_id"
        )
        or params.get(
            "sede"
        )
    )

    facultad_id = _safe_int(
        params.get(
            "facultad_id"
        )
    )

    carrera_id = _safe_int(
        params.get(
            "carrera_id"
        )
    )

    period = (
        _resolve_period_filters(
            params
        )
        if apply_period
        else {
            "mes_desde": None,
            "mes_hasta": None,
            "anio_desde": None,
            "anio_hasta": None,
            "modo": None,
        }
    )

    # --------------------------------------------------------
    # Sede
    #
    # Publicacion posee una relación institucional directa con
    # Sede. Se acepta ``sede`` como alias temporal de
    # ``sede_id`` para mantener compatibilidad entre clientes.
    # --------------------------------------------------------

    if sede_id:
        queryset = queryset.filter(
            sede_id=sede_id
        )

    # --------------------------------------------------------
    # Facultad
    #
    # Publicacion NO tiene facultad directa.
    #
    # Publicacion -> Carrera -> Facultad
    # --------------------------------------------------------

    if facultad_id:
        queryset = queryset.filter(
            carrera__facultad_id=(
                facultad_id
            )
        )

    # --------------------------------------------------------
    # Carrera
    # --------------------------------------------------------

    if carrera_id:
        queryset = queryset.filter(
            carrera_id=carrera_id
        )

    # --------------------------------------------------------
    # Período mensual sobre año + mes.
    # --------------------------------------------------------

    if (
        period["mes_desde"]
        or period["mes_hasta"]
    ):
        queryset = queryset.exclude(
            mes_publicacion__isnull=True
        )

    if period["mes_desde"]:
        (
            anio_desde,
            mes_desde,
        ) = period[
            "mes_desde"
        ]

        queryset = queryset.filter(
            Q(
                anio_publicacion__gt=(
                    anio_desde
                )
            )
            | Q(
                anio_publicacion=(
                    anio_desde
                ),
                mes_publicacion__gte=(
                    mes_desde
                ),
            )
        )

    if period["mes_hasta"]:
        (
            anio_hasta,
            mes_hasta,
        ) = period[
            "mes_hasta"
        ]

        queryset = queryset.filter(
            Q(
                anio_publicacion__lt=(
                    anio_hasta
                )
            )
            | Q(
                anio_publicacion=(
                    anio_hasta
                ),
                mes_publicacion__lte=(
                    mes_hasta
                ),
            )
        )

    # --------------------------------------------------------
    # Compatibilidad con filtros por rango de años.
    #
    # Solo se aplican cuando no existe rango mensual.
    # --------------------------------------------------------

    if period["anio_desde"]:
        queryset = queryset.filter(
            anio_publicacion__gte=(
                period["anio_desde"]
            )
        )

    if period["anio_hasta"]:
        queryset = queryset.filter(
            anio_publicacion__lte=(
                period["anio_hasta"]
            )
        )

    return queryset



# ============================================================
# RESOLUCIÓN DE TIPOS CANÓNICOS
# ============================================================


def _build_canonical_id_sources(
    publicaciones,
):
    Articulo = (
        get_articulo_model()
    )

    Ponencia = (
        get_ponencia_model()
    )

    Libro = (
        get_libro_model()
    )

    CapituloLibro = (
        get_capitulo_libro_model()
    )

    return {
        "AAI": (
            Articulo.objects
            .filter(
                publicacion__in=(
                    publicaciones
                ),
                tipo_articulo=(
                    "alto_impacto"
                ),
            )
            .values_list(
                "publicacion_id",
                flat=True,
            )
            .distinct()
        ),

        "AR": (
            Articulo.objects
            .filter(
                publicacion__in=(
                    publicaciones
                ),
                tipo_articulo="regional",
            )
            .values_list(
                "publicacion_id",
                flat=True,
            )
            .distinct()
        ),

        "PON": (
            Ponencia.objects
            .filter(
                publicacion__in=(
                    publicaciones
                )
            )
            .values_list(
                "publicacion_id",
                flat=True,
            )
            .distinct()
        ),

        "CAP": (
            CapituloLibro.objects
            .filter(
                publicacion__in=(
                    publicaciones
                )
            )
            .values_list(
                "publicacion_id",
                flat=True,
            )
            .distinct()
        ),

        "LIB": (
            Libro.objects
            .filter(
                publicacion__in=(
                    publicaciones
                )
            )
            .values_list(
                "publicacion_id",
                flat=True,
            )
            .distinct()
        ),
    }


def _canonical_queryset(
    publicaciones,
    code,
    id_sources=None,
):
    normalized = (
        _normalize_canonical_code(
            code
        )
    )

    if not normalized:
        return publicaciones.none()

    if id_sources is None:
        id_sources = (
            _build_canonical_id_sources(
                publicaciones
            )
        )

    source = id_sources.get(
        normalized
    )

    if source is None:
        return publicaciones.none()

    return publicaciones.filter(
        id__in=source
    )



def _apply_canonical_type_filter(
    publicaciones,
    tipo_codigo=None,
    id_sources=None,
):
    normalized = (
        _normalize_canonical_code(
            tipo_codigo
        )
    )

    if not normalized:
        return publicaciones

    return _canonical_queryset(
        publicaciones,
        normalized,
        id_sources=id_sources,
    )


def _count_by_canonical_type(
    publicaciones,
    id_sources=None,
):
    if id_sources is None:
        id_sources = (
            _build_canonical_id_sources(
                publicaciones
            )
        )

    counts = {}

    for item in CANONICAL_TYPES:
        code = item["codigo"]

        counts[code] = (
            _canonical_queryset(
                publicaciones,
                code,
                id_sources=id_sources,
            )
            .count()
        )

    return counts


# ============================================================
# PUBLICACIONES POR AÑO
# ============================================================


def _build_publicaciones_por_anio(
    publicaciones,
):
    rows = (
        publicaciones
        .exclude(
            anio_publicacion__isnull=True
        )
        .values(
            "anio_publicacion"
        )
        .annotate(
            total=Count("id")
        )
        .order_by(
            "anio_publicacion"
        )
    )

    return [
        {
            "label": str(
                row[
                    "anio_publicacion"
                ]
            ),
            "value": int(
                row["total"]
                or 0
            ),
        }
        for row in rows
    ]


# ============================================================
# PUBLICACIONES POR MES
# ============================================================


def _resolve_anio_base_mensual(
    publicaciones,
    explicit_year=None,
):
    if explicit_year:
        return explicit_year

    ultimo_anio_con_mes = (
        publicaciones
        .exclude(
            mes_publicacion__isnull=True
        )
        .values_list(
            "anio_publicacion",
            flat=True,
        )
        .distinct()
        .order_by(
            "-anio_publicacion"
        )
        .first()
    )

    if ultimo_anio_con_mes:
        return ultimo_anio_con_mes

    return (
        publicaciones
        .values_list(
            "anio_publicacion",
            flat=True,
        )
        .order_by(
            "-anio_publicacion"
        )
        .first()
    )


def _build_publicaciones_por_mes(
    publicaciones,
    explicit_year=None,
):
    anio_base = (
        _resolve_anio_base_mensual(
            publicaciones,
            explicit_year=(
                explicit_year
            ),
        )
    )

    if anio_base:
        qs_anio = (
            publicaciones.filter(
                anio_publicacion=(
                    anio_base
                )
            )
        )
    else:
        qs_anio = (
            publicaciones.none()
        )

    qs_con_mes = (
        qs_anio.exclude(
            mes_publicacion__isnull=True
        )
    )

    total_publicaciones_anio = (
        qs_anio.count()
    )

    total_con_mes = (
        qs_con_mes.count()
    )

    total_sin_mes = max(
        total_publicaciones_anio
        - total_con_mes,
        0,
    )

    rows = (
        qs_con_mes
        .values(
            "mes_publicacion"
        )
        .annotate(
            total=Count("id")
        )
        .order_by(
            "mes_publicacion"
        )
    )

    totals = {
        row[
            "mes_publicacion"
        ]: int(
            row["total"]
            or 0
        )
        for row in rows
        if row[
            "mes_publicacion"
        ]
    }

    return {
        "anio_base": (
            anio_base
        ),

        "items": [
            {
                "mes": i,
                "label": (
                    MONTH_LABELS_ES[i]
                ),
                "value": (
                    totals.get(
                        i,
                        0,
                    )
                ),
            }
            for i in range(
                1,
                13,
            )
        ],

        "total_publicaciones_anio": (
            total_publicaciones_anio
        ),

        "total_con_mes": (
            total_con_mes
        ),

        "total_sin_mes": (
            total_sin_mes
        ),
    }


# ============================================================
# PUBLICACIONES POR TIPO
# ============================================================


def _build_publicaciones_por_tipo(
    publicaciones,
    selected_tipo_codigo=None,
    id_sources=None,
):
    total_publicaciones = (
        publicaciones.count()
    )

    counts = (
        _count_by_canonical_type(
            publicaciones,
            id_sources=id_sources,
        )
    )

    all_items = []

    for item in CANONICAL_TYPES:
        code = item["codigo"]

        total = int(
            counts.get(
                code,
                0,
            )
            or 0
        )

        porcentaje = (
            round(
                (
                    total
                    / total_publicaciones
                )
                * 100,
                2,
            )
            if total_publicaciones
            else 0
        )

        all_items.append(
            {
                "tipo_id": code,
                "tipo_codigo": code,
                "tipo_nombre": (
                    item["nombre"]
                ),
                "categoria": (
                    item["categoria"]
                ),
                "total": total,
                "porcentaje": (
                    porcentaje
                ),
            }
        )

    items = [
        item
        for item in all_items
        if item["total"] > 0
    ]

    selected_code = (
        _normalize_canonical_code(
            selected_tipo_codigo
        )
    )

    seleccionado = None

    if selected_code:
        seleccionado = next(
            (
                item
                for item in all_items
                if (
                    item["tipo_codigo"]
                    == selected_code
                )
            ),
            None,
        )

    return {
        "total_publicaciones": (
            total_publicaciones
        ),
        "seleccionado": (
            seleccionado
        ),
        "items": items,
    }


# ============================================================
# TIPO POR AÑO
# ============================================================


def _build_publicaciones_por_tipo_anual(
    publicaciones,
    id_sources=None,
):
    categorias = list(
        publicaciones
        .exclude(
            anio_publicacion__isnull=True
        )
        .values_list(
            "anio_publicacion",
            flat=True,
        )
        .distinct()
        .order_by(
            "anio_publicacion"
        )
    )

    if not categorias:
        return {
            "categorias": [],
            "series": [],
            "total_publicaciones": 0,
        }

    if id_sources is None:
        id_sources = (
            _build_canonical_id_sources(
                publicaciones
            )
        )

    series = []

    for item in CANONICAL_TYPES:
        code = item["codigo"]

        rows = (
            _canonical_queryset(
                publicaciones,
                code,
                id_sources=id_sources,
            )
            .exclude(
                anio_publicacion__isnull=True
            )
            .values(
                "anio_publicacion"
            )
            .annotate(
                total=Count("id")
            )
            .order_by(
                "anio_publicacion"
            )
        )

        totals_by_year = {
            row["anio_publicacion"]: int(
                row["total"]
                or 0
            )
            for row in rows
        }

        data = [
            totals_by_year.get(
                anio,
                0,
            )
            for anio in categorias
        ]

        if sum(data) <= 0:
            continue

        series.append(
            {
                "id": code,
                "codigo": code,
                "label": (
                    item["nombre"]
                ),
                "categoria": (
                    item["categoria"]
                ),
                "data": data,
            }
        )

    return {
        "categorias": [
            str(anio)
            for anio in categorias
        ],

        "series": series,

        "total_publicaciones": (
            publicaciones.count()
        ),
    }


# ============================================================
# ÁREAS
# ============================================================


def _build_top_areas(
    publicaciones,
    limit,
):
    rows = (
        publicaciones
        .exclude(
            area__isnull=True
        )
        .values(
            "area__nombre"
        )
        .annotate(
            total=Count("id")
        )
        .order_by(
            "-total",
            "area__nombre",
        )[:limit]
    )

    return {
        "limite": limit,

        "items": [
            {
                "label": _label(
                    row["area__nombre"]
                ),
                "total": int(
                    row["total"]
                    or 0
                ),
            }
            for row in rows
        ],
    }


# ============================================================
# SEDES
# ============================================================


def _build_top_sedes(
    publicaciones,
    limit,
):
    """
    Construye el ranking de Sedes según la producción científica
    incluida en los filtros activos del dashboard.

    Las publicaciones históricas sin Sede permanecen disponibles
    en el resto de indicadores, pero no se atribuyen artificialmente
    a una ubicación institucional.
    """

    rows = (
        publicaciones
        .exclude(
            sede__isnull=True
        )
        .values(
            "sede_id",
            "sede__nombre",
        )
        .annotate(
            total=Count("id")
        )
        .order_by(
            "-total",
            "sede__nombre",
        )[:limit]
    )

    return {
        "limite": limit,

        "items": [
            {
                "sede_id": (
                    row["sede_id"]
                ),

                "sede": _label(
                    row["sede__nombre"]
                ),

                "label": _label(
                    row["sede__nombre"]
                ),

                "total": int(
                    row["total"]
                    or 0
                ),
            }

            for row in rows
        ],
    }


# ============================================================
# FACULTADES
# ============================================================


def _build_top_facultades(
    publicaciones,
    limit,
):
    rows = (
        publicaciones
        .exclude(
            carrera__facultad_id__isnull=True
        )
        .values(
            facultad_id=F(
                "carrera__facultad_id"
            ),
            facultad_nombre=F(
                "carrera__facultad__nombre"
            ),
        )
        .annotate(
            total=Count("id")
        )
        .order_by(
            "-total",
            "facultad_nombre",
        )[:limit]
    )

    return {
        "limite": limit,

        "items": [
            {
                "facultad_id": (
                    row["facultad_id"]
                ),

                "facultad": _label(
                    row["facultad_nombre"]
                ),

                "total": int(
                    row["total"]
                    or 0
                ),
            }

            for row in rows
        ],
    }


# ============================================================
# CARRERAS
# ============================================================


def _build_top_carreras(
    publicaciones,
    limit,
):
    rows = (
        publicaciones
        .exclude(
            carrera__isnull=True
        )
        .values(
            "carrera_id",
            "carrera__nombre",
        )
        .annotate(
            total=Count("id")
        )
        .order_by(
            "-total",
            "carrera__nombre",
        )[:limit]
    )

    return {
        "limite": limit,

        "items": [
            {
                "carrera_id": (
                    row["carrera_id"]
                ),

                "carrera": _label(
                    row["carrera__nombre"]
                ),

                "total": int(
                    row["total"]
                    or 0
                ),
            }

            for row in rows
        ],
    }


# ============================================================
# AUTORES
# ============================================================


def _build_top_autores(
    publicaciones,
    limit,
):
    """
    Ranking de autores por cantidad de publicaciones en las que
    participan. El orden bibliográfico no altera el conteo.
    """

    PublicacionAutor = (
        get_publicacion_autor_model()
    )

    base_qs = (
        PublicacionAutor.objects
        .filter(
            publicacion__in=(
                publicaciones
            )
        )
        .exclude(
            autor_id__isnull=True
        )
    )

    rows = (
        base_qs
        .values(
            "autor_id",
            "autor__nombres",
            "autor__apellidos",
        )
        .annotate(
            total_publicaciones=Count(
                "publicacion",
                distinct=True,
            )
        )
        .order_by(
            "-total_publicaciones",
            "autor__apellidos",
            "autor__nombres",
        )[:limit]
    )

    total_autores_activos = (
        base_qs
        .values(
            "autor_id"
        )
        .distinct()
        .count()
    )

    items = []

    for row in rows:
        nombre_autor = (
            f"{str(row['autor__nombres'] or '').strip()} "
            f"{str(row['autor__apellidos'] or '').strip()}"
        ).strip()

        if not nombre_autor:
            nombre_autor = (
                "Autor sin nombre"
            )

        total_publicaciones = int(
            row[
                "total_publicaciones"
            ]
            or 0
        )

        items.append(
            {
                "autor_id": (
                    row["autor_id"]
                ),

                "autor": (
                    nombre_autor
                ),

                "label": (
                    nombre_autor
                ),

                "total_publicaciones": (
                    total_publicaciones
                ),

                "total": (
                    total_publicaciones
                ),
            }
        )

    return {
        "limite": limit,

        "total_autores_activos": (
            total_autores_activos
        ),

        "items": items,
    }


# ============================================================
# REVISTAS
# ============================================================


def _build_journals(
    publicaciones,
    limit,
):
    Articulo = (
        get_articulo_model()
    )

    rows = (
        Articulo.objects
        .filter(
            publicacion__in=(
                publicaciones
            )
        )
        .exclude(
            nombre_revista__isnull=True
        )
        .exclude(
            nombre_revista=""
        )
        .values(
            "nombre_revista"
        )
        .annotate(
            total=Count("id")
        )
        .order_by(
            "-total",
            "nombre_revista",
        )[:limit]
    )

    return {
        "limite": limit,

        "items": [
            {
                "label": _label(
                    row["nombre_revista"]
                ),

                "total": int(
                    row["total"]
                    or 0
                ),
            }

            for row in rows
        ],
    }


# ============================================================
# PROYECTOS
# ============================================================


def _build_projects(
    publicaciones,
    limit,
):
    rows = (
        publicaciones
        .exclude(
            proyecto__isnull=True
        )
        .values(
            "proyecto__nombre"
        )
        .annotate(
            total=Count("id")
        )
        .order_by(
            "-total",
            "proyecto__nombre",
        )[:limit]
    )

    return {
        "limite": limit,

        "items": [
            {
                "label": _label(
                    row[
                        "proyecto__nombre"
                    ]
                ),

                "total": int(
                    row["total"]
                    or 0
                ),
            }

            for row in rows
        ],
    }


# ============================================================
# SUMMARY
# ============================================================


def _build_summary(
    publicaciones,
):
    PublicacionAutor = (
        get_publicacion_autor_model()
    )

    Articulo = (
        get_articulo_model()
    )

    total_publicaciones = (
        publicaciones.count()
    )

    total_autores = (
        PublicacionAutor.objects
        .filter(
            publicacion__in=(
                publicaciones
            )
        )
        .exclude(
            autor_id__isnull=True
        )
        .values(
            "autor_id"
        )
        .distinct()
        .count()
    )

    total_sedes = (
        publicaciones
        .exclude(
            sede_id__isnull=True
        )
        .values(
            "sede_id"
        )
        .distinct()
        .count()
    )

    total_facultades = (
        publicaciones
        .exclude(
            carrera__facultad_id__isnull=True
        )
        .values(
            "carrera__facultad_id"
        )
        .distinct()
        .count()
    )

    total_carreras = (
        publicaciones
        .exclude(
            carrera_id__isnull=True
        )
        .values(
            "carrera_id"
        )
        .distinct()
        .count()
    )

    total_proyectos = (
        publicaciones
        .exclude(
            proyecto_id__isnull=True
        )
        .values(
            "proyecto_id"
        )
        .distinct()
        .count()
    )

    articulos_qs = (
        Articulo.objects
        .filter(
            publicacion__in=(
                publicaciones
            )
        )
    )

    total_articulos_alto_impacto = (
        articulos_qs.filter(
            tipo_articulo=(
                "alto_impacto"
            )
        )
        .count()
    )

    total_articulos_regionales = (
        articulos_qs.filter(
            tipo_articulo="regional"
        )
        .count()
    )

    return {
        "total_publicaciones": (
            total_publicaciones
        ),

        "total_autores": (
            total_autores
        ),

        "total_sedes": (
            total_sedes
        ),

        "total_facultades": (
            total_facultades
        ),

        "total_carreras": (
            total_carreras
        ),

        "total_proyectos": (
            total_proyectos
        ),

        "articulos_alto_impacto": (
            total_articulos_alto_impacto
        ),

        "articulos_regionales": (
            total_articulos_regionales
        ),
    }


# ============================================================
# FILTROS DISPONIBLES
# ============================================================


def _compose_period_metadata(
    rows,
    *,
    today=None,
):
    """
    Construye el catálogo temporal consumido por el frontend.

    ``rows`` debe contener diccionarios con:

        anio_publicacion
        mes_publicacion
        total

    El catálogo se completa año a año desde el primer año con
    producción hasta, como mínimo, el año actual. Esto permite que
    el selector crezca automáticamente sin codificar años en Vue.
    """

    today = today or timezone.localdate()

    normalized_rows = []

    for row in rows or []:
        year = _safe_int(
            row.get(
                "anio_publicacion"
            )
        )

        month = _safe_int(
            row.get(
                "mes_publicacion"
            )
        )

        total = _safe_int(
            row.get(
                "total"
            )
        ) or 0

        if not year:
            continue

        normalized_rows.append(
            {
                "year": year,
                "month": month,
                "total": total,
            }
        )

    data_years = [
        row["year"]
        for row in normalized_rows
    ]

    anio_min = (
        min(data_years)
        if data_years
        else int(today.year)
    )

    anio_max = max(
        max(data_years)
        if data_years
        else int(today.year),
        int(today.year),
    )

    years = list(
        range(
            anio_max,
            anio_min - 1,
            -1,
        )
    )

    meses_con_datos = {}
    total_con_mes = 0
    total_sin_mes = 0
    latest_with_month = None

    for row in normalized_rows:
        year = row["year"]
        month = row["month"]
        total = row["total"]

        if not month or not (1 <= month <= 12):
            total_sin_mes += total
            continue

        total_con_mes += total

        year_key = str(year)
        month_key = str(month)

        year_bucket = (
            meses_con_datos
            .setdefault(
                year_key,
                {},
            )
        )

        year_bucket[month_key] = (
            int(
                year_bucket.get(
                    month_key,
                    0,
                )
            )
            + total
        )

        candidate = (year, month)

        if (
            latest_with_month is None
            or candidate > latest_with_month
        ):
            latest_with_month = candidate

    mes_actual = (
        f"{int(today.year):04d}-"
        f"{int(today.month):02d}"
    )

    return {
        "anio_min": anio_min,
        "anio_max": anio_max,
        "mes_min": (
            f"{anio_min:04d}-01"
        ),
        "mes_max": (
            f"{anio_max:04d}-12"
        ),
        "mes_actual": mes_actual,
        "ultimo_mes_con_datos": (
            _period_month_to_string(
                latest_with_month
            )
            if latest_with_month
            else None
        ),
        "anios": [
            {
                "value": year,
                "label": str(year),
            }
            for year in years
        ],
        "meses": [
            {
                "value": month,
                "label": MONTH_NAMES_ES[month],
                "short_label": MONTH_LABELS_ES[month],
            }
            for month in range(1, 13)
        ],
        "meses_con_datos": meses_con_datos,
        "total_con_mes": total_con_mes,
        "total_sin_mes": total_sin_mes,
    }


def _build_period_metadata(
    publicaciones,
):
    rows = list(
        publicaciones
        .exclude(
            anio_publicacion__isnull=True
        )
        .values(
            "anio_publicacion",
            "mes_publicacion",
        )
        .annotate(
            total=Count(
                "id"
            )
        )
        .order_by(
            "anio_publicacion",
            "mes_publicacion",
        )
    )

    return _compose_period_metadata(
        rows
    )


def _build_filters_metadata(
    publicaciones,
    publicaciones_para_tipos=None,
    publicaciones_temporales=None,
    selected_sede_id=None,
    selected_facultad_id=None,
    anio_base_mensual=None,
):
    Publicacion = (
        get_publicacion_model()
    )

    Sede = (
        get_sede_model()
    )

    Facultad = (
        get_facultad_model()
    )

    Carrera = (
        get_carrera_model()
    )

    # IMPORTANTE:
    #
    # No utilizar:
    #
    # publicaciones_para_tipos = publicaciones_para_tipos or publicaciones
    #
    # porque un QuerySet vacío se evaluaría y podría
    # sustituirse incorrectamente.
    if publicaciones_para_tipos is None:
        publicaciones_para_tipos = (
            publicaciones
        )

    if publicaciones_temporales is None:
        publicaciones_temporales = (
            publicaciones
        )

    periodo = (
        _build_period_metadata(
            publicaciones_temporales
        )
    )

    anios = (
        periodo.get(
            "anios",
            [],
        )
    )

    # --------------------------------------------------------
    # Sedes
    # --------------------------------------------------------

    sedes = list(
        Sede.objects
        .filter(
            activa=True
        )
        .order_by(
            "nombre"
        )
        .values(
            "id",
            "nombre",
            "codigo",
            "ciudad",
        )
    )

    # --------------------------------------------------------
    # Facultades
    #
    # Cuando existe una Sede seleccionada se ofrecen únicamente
    # facultades que poseen al menos una Carrera habilitada en
    # dicha Sede.
    # --------------------------------------------------------

    facultades_qs = (
        Facultad.objects
        .all()
        .order_by(
            "nombre"
        )
    )

    if selected_sede_id:
        facultades_qs = (
            facultades_qs
            .filter(
                carreras__sedes_carrera__sede_id=(
                    selected_sede_id
                ),
                carreras__sedes_carrera__activa=True,
                carreras__sedes_carrera__sede__activa=True,
            )
            .distinct()
        )

    facultades = list(
        facultades_qs.values(
            "id",
            "nombre",
        )
    )

    # --------------------------------------------------------
    # Carreras
    #
    # La relación Sede -> Carrera se valida mediante CarreraSede
    # activa. La Facultad continúa siendo una propiedad de la
    # Carrera y puede combinarse con el filtro de Sede.
    # --------------------------------------------------------

    carreras_qs = (
        Carrera.objects
        .select_related(
            "facultad"
        )
        .all()
        .order_by(
            "nombre"
        )
    )

    if selected_sede_id:
        carreras_qs = (
            carreras_qs
            .filter(
                sedes_carrera__sede_id=(
                    selected_sede_id
                ),
                sedes_carrera__activa=True,
                sedes_carrera__sede__activa=True,
            )
            .distinct()
        )

    if selected_facultad_id:
        carreras_qs = (
            carreras_qs.filter(
                facultad_id=(
                    selected_facultad_id
                )
            )
        )

    carreras = list(
        carreras_qs.values(
            "id",
            "nombre",
            "facultad_id",
        )
    )

    # --------------------------------------------------------
    # Tipos
    # --------------------------------------------------------

    canonical_counts = (
        _count_by_canonical_type(
            publicaciones_para_tipos
        )
    )

    return {
        "tipos": [
            {
                "id": (
                    item["codigo"]
                ),

                "codigo": (
                    item["codigo"]
                ),

                "nombre": (
                    item["nombre"]
                ),

                "categoria": (
                    item["categoria"]
                ),

                "total": int(
                    canonical_counts.get(
                        item["codigo"],
                        0,
                    )
                    or 0
                ),
            }

            for item in CANONICAL_TYPES
        ],

        "sedes": [
            {
                "id": (
                    row["id"]
                ),

                "nombre": (
                    row["nombre"]
                ),

                "codigo": (
                    row["codigo"]
                ),

                "ciudad": (
                    row["ciudad"]
                ),
            }

            for row in sedes
        ],

        "facultades": [
            {
                "id": (
                    row["id"]
                ),
                "nombre": (
                    row["nombre"]
                ),
            }

            for row in facultades
        ],

        "carreras": [
            {
                "id": (
                    row["id"]
                ),

                "nombre": (
                    row["nombre"]
                ),

                "facultad_id": (
                    row["facultad_id"]
                ),
            }

            for row in carreras
        ],

        "anios": anios,

        "periodo": periodo,

        "anio_base_mensual": (
            anio_base_mensual
        ),
    }


# ============================================================
# FILTROS APLICADOS
# ============================================================


def _build_filtros_aplicados(
    params,
    anio_base_mensual,
):
    Sede = (
        get_sede_model()
    )

    Facultad = (
        get_facultad_model()
    )

    Carrera = (
        get_carrera_model()
    )

    sede_id = _safe_int(
        params.get(
            "sede_id"
        )
        or params.get(
            "sede"
        )
    )

    facultad_id = _safe_int(
        params.get(
            "facultad_id"
        )
    )

    carrera_id = _safe_int(
        params.get(
            "carrera_id"
        )
    )

    tipo_codigo = (
        _normalize_canonical_code(
            params.get(
                "tipo_codigo"
            )
        )
    )

    period = _resolve_period_filters(
        params
    )

    anio = _safe_int(
        params.get(
            "anio"
        )
    )

    top = _safe_top(
        params.get(
            "top"
        )
    )

    sede_nombre = None
    facultad_nombre = None
    carrera_nombre = None

    if sede_id:
        sede_nombre = (
            Sede.objects
            .filter(
                id=sede_id
            )
            .values_list(
                "nombre",
                flat=True,
            )
            .first()
        )

    if facultad_id:
        facultad_nombre = (
            Facultad.objects
            .filter(
                id=facultad_id
            )
            .values_list(
                "nombre",
                flat=True,
            )
            .first()
        )

    if carrera_id:
        carrera_nombre = (
            Carrera.objects
            .filter(
                id=carrera_id
            )
            .values_list(
                "nombre",
                flat=True,
            )
            .first()
        )

    tipo_meta = (
        _canonical_meta(
            tipo_codigo
        )
        if tipo_codigo
        else None
    )

    mes_desde = period[
        "mes_desde"
    ]

    mes_hasta = period[
        "mes_hasta"
    ]

    return {
        "sede_id": (
            sede_id
        ),

        "sede_nombre": (
            sede_nombre
        ),

        "facultad_id": (
            facultad_id
        ),

        "facultad_nombre": (
            facultad_nombre
        ),

        "carrera_id": (
            carrera_id
        ),

        "carrera_nombre": (
            carrera_nombre
        ),

        "tipo_codigo": (
            tipo_codigo
        ),

        "tipo_nombre": (
            tipo_meta["nombre"]
            if tipo_meta
            else None
        ),

        "mes_desde": (
            _period_month_to_string(
                mes_desde
            )
        ),

        "mes_hasta": (
            _period_month_to_string(
                mes_hasta
            )
        ),

        "anio_desde": (
            period["anio_desde"]
        ),

        "anio_hasta": (
            period["anio_hasta"]
        ),

        "periodo_modo": (
            period["modo"]
        ),

        "anio": (
            anio
        ),

        "anio_base_mensual": (
            anio_base_mensual
        ),

        "top": (
            top
        ),
    }



# ============================================================
# PAYLOAD PRINCIPAL
# ============================================================


def _build_dashboard_payload(
    params,
):
    selected_sede_id = (
        _safe_int(
            params.get(
                "sede_id"
            )
            or params.get(
                "sede"
            )
        )
    )

    selected_facultad_id = (
        _safe_int(
            params.get(
                "facultad_id"
            )
        )
    )

    selected_tipo_codigo = (
        _normalize_canonical_code(
            params.get(
                "tipo_codigo"
            )
        )
    )

    selected_month_year = (
        _safe_int(
            params.get(
                "anio"
            )
        )
    )

    top_limit = (
        _safe_top(
            params.get(
                "top"
            )
        )
    )

    # --------------------------------------------------------
    # Queryset general con filtros institucionales/periodo
    # --------------------------------------------------------

    publicaciones_base = (
        _build_base_queryset(
            params
        )
    )

    # --------------------------------------------------------
    # Resolución de subtipos
    # --------------------------------------------------------

    base_id_sources = (
        _build_canonical_id_sources(
            publicaciones_base
        )
    )

    # --------------------------------------------------------
    # Aplicar tipo
    # --------------------------------------------------------

    publicaciones = (
        _apply_canonical_type_filter(
            publicaciones_base,
            tipo_codigo=(
                selected_tipo_codigo
            ),
            id_sources=(
                base_id_sources
            ),
        )
    )

    # --------------------------------------------------------
    # Catálogo temporal
    #
    # Se construye SIN aplicar mes_desde/mes_hasta ni rango de
    # años. De este modo el filtro temporal no se limita a sí
    # mismo. Sí respeta Sede, Facultad, Carrera, tipo y estado
    # aprobado, por lo que las opciones siguen siendo coherentes
    # con el contexto institucional seleccionado.
    # --------------------------------------------------------

    publicaciones_temporales_base = (
        _build_base_queryset(
            params,
            apply_period=False,
        )
    )

    temporal_id_sources = (
        _build_canonical_id_sources(
            publicaciones_temporales_base
        )
    )

    publicaciones_temporales = (
        _apply_canonical_type_filter(
            publicaciones_temporales_base,
            tipo_codigo=(
                selected_tipo_codigo
            ),
            id_sources=(
                temporal_id_sources
            ),
        )
    )

    # --------------------------------------------------------
    # Mensual
    # --------------------------------------------------------

    publicaciones_por_mes = (
        _build_publicaciones_por_mes(
            publicaciones,
            explicit_year=(
                selected_month_year
            ),
        )
    )

    # --------------------------------------------------------
    # Autores
    # --------------------------------------------------------

    top_autores = (
        _build_top_autores(
            publicaciones,
            top_limit,
        )
    )

    # --------------------------------------------------------
    # Payload
    # --------------------------------------------------------

    return {
        "ok": True,

        "summary": (
            _build_summary(
                publicaciones
            )
        ),

        "dashboards": {
            "publicaciones_por_anio": (
                _build_publicaciones_por_anio(
                    publicaciones
                )
            ),

            "publicaciones_por_mes": (
                publicaciones_por_mes
            ),

            "publicaciones_por_tipo": (
                _build_publicaciones_por_tipo(
                    publicaciones,
                    selected_tipo_codigo=(
                        selected_tipo_codigo
                    ),
                    id_sources=(
                        base_id_sources
                    ),
                )
            ),

            "publicaciones_por_tipo_anual": (
                _build_publicaciones_por_tipo_anual(
                    publicaciones,
                    id_sources=(
                        base_id_sources
                    ),
                )
            ),

            "areas": (
                _build_top_areas(
                    publicaciones,
                    top_limit,
                )
            ),

            "top_sedes": (
                _build_top_sedes(
                    publicaciones,
                    top_limit,
                )
            ),

            "top_facultades": (
                _build_top_facultades(
                    publicaciones,
                    top_limit,
                )
            ),

            "top_carreras": (
                _build_top_carreras(
                    publicaciones,
                    top_limit,
                )
            ),

            "top_autores": (
                top_autores
            ),

            "journals": (
                _build_journals(
                    publicaciones,
                    top_limit,
                )
            ),

            "projects": (
                _build_projects(
                    publicaciones,
                    top_limit,
                )
            ),
        },

        "filtros_disponibles": (
            _build_filters_metadata(
                publicaciones,
                publicaciones_para_tipos=(
                    publicaciones_base
                ),
                publicaciones_temporales=(
                    publicaciones_temporales
                ),
                selected_sede_id=(
                    selected_sede_id
                ),
                selected_facultad_id=(
                    selected_facultad_id
                ),
                anio_base_mensual=(
                    publicaciones_por_mes.get(
                        "anio_base"
                    )
                ),
            )
        ),

        "filtros_aplicados": (
            _build_filtros_aplicados(
                params,
                anio_base_mensual=(
                    publicaciones_por_mes.get(
                        "anio_base"
                    )
                ),
            )
        ),
    }


# ============================================================
# API — DASHBOARD
# ============================================================


class DashboardResumenView(
    APIView
):
    authentication_classes = [
        JWTAuthentication
    ]

    permission_classes = [
        IsAuthenticated
    ]

    def get(
        self,
        request,
        *args,
        **kwargs,
    ):
        payload = (
            _build_dashboard_payload(
                request.query_params
            )
        )

        return Response(
            payload
        )


# ============================================================
# REPORTE EXCEL — HELPERS
# ============================================================

REPORT_COLORS = {
    "navy": "0F172A",
    "primary": "2563EB",
    "primary_dark": "1D4ED8",
    "primary_soft": "EFF6FF",
    "green": "15803D",
    "purple": "7C3AED",
    "orange": "C2410C",
    "teal": "0F766E",
    "text": "172033",
    "muted": "64748B",
    "surface": "FFFFFF",
    "background": "F8FAFC",
    "border": "D8E1EC",
    "border_strong": "CBD5E1",
}

REPORT_TYPE_COLORS = {
    "AAI": "3366CC",
    "AR": "3F7F5A",
    "PON": "7251B5",
    "CAP": "B85C2E",
    "LIB": "1F7A73",
}

REPORT_KPI_COLORS = (
    "2563EB",
    "0F766E",
    "7C3AED",
    "C2410C",
    "0F766E",
    "2563EB",
    "7C3AED",
    "C2410C",
)


def _safe_excel_value(value):
    return "" if value is None else value


def _report_generated_at():
    return timezone.localtime().strftime("%d/%m/%Y %H:%M")


def _excel_report_filename():
    stamp = timezone.localtime().strftime("%Y-%m-%d_%H-%M-%S")
    return f"reporte-dashboard-sgpc-uleam-{stamp}.xlsx"


def _solid_fill(color):
    return PatternFill("solid", fgColor=color)


def _thin_border(color=None):
    side = Side(style="thin", color=color or REPORT_COLORS["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def _safe_table_name(value):
    normalized = (
        unicodedata.normalize("NFKD", str(value or "Datos"))
        .encode("ascii", "ignore")
        .decode("ascii")
    )
    token = "".join(ch for ch in normalized if ch.isalnum()) or "Datos"
    if token[0].isdigit():
        token = f"T{token}"
    return f"Tbl{token}"[:240]


def _set_range_fill(ws, cell_range, color):
    fill = _solid_fill(color)
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = fill


def _set_range_border(ws, cell_range, color=None):
    border = _thin_border(color)
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border


def _merge_block(
    ws,
    cell_range,
    value,
    *,
    fill=None,
    font=None,
    alignment=None,
    border=None,
):
    if fill:
        _set_range_fill(ws, cell_range, fill)
    if border:
        _set_range_border(ws, cell_range, border)

    ws.merge_cells(cell_range)
    anchor = ws[cell_range.split(":")[0]]
    anchor.value = value

    if font:
        anchor.font = font
    if alignment:
        anchor.alignment = alignment

    return anchor


def _format_month_period_label(
    value,
):
    parsed = _safe_period_month(
        value,
        field_name="periodo",
    )

    if not parsed:
        return ""

    year, month = parsed

    return (
        f"{MONTH_NAMES_ES[month]} "
        f"de {year}"
    )


def _format_periodo(
    filtros,
):
    mes_desde_raw = (
        filtros.get(
            "mes_desde"
        )
        or filtros.get(
            "fecha_desde"
        )
    )

    mes_hasta_raw = (
        filtros.get(
            "mes_hasta"
        )
        or filtros.get(
            "fecha_hasta"
        )
    )

    if (
        mes_desde_raw
        or mes_hasta_raw
    ):
        desde = (
            _format_month_period_label(
                mes_desde_raw
            )
            if mes_desde_raw
            else ""
        )

        hasta = (
            _format_month_period_label(
                mes_hasta_raw
            )
            if mes_hasta_raw
            else ""
        )

        if (
            desde
            and hasta
        ):
            return (
                desde
                if desde == hasta
                else (
                    f"{desde} – "
                    f"{hasta}"
                )
            )

        if desde:
            return (
                f"Desde {desde}"
            )

        if hasta:
            return (
                f"Hasta {hasta}"
            )

    desde = filtros.get(
        "anio_desde"
    )

    hasta = filtros.get(
        "anio_hasta"
    )

    if (
        desde
        and hasta
    ):
        return (
            str(desde)
            if desde == hasta
            else (
                f"{desde} – "
                f"{hasta}"
            )
        )

    if desde:
        return (
            f"Desde {desde}"
        )

    if hasta:
        return (
            f"Hasta {hasta}"
        )

    return "Todos los periodos"



def _set_sheet_print_settings(ws, *, print_area=None):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.45
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2

    ws.oddHeader.center.text = "SGPC ULEAM"
    ws.oddHeader.center.size = 9
    ws.oddHeader.center.font = "Calibri,Bold"
    ws.oddFooter.left.text = "Producción científica"
    ws.oddFooter.right.text = "Página &[Page] de &N"

    if print_area:
        ws.print_area = print_area


# ============================================================
# ESTILO DE HOJAS DE DATOS
# ============================================================


def _style_report_sheet(ws, headers, *, table_name, tab_color=None):
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    last_col = get_column_letter(max_col)

    title_fill = _solid_fill(REPORT_COLORS["navy"])
    subtitle_fill = _solid_fill(REPORT_COLORS["primary"])
    meta_fill = _solid_fill(REPORT_COLORS["primary_soft"])
    header_fill = _solid_fill(REPORT_COLORS["primary_dark"])
    body_border = _thin_border()

    if max_col > 1:
        for row_number in (1, 2, 3):
            ws.merge_cells(
                start_row=row_number,
                start_column=1,
                end_row=row_number,
                end_column=max_col,
            )

    for cell in ws[1]:
        cell.fill = title_fill
    for cell in ws[2]:
        cell.fill = subtitle_fill
    for cell in ws[3]:
        cell.fill = meta_fill

    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A2"].font = Font(name="Calibri", size=10.5, bold=True, color="FFFFFF")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A3"].font = Font(
        name="Calibri", size=9, italic=True, color=REPORT_COLORS["muted"]
    )
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 19
    ws.row_dimensions[4].height = 8
    ws.row_dimensions[5].height = 25

    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Calibri", size=10, color=REPORT_COLORS["text"])

    for cell in ws[5]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = body_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=6, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.font = body_font
            cell.border = body_border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    if max_row >= 6:
        table = ExcelTable(displayName=table_name, ref=f"A5:{last_col}{max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    else:
        ws.auto_filter.ref = f"A5:{last_col}5"

    numeric_headers = {"total", "total publicaciones", "total participaciones", "valor"}

    for column_index, header in enumerate(headers, start=1):
        normalized = str(header or "").strip().lower()
        letter = get_column_letter(column_index)

        if max_row < 6:
            continue

        cell_range = f"{letter}6:{letter}{max_row}"
        cells = [row[0] for row in ws[cell_range]]

        if normalized in numeric_headers:
            for cell in cells:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")

            ws.conditional_formatting.add(
                cell_range,
                DataBarRule(
                    start_type="num",
                    start_value=0,
                    end_type="max",
                    color=REPORT_COLORS["primary"],
                    showValue=True,
                ),
            )

        elif normalized == "porcentaje":
            for cell in cells:
                cell.number_format = '0.0"%"'
                cell.alignment = Alignment(horizontal="right", vertical="center")

        elif normalized in {"año", "código", "categoría"}:
            for cell in cells:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A6"

    for column_index in range(1, max_col + 1):
        header_value = headers[column_index - 1] if column_index - 1 < len(headers) else ""
        max_length = len(str(header_value or ""))

        for row_index in range(6, max_row + 1):
            value = ws.cell(row=row_index, column=column_index).value
            max_length = max(max_length, len(str(value or "")))

        ws.column_dimensions[get_column_letter(column_index)].width = min(
            max(max_length + 4, 12),
            42,
        )

    for row_index in range(6, max_row + 1):
        ws.row_dimensions[row_index].height = 20

    if tab_color:
        ws.sheet_properties.tabColor = tab_color

    ws.print_title_rows = "1:5"
    _set_sheet_print_settings(ws, print_area=f"A1:{last_col}{max_row}")


# ============================================================
# CREACIÓN DE HOJAS
# ============================================================


def _create_sheet(
    workbook,
    title,
    subtitle,
    headers,
    rows,
    *,
    tab_color=None,
    filtros=None,
):
    """Crea una hoja de datos con la misma identidad visual del PDF."""
    filtros = filtros or {}
    ws = workbook.create_sheet(title=title[:31])
    ws.sheet_view.showGridLines = False

    max_col = max(len(headers), 12)
    last_col = get_column_letter(max_col)

    # Anchos base de informe. Las columnas de datos se ajustan después.
    for column in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(column)].width = 12

    # Cabecera equivalente al PDF: título limpio + contexto, sin grandes bandas oscuras.
    _merge_block(
        ws,
        f"A1:{last_col}1",
        "Informe de producción científica",
        fill=REPORT_COLORS["surface"],
        font=Font(name="Calibri", size=18, bold=True, color=REPORT_COLORS["text"]),
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    _merge_block(
        ws,
        f"A2:{last_col}2",
        f"Anexo de datos · {title}. {subtitle}",
        fill=REPORT_COLORS["surface"],
        font=Font(name="Calibri", size=9.5, color=REPORT_COLORS["muted"]),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
    )
    _set_range_border(ws, f"A2:{last_col}2", REPORT_COLORS["border"])

    # Contexto del reporte, organizado como el bloque superior del PDF.
    context = (
        ("Sede", filtros.get("sede_nombre") or "Todas las sedes", "A4:D4", "A5:D5"),
        ("Facultad", filtros.get("facultad_nombre") or "Todas las facultades", "E4:H4", "E5:H5"),
        ("Carrera", filtros.get("carrera_nombre") or "Todas las carreras", "I4:L4", "I5:L5"),
        ("Tipo", filtros.get("tipo_nombre") or "Todos los tipos", "A6:D6", "A7:D7"),
        ("Período", _format_periodo(filtros), "E6:H6", "E7:H7"),
        ("Generado", _report_generated_at(), "I6:L6", "I7:L7"),
    )
    for label, value, label_range, value_range in context:
        _write_filter_card(
            ws,
            label,
            value,
            label_range=label_range,
            value_range=value_range,
        )

    _write_section_band(
        ws,
        9,
        title,
        subtitle,
        end_column=max_col,
    )

    header_row = 11
    data_start = 12
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col_index, header)
        cell.fill = _solid_fill(REPORT_COLORS["primary_soft"])
        cell.font = Font(name="Calibri", size=9, bold=True, color=REPORT_COLORS["primary_dark"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()

    for row_offset, values in enumerate(rows, start=0):
        row_index = data_start + row_offset
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row_index, col_index, _safe_excel_value(value))
            cell.fill = _solid_fill(
                REPORT_COLORS["surface"] if row_offset % 2 == 0 else REPORT_COLORS["background"]
            )
            cell.font = Font(name="Calibri", size=9, color=REPORT_COLORS["text"])
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = _thin_border()

    last_row = max(header_row, data_start + len(rows) - 1)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{last_row}"

    numeric_headers = {"total", "total publicaciones", "total participaciones", "valor", "publicaciones", "artículos"}
    for column_index, header in enumerate(headers, start=1):
        normalized = str(header or "").strip().lower()
        letter = get_column_letter(column_index)
        if last_row < data_start:
            continue
        cells = [row[0] for row in ws[f"{letter}{data_start}:{letter}{last_row}"]]
        if normalized in numeric_headers:
            for cell in cells:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="top")
        elif normalized == "porcentaje":
            for cell in cells:
                cell.number_format = '0.0"%"'
                cell.alignment = Alignment(horizontal="right", vertical="top")
        elif normalized == "año":
            for cell in cells:
                cell.alignment = Alignment(horizontal="center", vertical="top")

    # Ajuste de ancho con límites para conservar proporciones tipo informe.
    for column_index, header in enumerate(headers, start=1):
        max_length = len(str(header or ""))
        for row_index in range(data_start, last_row + 1):
            max_length = max(max_length, len(str(ws.cell(row_index, column_index).value or "")))
        ws.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 3, 12), 38)

    for row_index in range(data_start, last_row + 1):
        ws.row_dimensions[row_index].height = 22

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 24
    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 24
    ws.row_dimensions[9].height = 22
    ws.row_dimensions[10].height = 18
    ws.row_dimensions[11].height = 24

    if tab_color:
        ws.sheet_properties.tabColor = tab_color

    # Sin paneles congelados: evita la línea que atravesaba el diseño.
    ws.freeze_panes = None
    ws.print_title_rows = "1:11"
    _set_sheet_print_settings(ws, print_area=f"A1:{last_col}{last_row}")
    return ws

def _build_tipo_anual_rows(publicaciones_por_tipo_anual):
    """Filas visibles para usuario final: año, tipo y cantidad."""
    rows = []
    categorias = publicaciones_por_tipo_anual.get("categorias", []) or []

    for serie in publicaciones_por_tipo_anual.get("series", []) or []:
        data = serie.get("data", []) or []
        for index, anio in enumerate(categorias):
            rows.append(
                [
                    anio,
                    serie.get("label"),
                    data[index] if index < len(data) else 0,
                ]
            )

    return rows


# ============================================================
# DATOS OCULTOS PARA GRÁFICOS
# ============================================================


def _build_chart_data_sheet(workbook, dashboards):
    """Datos técnicos ocultos usados únicamente por los gráficos de Excel."""
    ws = workbook.create_sheet("_DatosGraficos")

    publicaciones_por_tipo = dashboards.get("publicaciones_por_tipo", {}) or {}
    publicaciones_por_anio = dashboards.get("publicaciones_por_anio", []) or []
    publicaciones_por_mes = dashboards.get("publicaciones_por_mes", {}) or {}
    publicaciones_por_tipo_anual = dashboards.get("publicaciones_por_tipo_anual", {}) or {}

    top_sedes = dashboards.get("top_sedes", {}) or {}
    top_facultades = dashboards.get("top_facultades", {}) or {}
    top_carreras = dashboards.get("top_carreras", {}) or {}
    areas = dashboards.get("areas", {}) or {}
    top_autores = dashboards.get("top_autores", {}) or {}
    journals = dashboards.get("journals", {}) or {}
    projects = dashboards.get("projects", {}) or {}

    # A:C  -> tipos
    ws.cell(1, 1, "Tipo")
    ws.cell(1, 2, "Total")
    ws.cell(1, 3, "Código")
    for row_index, item in enumerate(publicaciones_por_tipo.get("items", []) or [], start=2):
        ws.cell(row_index, 1, item.get("tipo_nombre"))
        ws.cell(row_index, 2, item.get("total", 0))
        ws.cell(row_index, 3, item.get("tipo_codigo"))

    # E:F -> producción por año
    ws.cell(1, 5, "Año")
    ws.cell(1, 6, "Total")
    for row_index, item in enumerate(publicaciones_por_anio, start=2):
        ws.cell(row_index, 5, item.get("label"))
        ws.cell(row_index, 6, item.get("value", 0))

    # H:M -> tipos por año (matriz)
    categories = publicaciones_por_tipo_anual.get("categorias", []) or []
    series = publicaciones_por_tipo_anual.get("series", []) or []
    ws.cell(1, 8, "Año")
    for series_index, serie in enumerate(series, start=9):
        ws.cell(1, series_index, serie.get("label"))
    for row_index, category in enumerate(categories, start=2):
        ws.cell(row_index, 8, category)
        for series_offset, serie in enumerate(series, start=9):
            values = serie.get("data", []) or []
            value_index = row_index - 2
            ws.cell(
                row_index,
                series_offset,
                values[value_index] if value_index < len(values) else 0,
            )

    # O:P -> producción por mes
    ws.cell(1, 15, "Mes")
    ws.cell(1, 16, "Total")
    for row_index, item in enumerate(publicaciones_por_mes.get("items", []) or [], start=2):
        ws.cell(row_index, 15, item.get("label"))
        ws.cell(row_index, 16, item.get("value", 0))

    # Rankings / listas. Cada conjunto usa dos columnas.
    blocks = (
        (18, "Sede", top_sedes.get("items", []) or [], ("sede", "label"), "total"),
        (21, "Facultad", top_facultades.get("items", []) or [], ("facultad", "label"), "total"),
        (24, "Carrera", top_carreras.get("items", []) or [], ("carrera", "label"), "total"),
        (27, "Área", areas.get("items", []) or [], ("label",), "total"),
        (30, "Autor", top_autores.get("items", []) or [], ("autor", "label"), "total"),
        (33, "Revista", journals.get("items", []) or [], ("label",), "total"),
        (36, "Proyecto", projects.get("items", []) or [], ("label",), "total"),
    )

    for start_col, label_header, items, label_keys, value_key in blocks:
        ws.cell(1, start_col, label_header)
        ws.cell(1, start_col + 1, "Total")
        for row_index, item in enumerate(items, start=2):
            label = ""
            for key in label_keys:
                label = item.get(key)
                if label:
                    break
            ws.cell(row_index, start_col, label)
            ws.cell(
                row_index,
                start_col + 1,
                item.get("total_publicaciones", item.get(value_key, 0)),
            )

    ws.sheet_state = "hidden"
    return ws


# ============================================================
# HOJAS VISUALES — MISMA LÓGICA DEL DASHBOARD
# ============================================================


def _write_filter_card(ws, label, value, *, label_range, value_range):
    _merge_block(
        ws,
        label_range,
        label,
        fill=REPORT_COLORS["primary_soft"],
        font=Font(name="Calibri", size=8.5, bold=True, color=REPORT_COLORS["primary_dark"]),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=REPORT_COLORS["border"],
    )
    _merge_block(
        ws,
        value_range,
        _safe_excel_value(value),
        fill=REPORT_COLORS["surface"],
        font=Font(name="Calibri", size=9.5, bold=True, color=REPORT_COLORS["text"]),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
        border=REPORT_COLORS["border"],
    )


def _write_kpi_card(ws, label, value, *, label_range, value_range, accent):
    _merge_block(
        ws,
        label_range,
        label,
        fill=REPORT_COLORS["surface"],
        font=Font(name="Calibri", size=8.5, bold=True, color=REPORT_COLORS["muted"]),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=REPORT_COLORS["border"],
    )
    value_cell = _merge_block(
        ws,
        value_range,
        int(value or 0),
        fill=REPORT_COLORS["surface"],
        font=Font(name="Calibri", size=18, bold=True, color=REPORT_COLORS["text"]),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=REPORT_COLORS["border"],
    )
    value_cell.number_format = "#,##0"


def _prepare_visual_sheet(workbook, name, view_label, filtros):
    """Hoja visual que reproduce la composición del PDF institucional."""
    ws = workbook.create_sheet(name)
    ws.sheet_properties.tabColor = REPORT_COLORS["primary"]
    ws.sheet_view.showGridLines = False

    for column in range(1, 13):
        ws.column_dimensions[get_column_letter(column)].width = 12.5

    _merge_block(
        ws,
        "A1:L1",
        "Informe de producción científica",
        fill=REPORT_COLORS["surface"],
        font=Font(name="Calibri", size=18, bold=True, color=REPORT_COLORS["text"]),
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    _merge_block(
        ws,
        "A2:L2",
        "Publicaciones aprobadas según las opciones seleccionadas en el panel.",
        fill=REPORT_COLORS["surface"],
        font=Font(name="Calibri", size=9.5, color=REPORT_COLORS["muted"]),
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    _set_range_border(ws, "A2:L2", REPORT_COLORS["border"])

    # Mismo bloque de contexto del PDF: 3 columnas x 2 filas.
    filter_specs = (
        ("Sede", filtros.get("sede_nombre") or "Todas las sedes", "A4:D4", "A5:D5"),
        ("Facultad", filtros.get("facultad_nombre") or "Todas las facultades", "E4:H4", "E5:H5"),
        ("Carrera", filtros.get("carrera_nombre") or "Todas las carreras", "I4:L4", "I5:L5"),
        ("Tipo", filtros.get("tipo_nombre") or "Todos los tipos", "A6:D6", "A7:D7"),
        ("Período", _format_periodo(filtros), "E6:H6", "E7:H7"),
        ("Generado", _report_generated_at(), "I6:L6", "I7:L7"),
    )
    for label, value, label_range, value_range in filter_specs:
        _write_filter_card(
            ws,
            label,
            value,
            label_range=label_range,
            value_range=value_range,
        )

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 24
    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 24
    ws.freeze_panes = None
    return ws

def _write_section_band(ws, row, title, subtitle=None, *, end_column=12):
    """Encabezado de sección equivalente al usado en el PDF."""
    last_col = get_column_letter(end_column)
    _merge_block(
        ws,
        f"A{row}:{last_col}{row}",
        title,
        fill=REPORT_COLORS["surface"],
        font=Font(name="Calibri", size=12, bold=True, color=REPORT_COLORS["text"]),
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    _set_range_border(ws, f"A{row}:{last_col}{row}", REPORT_COLORS["border"])
    ws.row_dimensions[row].height = 23

    if subtitle is not None:
        _merge_block(
            ws,
            f"A{row + 1}:{last_col}{row + 1}",
            subtitle,
            fill=REPORT_COLORS["surface"],
            font=Font(name="Calibri", size=8.5, color=REPORT_COLORS["muted"]),
            alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
        )
        ws.row_dimensions[row + 1].height = 18

def _write_compact_list(ws, start_row, start_col, end_col, title, items, *, label_keys, value_label="Publicaciones", limit=None):
    """Lista simple para autores, revistas y proyectos; evita barras innecesarias."""
    if limit is not None:
        items = list(items or [])[:limit]
    else:
        items = list(items or [])

    title_start = get_column_letter(start_col)
    title_end = get_column_letter(end_col)
    ws.merge_cells(f"{title_start}{start_row}:{title_end}{start_row}")
    cell = ws.cell(start_row, start_col, title)
    cell.font = Font(name="Calibri", size=10.5, bold=True, color=REPORT_COLORS["text"])
    cell.fill = _solid_fill(REPORT_COLORS["surface"])
    cell.alignment = Alignment(horizontal="left", vertical="center")
    _set_range_border(ws, f"{title_start}{start_row}:{title_end}{start_row}", REPORT_COLORS["border"])

    header_row = start_row + 1
    headers = ("Pos.", title, value_label)
    ws.cell(header_row, start_col, headers[0])
    ws.merge_cells(
        start_row=header_row,
        start_column=start_col + 1,
        end_row=header_row,
        end_column=end_col - 1,
    )
    ws.cell(header_row, start_col + 1, headers[1])
    ws.cell(header_row, end_col, headers[2])

    for col in range(start_col, end_col + 1):
        hcell = ws.cell(header_row, col)
        hcell.fill = _solid_fill(REPORT_COLORS["primary_soft"])
        hcell.font = Font(name="Calibri", size=8.5, bold=True, color=REPORT_COLORS["primary_dark"])
        hcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hcell.border = _thin_border()

    if not items:
        data_row = header_row + 1
        ws.merge_cells(
            start_row=data_row,
            start_column=start_col,
            end_row=data_row,
            end_column=end_col,
        )
        ws.cell(data_row, start_col, "No hay información para esta selección.")
        ws.cell(data_row, start_col).font = Font(name="Calibri", size=9, color=REPORT_COLORS["muted"])
        ws.cell(data_row, start_col).alignment = Alignment(horizontal="left", vertical="center")
        _set_range_border(ws, f"{title_start}{data_row}:{title_end}{data_row}", REPORT_COLORS["border"])
        return data_row

    for index, item in enumerate(items, start=1):
        row = header_row + index
        label = ""
        for key in label_keys:
            label = item.get(key)
            if label:
                break
        total = item.get("total_publicaciones", item.get("total", 0))

        ws.cell(row, start_col, index)
        ws.merge_cells(
            start_row=row,
            start_column=start_col + 1,
            end_row=row,
            end_column=end_col - 1,
        )
        ws.cell(row, start_col + 1, label or "Sin información")
        ws.cell(row, end_col, int(total or 0))

        for col in range(start_col, end_col + 1):
            dcell = ws.cell(row, col)
            dcell.border = _thin_border()
            dcell.fill = _solid_fill(REPORT_COLORS["surface"])
            dcell.font = Font(name="Calibri", size=8.8, color=REPORT_COLORS["text"])
            dcell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row, start_col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, end_col).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row, end_col).number_format = "#,##0"
        ws.row_dimensions[row].height = 20

    return header_row + len(items)


def _build_summary_sheet(workbook, summary, filtros, dashboards):
    """
    Resumen visual del Excel.

    A diferencia del PDF, el Excel no fuerza una composición rígida de
    dos columnas. La dona utiliza el ancho completo disponible para que
    sus etiquetas y leyenda mantengan legibilidad, y los rankings se
    colocan debajo según el contenido real.
    """
    ws = _prepare_visual_sheet(workbook, "Resumen", "Resumen", filtros)

    _write_section_band(
        ws,
        9,
        "Resumen",
        "Vista general de la producción científica incluida en este informe.",
    )

    kpis = (
        ("Publicaciones", summary.get("total_publicaciones", 0)),
        ("Autores", summary.get("total_autores", 0)),
        ("Sedes", summary.get("total_sedes", 0)),
        ("Facultades", summary.get("total_facultades", 0)),
        ("Carreras", summary.get("total_carreras", 0)),
        ("Proyectos", summary.get("total_proyectos", 0)),
    )
    card_ranges = (
        ("A12:B12", "A13:B15"),
        ("C12:D12", "C13:D15"),
        ("E12:F12", "E13:F15"),
        ("G12:H12", "G13:H15"),
        ("I12:J12", "I13:J15"),
        ("K12:L12", "K13:L15"),
    )
    for index, (label, value) in enumerate(kpis):
        _write_kpi_card(
            ws,
            label,
            value,
            label_range=card_ranges[index][0],
            value_range=card_ranges[index][1],
            accent=REPORT_KPI_COLORS[index],
        )

    type_items = (
        dashboards.get("publicaciones_por_tipo", {}) or {}
    ).get("items", []) or []

    current_row = 17

    # La composición por tipo necesita espacio horizontal para la leyenda.
    if type_items:
        donut_rows = max(17, min(22, 15 + len(type_items)))
        donut_end = current_row + donut_rows - 1
        _set_range_fill(
            ws,
            f"A{current_row}:L{donut_end}",
            REPORT_COLORS["surface"],
        )
        _set_range_border(
            ws,
            f"A{current_row}:L{donut_end}",
            REPORT_COLORS["border"],
        )
        current_row = donut_end + 2

    ranking_specs = (
        (
            "Sedes",
            (dashboards.get("top_sedes", {}) or {}).get("items", []) or [],
            ("sede", "label"),
        ),
        (
            "Facultades",
            (dashboards.get("top_facultades", {}) or {}).get("items", []) or [],
            ("facultad", "label"),
        ),
        (
            "Autores",
            (dashboards.get("top_autores", {}) or {}).get("items", []) or [],
            ("autor", "label"),
        ),
    )

    for title, items, label_keys in ranking_specs:
        end_row = _write_compact_list(
            ws,
            current_row,
            1,
            12,
            title,
            items,
            label_keys=label_keys,
            value_label="Publicaciones",
            limit=3,
        )
        current_row = end_row + 2

    print_end = max(39, current_row - 1)
    _set_sheet_print_settings(
        ws,
        print_area=f"A1:L{print_end}",
    )
    return ws

def _build_evolution_sheet(workbook, filtros, dashboards):
    """
    Hoja temporal adaptativa.

    Las series temporales dejan de compartir obligatoriamente media hoja.
    Cada gráfico recibe el ancho completo porque el número de años/meses
    puede variar y las categorías deben conservar separación legible.
    """
    ws = _prepare_visual_sheet(workbook, "Evolución", "Evolución", filtros)
    _write_section_band(
        ws,
        9,
        "Evolución",
        "Comportamiento de la producción científica a lo largo del tiempo.",
    )

    layout = _evolution_chart_layout(dashboards)

    if not layout:
        _merge_block(
            ws,
            "A12:L16",
            "No hay información temporal suficiente para generar gráficos.",
            fill=REPORT_COLORS["surface"],
            font=Font(
                name="Calibri",
                size=9,
                color=REPORT_COLORS["muted"],
            ),
            alignment=Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            ),
            border=REPORT_COLORS["border"],
        )
        print_end = 16
    else:
        for item in layout:
            _set_range_fill(
                ws,
                f"A{item['start_row']}:L{item['end_row']}",
                REPORT_COLORS["surface"],
            )
            _set_range_border(
                ws,
                f"A{item['start_row']}:L{item['end_row']}",
                REPORT_COLORS["border"],
            )
        print_end = layout[-1]["end_row"]

    _set_sheet_print_settings(
        ws,
        print_area=f"A1:L{print_end}",
    )
    return ws

def _build_highlights_sheet(workbook, filtros, dashboards):
    """
    Comparativas institucionales con altura proporcional al número de filas.

    Los rankings dejan de estar encerrados en una cuadrícula 2x2 fija.
    Se apilan a ancho completo para evitar que nombres de facultades,
    carreras o áreas queden comprimidos.
    """
    ws = _prepare_visual_sheet(workbook, "Destacados", "Destacados", filtros)
    _write_section_band(
        ws,
        9,
        "Destacados",
        "Unidades académicas y áreas con mayor producción en la selección actual.",
    )

    layout = _highlight_chart_layout(dashboards)

    if not layout:
        _merge_block(
            ws,
            "A12:L16",
            "No hay información suficiente para generar comparativas.",
            fill=REPORT_COLORS["surface"],
            font=Font(
                name="Calibri",
                size=9,
                color=REPORT_COLORS["muted"],
            ),
            alignment=Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            ),
            border=REPORT_COLORS["border"],
        )
        print_end = 16
    else:
        for item in layout:
            _set_range_fill(
                ws,
                f"A{item['start_row']}:L{item['end_row']}",
                REPORT_COLORS["surface"],
            )
            _set_range_border(
                ws,
                f"A{item['start_row']}:L{item['end_row']}",
                REPORT_COLORS["border"],
            )
        print_end = layout[-1]["end_row"]

    _set_sheet_print_settings(
        ws,
        print_area=f"A1:L{print_end}",
    )
    return ws

def _build_people_projects_sheet(workbook, filtros, dashboards):
    """Página equivalente al bloque final del PDF: autores, revistas y proyectos."""
    ws = _prepare_visual_sheet(
        workbook,
        "Autores y proyectos",
        "Autores, revistas y proyectos",
        filtros,
    )
    _write_section_band(
        ws,
        9,
        "Autores, revistas y proyectos",
        "Principales participantes y espacios de publicación vinculados con la producción mostrada.",
    )

    top_autores = (dashboards.get("top_autores", {}) or {}).get("items", []) or []
    journals = (dashboards.get("journals", {}) or {}).get("items", []) or []
    projects = (dashboards.get("projects", {}) or {}).get("items", []) or []

    end_author = _write_compact_list(
        ws, 12, 1, 6, "Autores", top_autores,
        label_keys=("autor", "label"), value_label="Publicaciones",
    )
    end_journal = _write_compact_list(
        ws, 12, 7, 12, "Revistas", journals,
        label_keys=("label",), value_label="Artículos",
    )

    project_start = max(end_author, end_journal) + 3
    end_project = _write_compact_list(
        ws, project_start, 1, 12, "Proyectos", projects,
        label_keys=("label",), value_label="Publicaciones",
    )

    note_row = end_project + 2
    _merge_block(
        ws,
        f"A{note_row}:L{note_row}",
        "El Excel conserva el detalle para ordenar, filtrar y trabajar con los datos; las primeras hojas replican la lectura visual del PDF.",
        fill=REPORT_COLORS["background"],
        font=Font(name="Calibri", size=8.5, italic=True, color=REPORT_COLORS["muted"]),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
        border=REPORT_COLORS["border"],
    )
    _set_sheet_print_settings(ws, print_area=f"A1:L{note_row}")
    return ws


# ============================================================
# GRÁFICOS EXCEL
# ============================================================


def _safe_chart_int(value):
    try:
        return max(0, int(value or 0))
    except (TypeError, ValueError, OverflowError):
        return 0


def _integer_chart_axis(axis, maximum, *, headroom=True):
    """
    Configura un eje de publicaciones como entero.

    Se añade margen superior para impedir que el valor máximo quede pegado
    al borde del gráfico, especialmente cuando existen muy pocos registros.
    """
    maximum = max(1, _safe_chart_int(maximum))

    if headroom:
        maximum += max(1, round(maximum * 0.15))

    axis.scaling.min = 0
    axis.scaling.max = maximum
    axis.majorUnit = (
        1
        if maximum <= 10
        else max(1, (maximum + 4) // 5)
    )
    axis.numFmt = "0"


def _adaptive_line_size(item_count):
    """
    Tamaño proporcional a la cantidad de categorías temporales.

    OpenPyXL expresa width/height en centímetros aproximados.
    """
    count = max(1, _safe_chart_int(item_count))
    width = min(20.2, max(14.8, 11.0 + (count * 0.72)))
    height = min(8.2, max(6.3, 5.9 + (min(count, 18) * 0.12)))
    return round(width, 2), round(height, 2)


def _adaptive_donut_size(item_count):
    count = max(1, _safe_chart_int(item_count))
    width = min(20.0, max(15.0, 13.2 + (count * 0.8)))
    height = min(8.8, max(6.8, 6.1 + (count * 0.42)))
    return round(width, 2), round(height, 2)


def _item_label(item, keys):
    for key in keys:
        value = item.get(key)
        if value not in (None, ""):
            return str(value).strip()
    return ""


def _adaptive_horizontal_size(items, *, label_keys):
    """
    Barras horizontales: la altura depende del número de elementos y
    el ancho crece cuando existen etiquetas largas.

    Se reserva más altura que en la primera versión adaptativa para que
    cada categoría tenga una línea visual propia incluso cuando existen
    8-10 elementos con nombres institucionales largos.
    """
    items = list(items or [])
    count = max(1, len(items))
    max_label = max(
        (len(_item_label(item, label_keys)) for item in items),
        default=0,
    )

    width = 20.6 if max_label >= 28 else 18.8
    height = min(15.4, max(6.4, 4.8 + (count * 0.68)))

    if max_label >= 55:
        height = min(16.2, height + 1.0)

    return round(width, 2), round(height, 2)


def _rows_for_chart_height(height_cm):
    """
    Reserva filas suficientes para que el objeto gráfico no invada
    el bloque siguiente. Se usa una conversión deliberadamente
    conservadora entre centímetros y altura de filas de Excel.
    """
    try:
        height = float(height_cm)
    except (TypeError, ValueError):
        height = 7.0

    return max(16, min(31, int(round(height * 2.45))))


def _evolution_chart_layout(dashboards):
    """
    Devuelve la disposición vertical de los gráficos temporales.
    """
    annual = dashboards.get("publicaciones_por_anio", []) or []
    monthly = (
        dashboards.get("publicaciones_por_mes", {}) or {}
    ).get("items", []) or []
    type_annual = dashboards.get("publicaciones_por_tipo_anual", {}) or {}
    categories = type_annual.get("categorias", []) or []
    series = type_annual.get("series", []) or []

    specs = []

    if annual:
        width, height = _adaptive_line_size(len(annual))
        specs.append({
            "kind": "annual",
            "count": len(annual),
            "width": width,
            "height": height,
        })

    if monthly:
        width, height = _adaptive_line_size(len(monthly))
        specs.append({
            "kind": "monthly",
            "count": len(monthly),
            "width": width,
            "height": height,
        })

    if categories and series:
        count = len(categories)
        width = min(20.2, max(16.5, 12.2 + count * 0.72))
        height = min(
            10.2,
            max(6.8, 6.0 + len(series) * 0.30 + count * 0.10),
        )
        specs.append({
            "kind": "types_by_year",
            "count": count,
            "width": round(width, 2),
            "height": round(height, 2),
        })

    current_row = 12
    layout = []

    for spec in specs:
        panel_rows = _rows_for_chart_height(spec["height"])
        item = dict(spec)
        item["start_row"] = current_row
        item["end_row"] = current_row + panel_rows - 1
        item["anchor"] = f"A{current_row}"
        layout.append(item)
        current_row = item["end_row"] + 2

    return layout


def _highlight_chart_layout(dashboards):
    base_specs = (
        (
            "top_sedes",
            18,
            "Sedes",
            ("sede", "label"),
            REPORT_COLORS["primary"],
        ),
        (
            "top_facultades",
            21,
            "Facultades",
            ("facultad", "label"),
            REPORT_COLORS["primary"],
        ),
        (
            "top_carreras",
            24,
            "Carreras",
            ("carrera", "label"),
            REPORT_COLORS["primary"],
        ),
        (
            "areas",
            27,
            "Áreas de conocimiento",
            ("label",),
            REPORT_COLORS["primary"],
        ),
    )

    current_row = 12
    layout = []

    for key, start_col, title, label_keys, color in base_specs:
        items = (dashboards.get(key, {}) or {}).get("items", []) or []

        if not items:
            continue

        width, height = _adaptive_horizontal_size(
            items,
            label_keys=label_keys,
        )
        panel_rows = _rows_for_chart_height(height)

        layout.append({
            "key": key,
            "start_col": start_col,
            "title": title,
            "label_keys": label_keys,
            "color": color,
            "items": items,
            "item_count": len(items),
            "start_row": current_row,
            "end_row": current_row + panel_rows - 1,
            "anchor": f"A{current_row}",
            "width": width,
            "height": height,
        })

        current_row += panel_rows + 2

    return layout


def _style_line_chart(
    chart,
    *,
    title,
    maximum,
    item_count,
    width=None,
    height=None,
):
    if width is None or height is None:
        width, height = _adaptive_line_size(item_count)

    chart.title = title
    chart.height = height
    chart.width = width
    chart.legend = None
    chart.y_axis.title = "Publicaciones"
    chart.x_axis.title = ""
    _integer_chart_axis(chart.y_axis, maximum)
    chart.dataLabels = DataLabelList()

    # Con demasiados puntos, mostrar cada valor reduce la legibilidad.
    chart.dataLabels.showVal = item_count <= 12

    if chart.series:
        serie = chart.series[0]
        serie.graphicalProperties.line.solidFill = REPORT_COLORS["primary"]
        serie.graphicalProperties.line.width = 24000
        serie.marker.symbol = "circle"
        serie.marker.size = 6
        serie.marker.graphicalProperties.solidFill = REPORT_COLORS["surface"]
        serie.marker.graphicalProperties.line.solidFill = REPORT_COLORS["primary"]


def _add_horizontal_bar_chart(
    ws,
    data_ws,
    *,
    start_col,
    item_count,
    title,
    anchor,
    color,
    width=None,
    height=None,
):
    """Inserta un ranking horizontal legible y proporcional.

    IMPORTANTE SOBRE OPENPYXL:
    Aunque Excel dibuje un ``BarChart(type="bar")`` rotado, OpenPyXL
    conserva ``x_axis`` como eje de categorías (texto) y ``y_axis`` como
    eje de valores (numérico). Configurarlos al revés comprime todas las
    categorías en la base del gráfico y puede invertir la escala, que fue
    el defecto de la primera versión adaptativa.
    """
    if item_count <= 0:
        return False

    if width is None:
        width = 18.8
    if height is None:
        height = min(15.4, max(6.4, 4.8 + item_count * 0.68))

    max_row = item_count + 1
    values = [
        _safe_chart_int(
            data_ws.cell(row=row, column=start_col + 1).value
        )
        for row in range(2, max_row + 1)
    ]
    maximum = max(values or [0])

    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "clustered"
    chart.overlap = 0

    chart.add_data(
        Reference(
            data_ws,
            min_col=start_col + 1,
            min_row=1,
            max_row=max_row,
        ),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(
            data_ws,
            min_col=start_col,
            min_row=2,
            max_row=max_row,
        )
    )

    chart.title = title
    chart.height = height
    chart.width = width
    chart.legend = None

    # En BarChart horizontal de OpenPyXL:
    #   x_axis -> categorías (se dibuja verticalmente)
    #   y_axis -> valores numéricos (se dibuja horizontalmente)
    # La primera versión adaptativa los trataba al revés.
    chart.x_axis.title = ""
    chart.y_axis.title = ""

    # El eje numérico debe crecer de izquierda a derecha: 0 -> máximo.
    _integer_chart_axis(chart.y_axis, maximum)
    chart.y_axis.scaling.orientation = "minMax"
    chart.y_axis.majorGridlines = None

    # El primer elemento del ranking se presenta arriba y las etiquetas
    # permanecen al lado izquierdo del área de trazado.
    chart.x_axis.scaling.orientation = "maxMin"
    chart.x_axis.tickLblPos = "low"

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = item_count <= 20

    if chart.series:
        series = chart.series[0]
        series.graphicalProperties.solidFill = color
        series.graphicalProperties.line.solidFill = color

    ws.add_chart(chart, anchor)
    return True


def _add_summary_charts(summary_ws, data_ws, dashboards):
    publicaciones_por_tipo = dashboards.get("publicaciones_por_tipo", {}) or {}
    type_items = publicaciones_por_tipo.get("items", []) or []

    if not type_items:
        return

    chart = DoughnutChart()
    chart.add_data(
        Reference(
            data_ws,
            min_col=2,
            min_row=1,
            max_row=len(type_items) + 1,
        ),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(
            data_ws,
            min_col=1,
            min_row=2,
            max_row=len(type_items) + 1,
        )
    )

    width, height = _adaptive_donut_size(len(type_items))

    chart.title = "Publicaciones por tipo"
    chart.holeSize = 62
    chart.height = height
    chart.width = width

    # Con el gráfico a ancho completo, la leyenda inferior es más legible
    # y evita recortar nombres largos.
    chart.legend.position = "b"

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = len(type_items) <= 6
    chart.dataLabels.showLeaderLines = len(type_items) <= 6

    if chart.series:
        points = []
        for index, item in enumerate(type_items):
            point = DataPoint(idx=index)
            color = REPORT_TYPE_COLORS.get(
                item.get("tipo_codigo"),
                REPORT_COLORS["muted"],
            )
            point.graphicalProperties.solidFill = color
            point.graphicalProperties.line.solidFill = color
            points.append(point)
        chart.series[0].data_points = points

    summary_ws.add_chart(chart, "A17")


def _add_evolution_charts(ws, data_ws, dashboards):
    publicaciones_por_anio = dashboards.get("publicaciones_por_anio", []) or []
    tipo_anual = dashboards.get("publicaciones_por_tipo_anual", {}) or {}
    mensual = dashboards.get("publicaciones_por_mes", {}) or {}

    layout = {
        item["kind"]: item
        for item in _evolution_chart_layout(dashboards)
    }

    annual_layout = layout.get("annual")
    if publicaciones_por_anio and annual_layout:
        values = [
            _safe_chart_int(item.get("value"))
            for item in publicaciones_por_anio
        ]

        chart = LineChart()
        chart.add_data(
            Reference(
                data_ws,
                min_col=6,
                min_row=1,
                max_row=len(publicaciones_por_anio) + 1,
            ),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(
                data_ws,
                min_col=5,
                min_row=2,
                max_row=len(publicaciones_por_anio) + 1,
            )
        )
        _style_line_chart(
            chart,
            title="Producción por año",
            maximum=max(values or [0]),
            item_count=len(publicaciones_por_anio),
            width=annual_layout["width"],
            height=annual_layout["height"],
        )
        ws.add_chart(chart, annual_layout["anchor"])

    month_items = mensual.get("items", []) or []
    monthly_layout = layout.get("monthly")
    if month_items and monthly_layout:
        values = [
            _safe_chart_int(item.get("value"))
            for item in month_items
        ]

        chart = LineChart()
        chart.add_data(
            Reference(
                data_ws,
                min_col=16,
                min_row=1,
                max_row=len(month_items) + 1,
            ),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(
                data_ws,
                min_col=15,
                min_row=2,
                max_row=len(month_items) + 1,
            )
        )
        title = (
            f"Producción por mes · "
            f"{mensual.get('anio_base') or 'último año con información'}"
        )
        _style_line_chart(
            chart,
            title=title,
            maximum=max(values or [0]),
            item_count=len(month_items),
            width=monthly_layout["width"],
            height=monthly_layout["height"],
        )
        ws.add_chart(chart, monthly_layout["anchor"])

    categories = tipo_anual.get("categorias", []) or []
    series = tipo_anual.get("series", []) or []
    type_layout = layout.get("types_by_year")

    if categories and series and type_layout:
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.overlap = 0

        chart.add_data(
            Reference(
                data_ws,
                min_col=9,
                max_col=8 + len(series),
                min_row=1,
                max_row=len(categories) + 1,
            ),
            titles_from_data=True,
            from_rows=False,
        )
        chart.set_categories(
            Reference(
                data_ws,
                min_col=8,
                min_row=2,
                max_row=len(categories) + 1,
            )
        )

        maximum = max(
            (
                _safe_chart_int(value)
                for serie in series
                for value in (serie.get("data", []) or [])
            ),
            default=0,
        )

        chart.title = "Tipos por año"
        chart.height = type_layout["height"]
        chart.width = type_layout["width"]
        chart.legend.position = "b"
        chart.y_axis.title = "Publicaciones"
        chart.x_axis.title = ""
        _integer_chart_axis(chart.y_axis, maximum)

        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = (
            len(categories) * len(series)
        ) <= 25

        for index, serie in enumerate(series):
            if index >= len(chart.series):
                break

            color = REPORT_TYPE_COLORS.get(
                serie.get("codigo"),
                REPORT_COLORS["primary"],
            )
            chart.series[index].graphicalProperties.solidFill = color
            chart.series[index].graphicalProperties.line.solidFill = color

        ws.add_chart(chart, type_layout["anchor"])


def _add_highlight_charts(ws, data_ws, dashboards):
    for item in _highlight_chart_layout(dashboards):
        _add_horizontal_bar_chart(
            ws,
            data_ws,
            start_col=item["start_col"],
            item_count=item["item_count"],
            title=item["title"],
            anchor=item["anchor"],
            color=item["color"],
            width=item["width"],
            height=item["height"],
        )

# ============================================================
# WORKBOOK
# ============================================================


def _build_report_workbook(payload):
    summary = payload.get("summary", {}) or {}
    dashboards = payload.get("dashboards", {}) or {}
    filtros = payload.get("filtros_aplicados", {}) or {}

    workbook = Workbook()
    workbook.remove(workbook.active)

    workbook.properties.title = "Informe de producción científica - SGPC ULEAM"
    workbook.properties.subject = "Producción científica institucional"
    workbook.properties.creator = "SGPC ULEAM"
    workbook.properties.description = (
        "Informe de producción científica con la misma organización visual del PDF institucional."
    )
    workbook.properties.keywords = "SGPC ULEAM, producción científica, informe"

    publicaciones_por_tipo = dashboards.get("publicaciones_por_tipo", {}) or {}
    publicaciones_por_anio = dashboards.get("publicaciones_por_anio", []) or []
    publicaciones_por_mes = dashboards.get("publicaciones_por_mes", {}) or {}
    publicaciones_por_tipo_anual = dashboards.get("publicaciones_por_tipo_anual", {}) or {}
    top_sedes = dashboards.get("top_sedes", {}) or {}
    top_facultades = dashboards.get("top_facultades", {}) or {}
    top_carreras = dashboards.get("top_carreras", {}) or {}
    top_autores = dashboards.get("top_autores", {}) or {}
    journals = dashboards.get("journals", {}) or {}
    projects = dashboards.get("projects", {}) or {}
    areas = dashboards.get("areas", {}) or {}

    # Las cuatro primeras hojas replican la composición del PDF institucional.
    resumen_ws = _build_summary_sheet(workbook, summary, filtros, dashboards)
    evolucion_ws = _build_evolution_sheet(workbook, filtros, dashboards)
    destacados_ws = _build_highlights_sheet(workbook, filtros, dashboards)
    _build_people_projects_sheet(workbook, filtros, dashboards)

    # Hojas de datos: se conserva el detalle para trabajar en Excel, pero sin
    # exponer códigos internos que no aportan al usuario final.
    _create_sheet(
        workbook,
        "Tipos de publicación",
        "Distribución de publicaciones por tipo.",
        ["Tipo", "Total", "Porcentaje"],
        [
            [
                item.get("tipo_nombre"),
                item.get("total", 0),
                item.get("porcentaje", 0),
            ]
            for item in publicaciones_por_tipo.get("items", []) or []
        ],
        tab_color=REPORT_COLORS["primary"],
        filtros=filtros,
    )

    _create_sheet(
        workbook,
        "Producción por año",
        "Evolución anual de las publicaciones aprobadas.",
        ["Año", "Total"],
        [[item.get("label"), item.get("value", 0)] for item in publicaciones_por_anio],
        tab_color=REPORT_COLORS["green"],
        filtros=filtros,
    )

    _create_sheet(
        workbook,
        "Producción por mes",
        f"Evolución mensual del año {publicaciones_por_mes.get('anio_base') or 'seleccionado'}.",
        ["Mes", "Total"],
        [[item.get("label"), item.get("value", 0)] for item in publicaciones_por_mes.get("items", []) or []],
        tab_color=REPORT_COLORS["teal"],
        filtros=filtros,
    )

    _create_sheet(
        workbook,
        "Tipos por año",
        "Comparación anual de los tipos de publicación.",
        ["Año", "Tipo", "Total"],
        _build_tipo_anual_rows(publicaciones_por_tipo_anual),
        tab_color=REPORT_COLORS["purple"],
        filtros=filtros,
    )

    _create_sheet(
        workbook,
        "Sedes",
        "Sedes ordenadas por número de publicaciones.",
        ["Sede", "Total"],
        [[item.get("sede") or item.get("label"), item.get("total", 0)] for item in top_sedes.get("items", []) or []],
        tab_color=REPORT_COLORS["primary"],
        filtros=filtros,
    )

    _create_sheet(
        workbook,
        "Facultades",
        "Facultades ordenadas por número de publicaciones.",
        ["Facultad", "Total"],
        [[item.get("facultad") or item.get("label"), item.get("total", 0)] for item in top_facultades.get("items", []) or []],
        tab_color=REPORT_COLORS["teal"],
        filtros=filtros,
    )

    _create_sheet(
        workbook,
        "Carreras",
        "Carreras ordenadas por número de publicaciones.",
        ["Carrera", "Total"],
        [[item.get("carrera") or item.get("label"), item.get("total", 0)] for item in top_carreras.get("items", []) or []],
        tab_color=REPORT_COLORS["orange"],
        filtros=filtros,
    )

    _create_sheet(
        workbook,
        "Autores",
        "Autores ordenados por número de publicaciones.",
        ["Autor", "Total publicaciones"],
        [
            [
                item.get("autor") or item.get("label"),
                item.get("total_publicaciones", item.get("total", 0)),
            ]
            for item in top_autores.get("items", []) or []
        ],
        tab_color=REPORT_COLORS["purple"],
        filtros=filtros,
    )

    _create_sheet(
        workbook,
        "Revistas",
        "Revistas ordenadas por número de artículos.",
        ["Revista", "Total"],
        [[item.get("label"), item.get("total", 0)] for item in journals.get("items", []) or []],
        tab_color=REPORT_COLORS["primary"],
        filtros=filtros,
    )

    _create_sheet(
        workbook,
        "Proyectos",
        "Proyectos ordenados por número de publicaciones relacionadas.",
        ["Proyecto", "Total"],
        [[item.get("label"), item.get("total", 0)] for item in projects.get("items", []) or []],
        tab_color=REPORT_COLORS["green"],
        filtros=filtros,
    )

    _create_sheet(
        workbook,
        "Áreas de conocimiento",
        "Áreas de conocimiento ordenadas por número de publicaciones.",
        ["Área", "Total"],
        [[item.get("label"), item.get("total", 0)] for item in areas.get("items", []) or []],
        tab_color=REPORT_COLORS["orange"],
        filtros=filtros,
    )

    chart_data_ws = _build_chart_data_sheet(workbook, dashboards)
    _add_summary_charts(resumen_ws, chart_data_ws, dashboards)
    _add_evolution_charts(evolucion_ws, chart_data_ws, dashboards)
    _add_highlight_charts(destacados_ws, chart_data_ws, dashboards)

    workbook.active = 0
    return workbook


# ============================================================
# INFORME PDF - HELPERS
# ============================================================

PDF_PAGE_WIDTH = landscape(A4)[0] if REPORTLAB_AVAILABLE else 0
PDF_PAGE_HEIGHT = landscape(A4)[1] if REPORTLAB_AVAILABLE else 0
PDF_CONTENT_WIDTH_MM = 269
PDF_HALF_WIDTH_MM = 132
PDF_GAP_MM = 5

PDF_COLORS = {
    "primary": "2563EB",
    "primary_dark": "1D4ED8",
    "primary_soft": "EFF6FF",
    "green": "15803D",
    "purple": "7C3AED",
    "orange": "C2410C",
    "teal": "0F766E",
    "text": "172033",
    "muted": "64748B",
    "surface": "FFFFFF",
    "soft": "F1F5F9",
    "soft_alt": "F8FAFC",
    "line": "D8E1EC",
}


def _require_reportlab():
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError(
            "El generador PDF no está disponible."
        )


def _pdf_report_filename():
    stamp = timezone.localtime().strftime(
        "%Y-%m-%d_%H-%M-%S"
    )
    return (
        "informe-produccion-cientifica-"
        f"sgpc-uleam-{stamp}.pdf"
    )


def _pdf_hex(value):
    token = str(value or "000000").lstrip("#")
    return colors.HexColor(f"#{token}")


def _pdf_safe(value, fallback="No especificado"):
    text = str(value or "").strip()
    return text if text else fallback


def _pdf_styles():
    sample = getSampleStyleSheet()

    return {
        "title": ParagraphStyle(
            "DashboardPdfTitle",
            parent=sample["Title"],
            fontName="Helvetica-Bold",
            fontSize=20,
            leading=23,
            textColor=_pdf_hex(PDF_COLORS["text"]),
            alignment=TA_LEFT,
            spaceAfter=1.5 * mm,
        ),
        "subtitle": ParagraphStyle(
            "DashboardPdfSubtitle",
            parent=sample["Normal"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=11,
            textColor=_pdf_hex(PDF_COLORS["muted"]),
            spaceAfter=2.5 * mm,
        ),
        "section": ParagraphStyle(
            "DashboardPdfSection",
            parent=sample["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=12,
            leading=14,
            textColor=_pdf_hex(PDF_COLORS["text"]),
            spaceBefore=1.5 * mm,
            spaceAfter=1.5 * mm,
        ),
        "section_bar": ParagraphStyle(
            "DashboardPdfSectionBar",
            parent=sample["Normal"],
            fontName="Helvetica-Bold",
            fontSize=11.5,
            leading=13.5,
            textColor=colors.white,
        ),
        "section_bar_subtitle": ParagraphStyle(
            "DashboardPdfSectionBarSubtitle",
            parent=sample["Normal"],
            fontName="Helvetica",
            fontSize=7,
            leading=9,
            textColor=_pdf_hex("DBEAFE"),
        ),
        "body": ParagraphStyle(
            "DashboardPdfBody",
            parent=sample["Normal"],
            fontName="Helvetica",
            fontSize=7.8,
            leading=10.2,
            textColor=_pdf_hex(PDF_COLORS["text"]),
        ),
        "small": ParagraphStyle(
            "DashboardPdfSmall",
            parent=sample["Normal"],
            fontName="Helvetica",
            fontSize=6.8,
            leading=8.6,
            textColor=_pdf_hex(PDF_COLORS["muted"]),
        ),
        "table_header": ParagraphStyle(
            "DashboardPdfTableHeader",
            parent=sample["Normal"],
            fontName="Helvetica-Bold",
            fontSize=7.2,
            leading=8.5,
            textColor=_pdf_hex(PDF_COLORS["text"]),
        ),
        "table": ParagraphStyle(
            "DashboardPdfTable",
            parent=sample["Normal"],
            fontName="Helvetica",
            fontSize=6.9,
            leading=8.3,
            textColor=_pdf_hex(PDF_COLORS["text"]),
        ),
        "kpi_value": ParagraphStyle(
            "DashboardPdfKpiValue",
            parent=sample["Normal"],
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=18,
            textColor=_pdf_hex(PDF_COLORS["text"]),
            alignment=TA_CENTER,
        ),
        "kpi_label": ParagraphStyle(
            "DashboardPdfKpiLabel",
            parent=sample["Normal"],
            fontName="Helvetica",
            fontSize=6.8,
            leading=8.2,
            textColor=_pdf_hex(PDF_COLORS["muted"]),
            alignment=TA_CENTER,
        ),
        "panel_title": ParagraphStyle(
            "DashboardPdfPanelTitle",
            parent=sample["Normal"],
            fontName="Helvetica-Bold",
            fontSize=7.6,
            leading=9,
            textColor=_pdf_hex(PDF_COLORS["text"]),
        ),
    }


def _pdf_p(value, style):
    return Paragraph(
        _pdf_safe(value, ""),
        style,
    )


def _pdf_period_label(filtros):
    desde = filtros.get("mes_desde")
    hasta = filtros.get("mes_hasta")
    anio_desde = filtros.get("anio_desde")
    anio_hasta = filtros.get("anio_hasta")

    if desde and hasta:
        return desde if desde == hasta else f"{desde} a {hasta}"
    if desde:
        return f"Desde {desde}"
    if hasta:
        return f"Hasta {hasta}"
    if anio_desde and anio_hasta:
        return str(anio_desde) if anio_desde == anio_hasta else f"{anio_desde} a {anio_hasta}"
    if anio_desde:
        return f"Desde {anio_desde}"
    if anio_hasta:
        return f"Hasta {anio_hasta}"
    return "Todo el período"


def _pdf_header_footer(canvas, document):
    canvas.saveState()
    width, height = document.pagesize

    if document.page > 1:
        canvas.setFont("Helvetica-Bold", 7)
        canvas.setFillColor(_pdf_hex(PDF_COLORS["muted"]))
        canvas.drawString(
            document.leftMargin,
            height - 8 * mm,
            "SGPC ULEAM - Informe de producción científica",
        )
        canvas.setStrokeColor(_pdf_hex(PDF_COLORS["line"]))
        canvas.setLineWidth(0.45)
        canvas.line(
            document.leftMargin,
            height - 10 * mm,
            width - document.rightMargin,
            height - 10 * mm,
        )

    canvas.setStrokeColor(_pdf_hex(PDF_COLORS["line"]))
    canvas.setLineWidth(0.45)
    canvas.line(
        document.leftMargin,
        13 * mm,
        width - document.rightMargin,
        13 * mm,
    )

    canvas.setFont("Helvetica", 6.8)
    canvas.setFillColor(_pdf_hex(PDF_COLORS["muted"]))
    canvas.drawString(
        document.leftMargin,
        8 * mm,
        "SGPC ULEAM - Producción científica",
    )
    canvas.drawRightString(
        width - document.rightMargin,
        8 * mm,
        f"Página {document.page}",
    )
    canvas.restoreState()


def _pdf_section_header(title, styles, subtitle=None):
    content = [Paragraph(title, styles["section_bar"])]
    if subtitle:
        content.append(
            Paragraph(subtitle, styles["section_bar_subtitle"])
        )

    table = Table(
        [[content]],
        colWidths=[PDF_CONTENT_WIDTH_MM * mm],
        hAlign="LEFT",
    )
    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), _pdf_hex(PDF_COLORS["primary"])),
            ("LEFTPADDING", (0, 0), (-1, -1), 9),
            ("RIGHTPADDING", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])
    )
    return table


def _pdf_filters_table(filtros, styles):
    tipo = filtros.get("tipo_nombre") or "Todos los tipos"
    entries = [
        ("Sede", filtros.get("sede_nombre") or "Todas las sedes"),
        ("Facultad", filtros.get("facultad_nombre") or "Todas las facultades"),
        ("Carrera", filtros.get("carrera_nombre") or "Todas las carreras"),
        ("Tipo", tipo),
        ("Período", _pdf_period_label(filtros)),
        ("Generado", _report_generated_at()),
    ]

    cells = [
        Paragraph(
            f"<b>{label}</b><br/>{_pdf_safe(value)}",
            styles["body"],
        )
        for label, value in entries
    ]

    table = Table(
        [cells[:3], cells[3:]],
        colWidths=[(PDF_CONTENT_WIDTH_MM / 3) * mm] * 3,
        hAlign="LEFT",
    )
    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), _pdf_hex(PDF_COLORS["soft"])),
            ("BOX", (0, 0), (-1, -1), 0.45, _pdf_hex(PDF_COLORS["line"])),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, _pdf_hex(PDF_COLORS["line"])),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ])
    )
    return table


def _pdf_kpi_table(summary, styles):
    items = [
        ("Publicaciones", summary.get("total_publicaciones", 0)),
        ("Autores", summary.get("total_autores", 0)),
        ("Sedes", summary.get("total_sedes", 0)),
        ("Facultades", summary.get("total_facultades", 0)),
        ("Carreras", summary.get("total_carreras", 0)),
        ("Proyectos", summary.get("total_proyectos", 0)),
    ]

    row = []
    for label, value in items:
        row.append(
            Table(
                [
                    [Paragraph(f"{int(value or 0):,}", styles["kpi_value"])],
                    [Paragraph(label, styles["kpi_label"])],
                ],
                colWidths=[(PDF_CONTENT_WIDTH_MM / 6 - 1) * mm],
            )
        )

    table = Table(
        [row],
        colWidths=[(PDF_CONTENT_WIDTH_MM / 6) * mm] * 6,
        hAlign="LEFT",
    )
    style = [
        ("BACKGROUND", (0, 0), (-1, -1), _pdf_hex(PDF_COLORS["surface"])),
        ("BOX", (0, 0), (-1, -1), 0.45, _pdf_hex(PDF_COLORS["line"])),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, _pdf_hex(PDF_COLORS["line"])),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]
    for index in range(6):
        style.append(("LINEABOVE", (index, 0), (index, 0), 2.2, _pdf_hex(REPORT_KPI_COLORS[index])))
    table.setStyle(TableStyle(style))
    return table


def _pdf_axis_max(values):
    max_value = max([int(value or 0) for value in values] or [0])
    if max_value <= 5:
        return 5
    step = 5
    return ((max_value + step - 1) // step) * step


def _pdf_type_donut(publicaciones_por_tipo, *, width=340, height=205):
    items = publicaciones_por_tipo.get("items", []) or []
    drawing = Drawing(width, height)

    if not items:
        drawing.add(String(20, height / 2, "No hay publicaciones por tipo para esta selección.", fontName="Helvetica", fontSize=8, fillColor=_pdf_hex(PDF_COLORS["muted"])))
        return drawing

    pie = Pie()
    pie.x = 24
    pie.y = 26
    pie.width = 145
    pie.height = 145
    pie.data = [int(item.get("total", 0) or 0) for item in items]

    short_labels = {
        "AAI": "Alto impacto",
        "AR": "Regional",
        "PON": "Ponencia",
        "CAP": "Capítulo",
        "LIB": "Libro",
    }
    pie.labels = [short_labels.get(item.get("tipo_codigo"), item.get("tipo_nombre") or "Tipo") for item in items]
    pie.simpleLabels = 0
    pie.sideLabels = 1
    pie.slices.strokeWidth = 0.5
    pie.slices.strokeColor = colors.white

    for index, item in enumerate(items):
        color_value = REPORT_TYPE_COLORS.get(item.get("tipo_codigo"), REPORT_COLORS["primary"])
        pie.slices[index].fillColor = _pdf_hex(color_value)
        pie.slices[index].fontName = "Helvetica"
        pie.slices[index].fontSize = 6.3

    drawing.add(pie)
    drawing.add(Circle(96.5, 98.5, 32, fillColor=colors.white, strokeColor=colors.white))
    total = sum(pie.data)
    drawing.add(String(96.5, 101, str(total), textAnchor="middle", fontName="Helvetica-Bold", fontSize=14, fillColor=_pdf_hex(PDF_COLORS["text"])))
    drawing.add(String(96.5, 88, "publicaciones", textAnchor="middle", fontName="Helvetica", fontSize=6.3, fillColor=_pdf_hex(PDF_COLORS["muted"])))
    return drawing


def _pdf_line_chart(items, *, width=350, height=128):
    drawing = Drawing(width, height)
    values = [int(item.get("value", 0) or 0) for item in items]
    labels = [str(item.get("label", "")) for item in items]

    if not items:
        drawing.add(String(20, height / 2, "No hay información para esta selección.", fontName="Helvetica", fontSize=8, fillColor=_pdf_hex(PDF_COLORS["muted"])))
        return drawing

    chart = HorizontalLineChart()
    chart.x = 34
    chart.y = 25
    chart.width = width - 50
    chart.height = height - 44
    chart.data = [values]
    chart.categoryAxis.categoryNames = labels
    chart.categoryAxis.labels.fontName = "Helvetica"
    chart.categoryAxis.labels.fontSize = 6.2
    chart.categoryAxis.labels.boxAnchor = "n"
    chart.valueAxis.valueMin = 0
    chart.valueAxis.valueMax = _pdf_axis_max(values)
    chart.valueAxis.valueStep = max(1, chart.valueAxis.valueMax // 5)
    chart.valueAxis.labels.fontName = "Helvetica"
    chart.valueAxis.labels.fontSize = 6.2
    chart.valueAxis.strokeColor = _pdf_hex(PDF_COLORS["line"])
    chart.categoryAxis.strokeColor = _pdf_hex(PDF_COLORS["line"])
    chart.lines[0].strokeColor = _pdf_hex(PDF_COLORS["primary"])
    chart.lines[0].strokeWidth = 2

    try:
        marker = makeMarker("Circle")
        marker.size = 4
        marker.fillColor = colors.white
        marker.strokeColor = _pdf_hex(PDF_COLORS["primary"])
        marker.strokeWidth = 1.2
        chart.lines[0].symbol = marker
    except Exception:
        chart.lines[0].symbol = None

    drawing.add(chart)
    return drawing


def _pdf_short_chart_label(value, max_chars=34):
    text = _pdf_safe(value, "Sin dato")
    if len(text) <= max_chars:
        return text
    return text[: max_chars - 3].rstrip() + "..."


def _pdf_rank_bar_chart(
    items,
    *,
    label_keys=("label",),
    width=350,
    height=None,
):
    normalized = []
    for item in items or []:
        label = next(
            (item.get(key) for key in label_keys if item.get(key)),
            None,
        )
        normalized.append(
            (
                _pdf_short_chart_label(label),
                int(item.get("total", 0) or 0),
            )
        )

    if height is None:
        height = max(
            125,
            min(180, 90 + (len(normalized) * 4.5)),
        )

    drawing = Drawing(width, height)
    if not normalized:
        drawing.add(
            String(
                18,
                height / 2,
                "No hay información para esta selección.",
                fontName="Helvetica",
                fontSize=8,
                fillColor=_pdf_hex(PDF_COLORS["muted"]),
            )
        )
        return drawing

    # HorizontalBarChart coloca la primera categoría abajo; se invierte
    # para que el primer puesto se muestre visualmente arriba.
    labels = [row[0] for row in reversed(normalized)]
    values = [row[1] for row in reversed(normalized)]

    chart = HorizontalBarChart()
    chart.x = 132
    chart.y = 18
    chart.width = width - 165
    chart.height = height - 34
    chart.data = [values]
    chart.categoryAxis.categoryNames = labels
    chart.categoryAxis.labels.fontName = "Helvetica"
    chart.categoryAxis.labels.fontSize = 5.8 if len(labels) > 12 else 6.4
    chart.categoryAxis.labels.boxAnchor = "e"
    chart.categoryAxis.strokeColor = _pdf_hex(PDF_COLORS["line"])
    chart.valueAxis.valueMin = 0
    chart.valueAxis.valueMax = _pdf_axis_max(values)
    chart.valueAxis.valueStep = max(1, chart.valueAxis.valueMax // 4)
    chart.valueAxis.labels.fontName = "Helvetica"
    chart.valueAxis.labels.fontSize = 5.5
    chart.valueAxis.labels.fillColor = _pdf_hex(PDF_COLORS["muted"])
    chart.valueAxis.strokeColor = _pdf_hex(PDF_COLORS["line"])
    chart.bars[0].fillColor = _pdf_hex(PDF_COLORS["primary"])
    chart.bars[0].strokeColor = None
    chart.barLabelFormat = "%d"
    chart.barLabels.fontName = "Helvetica-Bold"
    chart.barLabels.fontSize = 5.8
    chart.barLabels.fillColor = _pdf_hex(PDF_COLORS["text"])
    chart.barLabels.nudge = 5
    chart.barLabels.boxAnchor = "w"

    drawing.add(chart)
    return drawing


def _pdf_type_year_chart(data, *, width=740, height=150):
    categories = data.get("categorias", []) or []
    series = data.get("series", []) or []
    drawing = Drawing(width, height)

    if not categories or not series:
        drawing.add(String(20, height / 2, "No hay información por tipo y año para esta selección.", fontName="Helvetica", fontSize=8, fillColor=_pdf_hex(PDF_COLORS["muted"])))
        return drawing

    chart = VerticalBarChart()
    chart.x = 42
    chart.y = 38
    chart.width = width - 66
    chart.height = height - 60
    chart.data = [[int(value or 0) for value in serie.get("data", [])] for serie in series]
    chart.categoryAxis.categoryNames = [str(value) for value in categories]
    chart.categoryAxis.labels.fontName = "Helvetica"
    chart.categoryAxis.labels.fontSize = 6.1
    chart.valueAxis.valueMin = 0
    all_values = [value for serie in chart.data for value in serie]
    chart.valueAxis.valueMax = _pdf_axis_max(all_values)
    chart.valueAxis.valueStep = max(1, chart.valueAxis.valueMax // 5)
    chart.valueAxis.labels.fontName = "Helvetica"
    chart.valueAxis.labels.fontSize = 6.1
    chart.valueAxis.strokeColor = _pdf_hex(PDF_COLORS["line"])
    chart.categoryAxis.strokeColor = _pdf_hex(PDF_COLORS["line"])
    chart.barSpacing = 1.4
    chart.groupSpacing = 6

    for index, serie in enumerate(series):
        code = serie.get("codigo")
        chart.bars[index].fillColor = _pdf_hex(REPORT_TYPE_COLORS.get(code, REPORT_COLORS["primary"]))
        chart.bars[index].strokeColor = None
    drawing.add(chart)

    legend_y = 10
    legend_slots = max(1, len(series))
    slot_width = (width - 40) / legend_slots
    for index, serie in enumerate(series):
        code = serie.get("codigo")
        label = _pdf_safe(serie.get("label"), "Tipo")
        item_x = 22 + (index * slot_width)
        drawing.add(Rect(item_x, legend_y, 6, 6, fillColor=_pdf_hex(REPORT_TYPE_COLORS.get(code, REPORT_COLORS["primary"])), strokeColor=None))
        drawing.add(String(item_x + 9, legend_y, label[:24], fontName="Helvetica", fontSize=5.7, fillColor=_pdf_hex(PDF_COLORS["muted"])))
    return drawing


def _pdf_chart_panel(title, drawing, styles, *, width_mm=132):
    table = Table(
        [[Paragraph(title, styles["panel_title"])], [drawing]],
        colWidths=[width_mm * mm],
        hAlign="LEFT",
    )
    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _pdf_hex(PDF_COLORS["soft"])),
            ("BACKGROUND", (0, 1), (-1, -1), _pdf_hex(PDF_COLORS["surface"])),
            ("BOX", (0, 0), (-1, -1), 0.45, _pdf_hex(PDF_COLORS["line"])),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ])
    )
    return table


def _pdf_ranking_table(
    column_label,
    items,
    styles,
    *,
    label_keys=("label",),
    value_key="total",
    value_label="Publicaciones",
    width_mm=132,
):
    position_width = 12
    value_width = 27
    label_width = max(45, width_mm - position_width - value_width)

    rows = [[
        _pdf_p("Pos.", styles["table_header"]),
        _pdf_p(column_label, styles["table_header"]),
        _pdf_p(value_label, styles["table_header"]),
    ]]

    for index, item in enumerate(items or [], start=1):
        label = next((item.get(key) for key in label_keys if item.get(key)), None)
        rows.append([
            _pdf_p(index, styles["table"]),
            _pdf_p(label, styles["table"]),
            _pdf_p(int(item.get(value_key, 0) or 0), styles["table"]),
        ])

    if len(rows) == 1:
        rows.append(["", _pdf_p("No hay información para esta selección.", styles["small"]), ""])

    table = Table(
        rows,
        colWidths=[position_width * mm, label_width * mm, value_width * mm],
        repeatRows=1,
        hAlign="LEFT",
    )
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), _pdf_hex(PDF_COLORS["soft"])),
        ("GRID", (0, 0), (-1, -1), 0.3, _pdf_hex(PDF_COLORS["line"])),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (2, 1), (2, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 3.3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3.3),
    ]
    for row_index in range(2, len(rows), 2):
        style.append(("BACKGROUND", (0, row_index), (-1, row_index), _pdf_hex(PDF_COLORS["soft_alt"])))
    table.setStyle(TableStyle(style))

    if len(rows) == 2 and not items:
        table.setStyle(TableStyle([("SPAN", (1, 1), (2, 1))]))
    return table


def _pdf_ranking_panel(
    title,
    column_label,
    items,
    styles,
    *,
    label_keys=("label",),
    value_key="total",
    value_label="Publicaciones",
    width_mm=132,
):
    inner = _pdf_ranking_table(
        column_label,
        items,
        styles,
        label_keys=label_keys,
        value_key=value_key,
        value_label=value_label,
        width_mm=width_mm,
    )
    panel = Table(
        [[Paragraph(title, styles["panel_title"])], [inner]],
        colWidths=[width_mm * mm],
        hAlign="LEFT",
    )
    panel.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _pdf_hex(PDF_COLORS["primary_soft"])),
            ("BOX", (0, 0), (-1, -1), 0.45, _pdf_hex(PDF_COLORS["line"])),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 1), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 0),
            ("LEFTPADDING", (0, 0), (-1, 0), 6),
            ("RIGHTPADDING", (0, 0), (-1, 0), 6),
            ("TOPPADDING", (0, 0), (-1, 0), 5),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ])
    )
    return panel


def _pdf_pair(left, right, *, left_width=132, right_width=132, gap_mm=5):
    pair = Table(
        [[left, "", right]],
        colWidths=[left_width * mm, gap_mm * mm, right_width * mm],
        hAlign="LEFT",
    )
    pair.setStyle(
        TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ])
    )
    return pair


def _pdf_stack(flowables, *, width_mm=132, gap_mm=2):
    rows = []
    for index, flowable in enumerate(flowables):
        if index:
            rows.append([Spacer(1, gap_mm * mm)])
        rows.append([flowable])
    table = Table(rows, colWidths=[width_mm * mm], hAlign="LEFT")
    table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    return table


def _pdf_note(text, styles):
    table = Table(
        [[Paragraph(text, styles["small"])]],
        colWidths=[PDF_CONTENT_WIDTH_MM * mm],
        hAlign="LEFT",
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), _pdf_hex(PDF_COLORS["soft_alt"])),
        ("BOX", (0, 0), (-1, -1), 0.35, _pdf_hex(PDF_COLORS["line"])),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    return table


def _build_report_pdf(payload):
    _require_reportlab()

    summary = payload.get("summary", {}) or {}
    dashboards = payload.get("dashboards", {}) or {}
    filtros = payload.get("filtros_aplicados", {}) or {}
    styles = _pdf_styles()

    publicaciones_por_tipo = dashboards.get("publicaciones_por_tipo", {}) or {}
    publicaciones_por_anio = dashboards.get("publicaciones_por_anio", []) or []
    publicaciones_por_mes = dashboards.get("publicaciones_por_mes", {}) or {}
    publicaciones_por_tipo_anual = dashboards.get("publicaciones_por_tipo_anual", {}) or {}
    top_sedes = (dashboards.get("top_sedes", {}) or {}).get("items", []) or []
    top_facultades = (dashboards.get("top_facultades", {}) or {}).get("items", []) or []
    top_carreras = (dashboards.get("top_carreras", {}) or {}).get("items", []) or []
    top_autores = (dashboards.get("top_autores", {}) or {}).get("items", []) or []
    areas = (dashboards.get("areas", {}) or {}).get("items", []) or []
    journals = (dashboards.get("journals", {}) or {}).get("items", []) or []
    projects = (dashboards.get("projects", {}) or {}).get("items", []) or []

    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=15 * mm,
        bottomMargin=18 * mm,
        title="Informe de producción científica",
        author="SGPC ULEAM",
        subject="Producción científica institucional",
    )

    # --------------------------------------------------------
    # PÁGINA 1 - RESUMEN
    # --------------------------------------------------------
    story = [
        Paragraph("Informe de producción científica", styles["title"]),
        Paragraph(
            "Publicaciones aprobadas según las opciones seleccionadas en el panel.",
            styles["subtitle"],
        ),
        _pdf_filters_table(filtros, styles),
        Spacer(1, 3 * mm),
        _pdf_section_header(
            "Resumen",
            styles,
            "Vista general de la producción científica incluida en este informe.",
        ),
        Spacer(1, 2.5 * mm),
        _pdf_kpi_table(summary, styles),
        Spacer(1, 3 * mm),
    ]

    donut_panel = _pdf_chart_panel(
        "Publicaciones por tipo",
        _pdf_type_donut(publicaciones_por_tipo, width=340, height=205),
        styles,
        width_mm=127,
    )
    summary_rankings = _pdf_stack(
        [
            _pdf_ranking_panel(
                "Sedes",
                "Sede",
                top_sedes[:3],
                styles,
                label_keys=("sede", "label"),
                width_mm=137,
            ),
            _pdf_ranking_panel(
                "Facultades",
                "Facultad",
                top_facultades[:3],
                styles,
                label_keys=("facultad", "label"),
                width_mm=137,
            ),
            _pdf_ranking_panel(
                "Autores",
                "Autor",
                top_autores[:3],
                styles,
                label_keys=("autor", "label"),
                width_mm=137,
            ),
        ],
        width_mm=137,
        gap_mm=2,
    )
    story.extend([
        _pdf_pair(
            donut_panel,
            summary_rankings,
            left_width=127,
            right_width=137,
            gap_mm=5,
        ),
        PageBreak(),
    ])

    # --------------------------------------------------------
    # PÁGINA 2 - EVOLUCIÓN
    # Dos líneas temporales arriba y comparación por tipo abajo.
    # --------------------------------------------------------
    story.extend([
        _pdf_section_header(
            "Evolución",
            styles,
            "Comportamiento de la producción científica a lo largo del tiempo.",
        ),
        Spacer(1, 3 * mm),
    ])

    annual_panel = _pdf_chart_panel(
        "Producción por año",
        _pdf_line_chart(publicaciones_por_anio, width=350, height=132),
        styles,
        width_mm=132,
    )
    monthly_panel = _pdf_chart_panel(
        f"Producción por mes - {publicaciones_por_mes.get('anio_base') or 'último año con información'}",
        _pdf_line_chart(publicaciones_por_mes.get("items", []) or [], width=350, height=132),
        styles,
        width_mm=132,
    )
    story.extend([
        _pdf_pair(annual_panel, monthly_panel, left_width=132, right_width=132, gap_mm=5),
        Spacer(1, 3 * mm),
        _pdf_chart_panel(
            "Tipos por año",
            _pdf_type_year_chart(publicaciones_por_tipo_anual, width=740, height=155),
            styles,
            width_mm=269,
        ),
        PageBreak(),
    ])

    # --------------------------------------------------------
    # DESTACADOS INSTITUCIONALES
    # Cada pareja usa todo el ancho. Si una segunda pareja no cabe,
    # ReportLab la mueve completa a la página siguiente.
    # --------------------------------------------------------
    story.extend([
        _pdf_section_header(
            "Destacados",
            styles,
            "Unidades académicas y áreas con mayor producción en la selección actual.",
        ),
        Spacer(1, 3 * mm),
        _pdf_pair(
            _pdf_chart_panel(
                "Sedes",
                _pdf_rank_bar_chart(
                    top_sedes,
                    label_keys=("sede", "label"),
                    width=350,
                ),
                styles,
                width_mm=132,
            ),
            _pdf_chart_panel(
                "Facultades",
                _pdf_rank_bar_chart(
                    top_facultades,
                    label_keys=("facultad", "label"),
                    width=350,
                ),
                styles,
                width_mm=132,
            ),
            left_width=132, right_width=132, gap_mm=5,
        ),
        Spacer(1, 3 * mm),
        _pdf_pair(
            _pdf_chart_panel(
                "Carreras",
                _pdf_rank_bar_chart(
                    top_carreras,
                    label_keys=("carrera", "label"),
                    width=350,
                ),
                styles,
                width_mm=132,
            ),
            _pdf_chart_panel(
                "Áreas de conocimiento",
                _pdf_rank_bar_chart(
                    areas,
                    label_keys=("label",),
                    width=350,
                ),
                styles,
                width_mm=132,
            ),
            left_width=132, right_width=132, gap_mm=5,
        ),
        PageBreak(),
    ])

    # --------------------------------------------------------
    # AUTORES, REVISTAS Y PROYECTOS
    # Autores y revistas comparten fila. Proyectos usa ancho completo,
    # lo que reduce saltos de línea en nombres largos.
    # --------------------------------------------------------
    story.extend([
        _pdf_section_header(
            "Autores, revistas y proyectos",
            styles,
            "Principales participantes y espacios de publicación vinculados con la producción mostrada.",
        ),
        Spacer(1, 3 * mm),
        _pdf_pair(
            _pdf_ranking_panel(
                "Autores", "Autor", top_autores, styles,
                label_keys=("autor", "label"), width_mm=132,
            ),
            _pdf_ranking_panel(
                "Revistas", "Revista", journals, styles,
                label_keys=("label",), value_label="Artículos", width_mm=132,
            ),
            left_width=132, right_width=132, gap_mm=5,
        ),
        Spacer(1, 3 * mm),
        CondPageBreak(70 * mm),
        _pdf_ranking_panel(
            "Proyectos",
            "Proyecto",
            projects,
            styles,
            label_keys=("label",),
            width_mm=269,
        ),
        Spacer(1, 3 * mm),
        _pdf_note(
            "El PDF está preparado para consulta y presentación. Para ordenar, filtrar o trabajar con los datos, utilice la opción Excel.",
            styles,
        ),
    ])

    document.build(
        story,
        onFirstPage=_pdf_header_footer,
        onLaterPages=_pdf_header_footer,
    )
    return output.getvalue()


# ============================================================
# API — INFORME PDF
# ============================================================


class DashboardReportePdfView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        payload = _build_dashboard_payload(
            request.query_params
        )

        try:
            file_bytes = _build_report_pdf(
                payload
            )
        except RuntimeError:
            return Response(
                {
                    "ok": False,
                    "detail": (
                        "El formato PDF no está "
                        "disponible temporalmente."
                    ),
                },
                status=503,
            )

        filename = _pdf_report_filename()

        response = HttpResponse(
            file_bytes,
            content_type="application/pdf",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{filename}"'
        )
        response["Content-Length"] = str(
            len(file_bytes)
        )
        response["Cache-Control"] = (
            "private, no-store, no-cache, "
            "must-revalidate"
        )
        response["Pragma"] = "no-cache"
        response["Expires"] = "0"
        response["X-Content-Type-Options"] = (
            "nosniff"
        )
        return response


# ============================================================
# API — REPORTE EXCEL
# ============================================================


class DashboardReporteExcelView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        payload = _build_dashboard_payload(request.query_params)
        workbook = _build_report_workbook(payload)

        output = BytesIO()
        workbook.save(output)
        file_bytes = output.getvalue()
        filename = _excel_report_filename()

        response = HttpResponse(
            file_bytes,
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = str(len(file_bytes))
        response["Cache-Control"] = "private, no-store"
        response["X-Content-Type-Options"] = "nosniff"
        return response