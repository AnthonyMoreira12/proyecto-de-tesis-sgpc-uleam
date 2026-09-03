"""
Reporte institucional de gestión científica.

Este servicio reutiliza exactamente la misma capa de cálculo del
Dashboard de Gestión para evitar divergencias entre:

- indicadores visualizados;
- vista previa del reporte;
- archivo Excel exportado.

El reporte NO sustituye la matriz detallada de publicaciones que ya
existe en:

    /reportes/publicaciones/excel/

Este reporte está orientado a gestión institucional y contiene:

- resumen ejecutivo;
- estados del flujo;
- alertas;
- producción anual;
- producción por tipo;
- comparativa por Sede;
- comparativa por Facultad;
- comparativa por Carrera;
- comparativa por Proyecto;
- proyectos sin producción;
- cola de revisión;
- actividad reciente.
"""

from datetime import datetime, time as datetime_time
from io import BytesIO

from django.utils import timezone

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from core.dashboard_gestion import (
    build_management_dashboard_payload,
)


# ============================================================
# ESTILO
# ============================================================

TITLE_FILL = "17365D"
TITLE_TEXT = "FFFFFF"
SECTION_FILL = "D9EAF7"
SECTION_TEXT = "17365D"
HEADER_FILL = "5B9BD5"
HEADER_TEXT = "FFFFFF"
ALT_FILL = "F3F6F9"
BORDER_COLOR = "B8C4CE"
ALERT_FILL = "FFF2CC"

THIN_BORDER = Border(
    left=Side(
        style="thin",
        color=BORDER_COLOR,
    ),
    right=Side(
        style="thin",
        color=BORDER_COLOR,
    ),
    top=Side(
        style="thin",
        color=BORDER_COLOR,
    ),
    bottom=Side(
        style="thin",
        color=BORDER_COLOR,
    ),
)


# ============================================================
# UTILIDADES
# ============================================================

def _safe_value(value):
    """
    Convierte valores del reporte a tipos compatibles con Excel.

    openpyxl no admite datetime/time con zona horaria. Los datetime
    conscientes se convierten primero a la zona horaria local de Django
    y luego se vuelven naive para conservar la hora visible esperada.
    """
    if value is None:
        return "—"

    if isinstance(
        value,
        bool,
    ):
        return (
            "Sí"
            if value
            else "No"
        )

    if isinstance(
        value,
        datetime,
    ):
        if timezone.is_aware(
            value
        ):
            value = timezone.localtime(
                value
            )

        return value.replace(
            tzinfo=None
        )

    if isinstance(
        value,
        datetime_time,
    ):
        if (
            value.tzinfo is not None
            and value.utcoffset() is not None
        ):
            return value.replace(
                tzinfo=None
            )

    return value


def _relation_name(value):
    """
    Normaliza relaciones del payload del dashboard para el Excel.

    El dashboard de gestión entrega sede, facultad y carrera como
    cadenas de texto. Se admite también un diccionario con la clave
    "nombre" para conservar compatibilidad con payloads anteriores.
    """
    if isinstance(value, dict):
        value = (
            value.get("nombre")
            or value.get("label")
            or value.get("name")
        )

    if value is None:
        return None

    normalized = str(value).strip()
    return normalized or None


def _sheet_title(
    worksheet,
    title,
    subtitle=None,
):
    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=8,
    )

    title_cell = worksheet.cell(
        row=1,
        column=1,
        value=title,
    )

    title_cell.fill = PatternFill(
        "solid",
        fgColor=TITLE_FILL,
    )

    title_cell.font = Font(
        bold=True,
        color=TITLE_TEXT,
        size=14,
    )

    title_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
    )

    worksheet.row_dimensions[
        1
    ].height = 24

    if subtitle:
        worksheet.merge_cells(
            start_row=2,
            start_column=1,
            end_row=2,
            end_column=8,
        )

        subtitle_cell = worksheet.cell(
            row=2,
            column=1,
            value=subtitle,
        )

        subtitle_cell.font = Font(
            italic=True,
            color="475569",
            size=10,
        )

        subtitle_cell.alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )


def _section_header(
    worksheet,
    row,
    title,
    end_column=8,
):
    worksheet.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=end_column,
    )

    cell = worksheet.cell(
        row=row,
        column=1,
        value=title,
    )

    cell.fill = PatternFill(
        "solid",
        fgColor=SECTION_FILL,
    )

    cell.font = Font(
        bold=True,
        color=SECTION_TEXT,
    )

    cell.alignment = Alignment(
        vertical="center",
    )

    return row + 1


def _write_table(
    worksheet,
    *,
    start_row,
    headers,
    rows,
):
    header_row = start_row

    for column_index, header in enumerate(
        headers,
        start=1,
    ):
        cell = worksheet.cell(
            row=header_row,
            column=column_index,
            value=header,
        )

        cell.fill = PatternFill(
            "solid",
            fgColor=HEADER_FILL,
        )

        cell.font = Font(
            bold=True,
            color=HEADER_TEXT,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        cell.border = THIN_BORDER

    current_row = header_row + 1

    for row_index, values in enumerate(
        rows,
        start=0,
    ):
        for column_index, value in enumerate(
            values,
            start=1,
        ):
            cell = worksheet.cell(
                row=current_row,
                column=column_index,
                value=_safe_value(
                    value
                ),
            )

            if row_index % 2 == 1:
                cell.fill = PatternFill(
                    "solid",
                    fgColor=ALT_FILL,
                )

            cell.border = THIN_BORDER

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

        current_row += 1

    return current_row


def _autofit(
    worksheet,
    *,
    min_width=12,
    max_width=42,
):
    for column_cells in worksheet.columns:
        length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:
            try:
                value = str(
                    cell.value
                    if cell.value is not None
                    else ""
                )
            except Exception:
                value = ""

            if len(value) > length:
                length = len(value)

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(
                length + 2,
                min_width,
            ),
            max_width,
        )


def _freeze_and_filter(
    worksheet,
    *,
    freeze_cell,
    auto_filter_range=None,
):
    worksheet.freeze_panes = freeze_cell

    if auto_filter_range:
        worksheet.auto_filter.ref = (
            auto_filter_range
        )


def _format_percent(
    value,
):
    try:
        return f"{float(value):.2f}%"
    except (
        TypeError,
        ValueError,
    ):
        return "0.00%"




def _add_gestion_line_chart(source_ws, target_ws, *, label_col, value_cols, title, anchor):
    if source_ws.max_row < 5:
        return

    chart = LineChart()
    chart.title = title
    chart.height = 7.2
    chart.width = 12.5
    chart.y_axis.title = "Publicaciones"
    chart.add_data(
        Reference(
            source_ws,
            min_col=min(value_cols),
            max_col=max(value_cols),
            min_row=4,
            max_row=source_ws.max_row,
        ),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(
            source_ws,
            min_col=label_col,
            min_row=5,
            max_row=source_ws.max_row,
        )
    )
    palette = ("2F66E8", "2F855A", "B85F2E", "7554B8")
    for index, series in enumerate(chart.series):
        series.graphicalProperties.line.solidFill = palette[index % len(palette)]
        series.marker.symbol = "circle"
        series.marker.size = 5
    target_ws.add_chart(chart, anchor)


def _add_gestion_donut_chart(source_ws, target_ws, *, label_col, value_col, title, anchor):
    if source_ws.max_row < 5:
        return
    chart = DoughnutChart()
    chart.title = title
    chart.height = 7.2
    chart.width = 10.5
    chart.holeSize = 62
    chart.legend.position = "b"
    chart.add_data(
        Reference(source_ws, min_col=value_col, min_row=4, max_row=source_ws.max_row),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(source_ws, min_col=label_col, min_row=5, max_row=source_ws.max_row)
    )
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showLeaderLines = True
    target_ws.add_chart(chart, anchor)


def _add_gestion_bar_chart(source_ws, target_ws, *, label_col, value_col, title, anchor, limit=10):
    if source_ws.max_row < 5:
        return
    max_row = min(source_ws.max_row, 4 + max(1, int(limit)))
    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "clustered"
    chart.title = title
    chart.height = 8
    chart.width = 13
    chart.legend = None
    chart.x_axis.title = "Publicaciones"
    chart.add_data(
        Reference(source_ws, min_col=value_col, min_row=4, max_row=max_row),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(source_ws, min_col=label_col, min_row=5, max_row=max_row)
    )
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = "2F66E8"
        chart.series[0].graphicalProperties.line.solidFill = "2F66E8"
    target_ws.add_chart(chart, anchor)


def _add_management_workbook_charts(workbook):
    summary_ws = workbook["Resumen"]
    summary_ws.sheet_view.showGridLines = False

    for letter in "EFGHIJKLMNOP":
        summary_ws.column_dimensions[letter].width = 11

    if "Estados" in workbook.sheetnames:
        states_ws = workbook["Estados"]
        _add_gestion_donut_chart(
            states_ws,
            summary_ws,
            label_col=1,
            value_col=2,
            title="Estados de publicaciones",
            anchor="E4",
        )
        _add_gestion_donut_chart(
            states_ws,
            states_ws,
            label_col=1,
            value_col=2,
            title="Estados de publicaciones",
            anchor="D4",
        )

    if "Por año" in workbook.sheetnames:
        annual_ws = workbook["Por año"]
        _add_gestion_line_chart(
            annual_ws,
            summary_ws,
            label_col=1,
            value_cols=(2, 3, 4, 5),
            title="Evolución de la producción",
            anchor="K4",
        )
        _add_gestion_line_chart(
            annual_ws,
            annual_ws,
            label_col=1,
            value_cols=(2, 3, 4, 5),
            title="Evolución de la producción",
            anchor="G4",
        )

    if "Por tipo" in workbook.sheetnames:
        type_ws = workbook["Por tipo"]
        _add_gestion_bar_chart(
            type_ws,
            summary_ws,
            label_col=3,
            value_col=5,
            title="Producción por tipo",
            anchor="E20",
            limit=10,
        )
        _add_gestion_bar_chart(
            type_ws,
            type_ws,
            label_col=3,
            value_col=5,
            title="Producción por tipo",
            anchor="H4",
            limit=10,
        )

    relation_specs = (
        ("Sedes", "Producción por sede"),
        ("Facultades", "Producción por facultad"),
        ("Carreras", "Producción por carrera"),
        ("Proyectos", "Producción por proyecto"),
    )
    for sheet_name, title in relation_specs:
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        _add_gestion_bar_chart(
            worksheet,
            worksheet,
            label_col=2,
            value_col=3,
            title=title,
            anchor="H4",
            limit=10,
        )

    if "Sedes" in workbook.sheetnames:
        _add_gestion_bar_chart(
            workbook["Sedes"],
            summary_ws,
            label_col=2,
            value_col=3,
            title="Sedes con mayor producción",
            anchor="K20",
            limit=8,
        )

def _report_filename():
    timestamp = timezone.localtime().strftime(
        "%Y%m%d_%H%M%S"
    )

    return (
        "reporte_institucional_gestion_"
        f"{timestamp}.xlsx"
    )


# ============================================================
# VISTA PREVIA
# ============================================================

def build_institutional_report_preview(
    params,
):
    payload = (
        build_management_dashboard_payload(
            params
        )
    )

    projects = (
        payload.get(
            "proyectos",
            {}
        )
        or {}
    )

    return {
        "ok": True,
        "filtros_aplicados": (
            payload.get(
                "filtros_aplicados",
                {},
            )
        ),
        "indicadores": (
            payload.get(
                "indicadores",
                {},
            )
        ),
        "alertas": (
            payload.get(
                "alertas",
                {},
            )
        ),
        "proyectos": (
            projects.get(
                "resumen",
                {},
            )
        ),
        "hojas_estimadas": 11,
        "incluye": [
            "Resumen ejecutivo",
            "Estados de publicaciones",
            "Producción anual",
            "Producción por tipo",
            "Sedes",
            "Facultades",
            "Carreras",
            "Proyectos",
            "Proyectos sin producción",
            "Cola de revisión",
            "Actividad reciente",
        ],
    }


# ============================================================
# WORKBOOK
# ============================================================

def build_institutional_report_workbook(
    params,
):
    payload = (
        build_management_dashboard_payload(
            params
        )
    )

    workbook = Workbook()

    # ========================================================
    # RESUMEN
    # ========================================================

    summary_ws = workbook.active
    summary_ws.title = "Resumen"

    _sheet_title(
        summary_ws,
        "SGPC ULEAM — Reporte institucional de gestión científica",
        (
            "Documento de apoyo para seguimiento, control y toma "
            "de decisiones sobre producción científica."
        ),
    )

    filters = (
        payload.get(
            "filtros_aplicados",
            {}
        )
        or {}
    )

    row = 4

    row = _section_header(
        summary_ws,
        row,
        "Filtros aplicados",
        end_column=4,
    )

    filter_rows = [
        [
            "Sede",
            filters.get(
                "sede_id"
            ),
        ],
        [
            "Facultad",
            filters.get(
                "facultad_id"
            ),
        ],
        [
            "Carrera",
            filters.get(
                "carrera_id"
            ),
        ],
        [
            "Proyecto",
            filters.get(
                "proyecto_id"
            ),
        ],
        [
            "Tipo de publicación",
            filters.get(
                "tipo_id"
            ),
        ],
        [
            "Estado publicación",
            filters.get(
                "estado"
            ),
        ],
        [
            "Estado proyecto",
            filters.get(
                "estado_proyecto"
            ),
        ],
        [
            "Año desde",
            filters.get(
                "anio_desde"
            ),
        ],
        [
            "Año hasta",
            filters.get(
                "anio_hasta"
            ),
        ],
    ]

    row = _write_table(
        summary_ws,
        start_row=row,
        headers=[
            "Filtro",
            "Valor",
        ],
        rows=filter_rows,
    )

    row += 1

    indicators = (
        payload.get(
            "indicadores",
            {}
        )
        or {}
    )

    row = _section_header(
        summary_ws,
        row,
        "Indicadores generales",
        end_column=4,
    )

    indicator_rows = [
        [
            "Total publicaciones",
            indicators.get(
                "total_publicaciones",
                0,
            ),
        ],
        [
            "Borrador",
            indicators.get(
                "borrador",
                0,
            ),
        ],
        [
            "En revisión",
            indicators.get(
                "en_revision",
                0,
            ),
        ],
        [
            "Observadas",
            indicators.get(
                "observada",
                0,
            ),
        ],
        [
            "Aprobadas",
            indicators.get(
                "aprobada",
                0,
            ),
        ],
        [
            "Rechazadas",
            indicators.get(
                "rechazada",
                0,
            ),
        ],
        [
            "Pendientes de gestión",
            indicators.get(
                "pendientes_gestion",
                0,
            ),
        ],
        [
            "Cobertura PDF",
            _format_percent(
                indicators.get(
                    "cobertura_pdf",
                    0,
                )
            ),
        ],
        [
            "Vinculación con proyectos",
            _format_percent(
                indicators.get(
                    "vinculacion_proyectos",
                    0,
                )
            ),
        ],
        [
            "Aprobación sobre resueltas",
            _format_percent(
                indicators.get(
                    "tasa_aprobacion_resueltas",
                    0,
                )
            ),
        ],
    ]

    row = _write_table(
        summary_ws,
        start_row=row,
        headers=[
            "Indicador",
            "Valor",
        ],
        rows=indicator_rows,
    )

    row += 1

    alerts = (
        payload.get(
            "alertas",
            {}
        )
        or {}
    )

    row = _section_header(
        summary_ws,
        row,
        "Alertas de gestión",
        end_column=4,
    )

    alert_rows = [
        [
            "Publicaciones en revisión",
            alerts.get(
                "publicaciones_en_revision",
                0,
            ),
        ],
        [
            "Publicaciones observadas",
            alerts.get(
                "publicaciones_observadas",
                0,
            ),
        ],
        [
            "Publicaciones sin PDF",
            alerts.get(
                "publicaciones_sin_pdf",
                0,
            ),
        ],
        [
            "Publicaciones sin proyecto",
            alerts.get(
                "publicaciones_sin_proyecto",
                0,
            ),
        ],
        [
            "Proyectos sin producción",
            alerts.get(
                "proyectos_sin_produccion",
                0,
            ),
        ],
    ]

    alert_start = row

    row = _write_table(
        summary_ws,
        start_row=row,
        headers=[
            "Alerta",
            "Cantidad",
        ],
        rows=alert_rows,
    )

    for excel_row in range(
        alert_start + 1,
        row,
    ):
        summary_ws.cell(
            row=excel_row,
            column=1,
        ).fill = PatternFill(
            "solid",
            fgColor=ALERT_FILL,
        )

        summary_ws.cell(
            row=excel_row,
            column=2,
        ).fill = PatternFill(
            "solid",
            fgColor=ALERT_FILL,
        )

    _autofit(
        summary_ws
    )

    summary_ws.freeze_panes = "A4"

    # ========================================================
    # ESTADOS
    # ========================================================

    states_ws = workbook.create_sheet(
        "Estados"
    )

    _sheet_title(
        states_ws,
        "Estados del ciclo de publicaciones",
        (
            "Distribución operativa del flujo de registro, revisión "
            "y resolución."
        ),
    )

    state_rows = [
        [
            "Borrador",
            indicators.get(
                "borrador",
                0,
            ),
        ],
        [
            "En revisión",
            indicators.get(
                "en_revision",
                0,
            ),
        ],
        [
            "Observada",
            indicators.get(
                "observada",
                0,
            ),
        ],
        [
            "Aprobada",
            indicators.get(
                "aprobada",
                0,
            ),
        ],
        [
            "Rechazada",
            indicators.get(
                "rechazada",
                0,
            ),
        ],
    ]

    _write_table(
        states_ws,
        start_row=4,
        headers=[
            "Estado",
            "Total",
        ],
        rows=state_rows,
    )

    _autofit(
        states_ws
    )

    # ========================================================
    # DISTRIBUCIONES
    # ========================================================

    distributions = (
        payload.get(
            "distribuciones",
            {}
        )
        or {}
    )

    annual_ws = workbook.create_sheet(
        "Por año"
    )

    _sheet_title(
        annual_ws,
        "Producción científica por año",
    )

    annual_rows = [
        [
            item.get(
                "anio"
            ),
            item.get(
                "total",
                0,
            ),
            item.get(
                "aprobadas",
                0,
            ),
            item.get(
                "observadas",
                0,
            ),
            item.get(
                "rechazadas",
                0,
            ),
        ]
        for item in (
            distributions.get(
                "por_anio",
                []
            )
            or []
        )
    ]

    end_row = _write_table(
        annual_ws,
        start_row=4,
        headers=[
            "Año",
            "Total",
            "Aprobadas",
            "Observadas",
            "Rechazadas",
        ],
        rows=annual_rows,
    )

    _freeze_and_filter(
        annual_ws,
        freeze_cell="A5",
        auto_filter_range=(
            f"A4:E{max(4, end_row - 1)}"
        ),
    )

    _autofit(
        annual_ws
    )

    type_ws = workbook.create_sheet(
        "Por tipo"
    )

    _sheet_title(
        type_ws,
        "Producción científica por tipo",
    )

    type_rows = [
        [
            item.get(
                "tipo_id"
            ),
            item.get(
                "codigo"
            ),
            item.get(
                "nombre"
            ),
            item.get(
                "categoria"
            ),
            item.get(
                "total",
                0,
            ),
            item.get(
                "aprobadas",
                0,
            ),
        ]
        for item in (
            distributions.get(
                "por_tipo",
                []
            )
            or []
        )
    ]

    end_row = _write_table(
        type_ws,
        start_row=4,
        headers=[
            "ID",
            "Código",
            "Tipo",
            "Categoría",
            "Total",
            "Aprobadas",
        ],
        rows=type_rows,
    )

    _freeze_and_filter(
        type_ws,
        freeze_cell="A5",
        auto_filter_range=(
            f"A4:F{max(4, end_row - 1)}"
        ),
    )

    _autofit(
        type_ws
    )

    def build_relation_sheet(
        *,
        sheet_name,
        title,
        items,
        id_key,
        label_key,
        id_header,
        label_header,
    ):
        worksheet = workbook.create_sheet(
            sheet_name
        )

        _sheet_title(
            worksheet,
            title,
        )

        rows = [
            [
                item.get(
                    id_key
                ),
                item.get(
                    label_key
                ),
                item.get(
                    "total",
                    0,
                ),
                item.get(
                    "aprobadas",
                    0,
                ),
                item.get(
                    "en_revision",
                    0,
                ),
                item.get(
                    "observadas",
                    0,
                ),
            ]
            for item in (
                items
                or []
            )
        ]

        end_row = _write_table(
            worksheet,
            start_row=4,
            headers=[
                id_header,
                label_header,
                "Total",
                "Aprobadas",
                "En revisión",
                "Observadas",
            ],
            rows=rows,
        )

        _freeze_and_filter(
            worksheet,
            freeze_cell="A5",
            auto_filter_range=(
                f"A4:F{max(4, end_row - 1)}"
            ),
        )

        _autofit(
            worksheet
        )

        return worksheet

    build_relation_sheet(
        sheet_name="Sedes",
        title="Producción por Sede",
        items=distributions.get(
            "por_sede",
            []
        ),
        id_key="sede_id",
        label_key="sede",
        id_header="ID Sede",
        label_header="Sede",
    )

    build_relation_sheet(
        sheet_name="Facultades",
        title="Producción por Facultad",
        items=distributions.get(
            "por_facultad",
            []
        ),
        id_key="facultad_id",
        label_key="facultad",
        id_header="ID Facultad",
        label_header="Facultad",
    )

    build_relation_sheet(
        sheet_name="Carreras",
        title="Producción por Carrera",
        items=distributions.get(
            "por_carrera",
            []
        ),
        id_key="carrera_id",
        label_key="carrera",
        id_header="ID Carrera",
        label_header="Carrera",
    )

    build_relation_sheet(
        sheet_name="Proyectos",
        title="Producción por Proyecto",
        items=distributions.get(
            "por_proyecto",
            []
        ),
        id_key="proyecto_id",
        label_key="proyecto",
        id_header="ID Proyecto",
        label_header="Proyecto",
    )

    # ========================================================
    # PROYECTOS SIN PRODUCCIÓN
    # ========================================================

    projects = (
        payload.get(
            "proyectos",
            {}
        )
        or {}
    )

    no_output_ws = workbook.create_sheet(
        "Sin producción"
    )

    _sheet_title(
        no_output_ws,
        "Proyectos sin producción científica registrada",
        (
            "La ausencia de producción no constituye por sí sola "
            "una anomalía; funciona como indicador para seguimiento."
        ),
    )

    no_output_rows = []

    for item in (
        projects.get(
            "proyectos_sin_produccion",
            []
        )
        or []
    ):
        site = _relation_name(
            item.get(
                "sede"
            )
        )

        faculty = _relation_name(
            item.get(
                "facultad"
            )
        )

        career = _relation_name(
            item.get(
                "carrera"
            )
        )

        no_output_rows.append(
            [
                item.get(
                    "proyecto_id"
                ),
                item.get(
                    "nombre"
                ),
                item.get(
                    "estado_label"
                ),
                site,
                faculty,
                career,
                item.get(
                    "total_publicaciones",
                    0,
                ),
            ]
        )

    end_row = _write_table(
        no_output_ws,
        start_row=4,
        headers=[
            "ID",
            "Proyecto",
            "Estado",
            "Sede",
            "Facultad",
            "Carrera",
            "Publicaciones",
        ],
        rows=no_output_rows,
    )

    _freeze_and_filter(
        no_output_ws,
        freeze_cell="A5",
        auto_filter_range=(
            f"A4:G{max(4, end_row - 1)}"
        ),
    )

    _autofit(
        no_output_ws
    )

    # ========================================================
    # COLA DE REVISIÓN
    # ========================================================

    queue_ws = workbook.create_sheet(
        "Cola revisión"
    )

    _sheet_title(
        queue_ws,
        "Publicaciones pendientes de revisión",
    )

    queue_rows = [
        [
            item.get(
                "publicacion_id"
            ),
            item.get(
                "numero"
            ),
            item.get(
                "tipo"
            ),
            item.get(
                "anio_publicacion"
            ),
            item.get(
                "sede"
            ),
            item.get(
                "carrera"
            ),
            item.get(
                "usuario_creador_email"
            ),
            item.get(
                "updated_at"
            ),
        ]
        for item in (
            payload.get(
                "cola_revision",
                []
            )
            or []
        )
    ]

    end_row = _write_table(
        queue_ws,
        start_row=4,
        headers=[
            "ID publicación",
            "Número",
            "Tipo",
            "Año",
            "Sede",
            "Carrera",
            "Usuario creador",
            "Última actualización",
        ],
        rows=queue_rows,
    )

    _freeze_and_filter(
        queue_ws,
        freeze_cell="A5",
        auto_filter_range=(
            f"A4:H{max(4, end_row - 1)}"
        ),
    )

    _autofit(
        queue_ws
    )

    # ========================================================
    # ACTIVIDAD RECIENTE
    # ========================================================

    activity_ws = workbook.create_sheet(
        "Actividad"
    )

    _sheet_title(
        activity_ws,
        "Actividad reciente de auditoría",
    )

    activity_rows = [
        [
            item.get(
                "id"
            ),
            item.get(
                "publicacion_id"
            ),
            item.get(
                "evento_label"
            ),
            item.get(
                "actor_nombre"
            ),
            item.get(
                "actor_email"
            ),
            item.get(
                "estado_anterior"
            ),
            item.get(
                "estado_resultante"
            ),
            item.get(
                "created_at"
            ),
        ]
        for item in (
            payload.get(
                "actividad_reciente",
                []
            )
            or []
        )
    ]

    end_row = _write_table(
        activity_ws,
        start_row=4,
        headers=[
            "ID evento",
            "ID publicación",
            "Evento",
            "Actor",
            "Correo",
            "Estado anterior",
            "Estado resultante",
            "Fecha",
        ],
        rows=activity_rows,
    )

    _freeze_and_filter(
        activity_ws,
        freeze_cell="A5",
        auto_filter_range=(
            f"A4:H{max(4, end_row - 1)}"
        ),
    )

    _autofit(
        activity_ws
    )

    _add_management_workbook_charts(workbook)
    workbook.active = 0

    return workbook


def build_institutional_report_file(
    params,
):
    workbook = (
        build_institutional_report_workbook(
            params
        )
    )

    output = BytesIO()

    workbook.save(
        output
    )

    return (
        output.getvalue(),
        _report_filename(),
    )