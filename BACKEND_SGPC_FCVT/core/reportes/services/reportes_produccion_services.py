"""Reportes de producción científica aprobada del SGPC ULEAM.

Este módulo implementa dos contratos que comparten exactamente la misma
lógica temporal y de agregación:

* reporte institucional (solo administradores);
* reporte personal del docente autenticado.

La producción oficial contabiliza exclusivamente publicaciones en estado
``aprobada``. Los estados del flujo de revisión pertenecen al dashboard de
gestión y no se mezclan con estos indicadores.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO

from django.db.models import Count, Max, Min, Q
from django.utils import timezone
from rest_framework.exceptions import ValidationError

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.models import Publicacion


# ============================================================
# CONSTANTES
# ============================================================

MONTHS = (
    (1, "Enero", "Ene"),
    (2, "Febrero", "Feb"),
    (3, "Marzo", "Mar"),
    (4, "Abril", "Abr"),
    (5, "Mayo", "May"),
    (6, "Junio", "Jun"),
    (7, "Julio", "Jul"),
    (8, "Agosto", "Ago"),
    (9, "Septiembre", "Sep"),
    (10, "Octubre", "Oct"),
    (11, "Noviembre", "Nov"),
    (12, "Diciembre", "Dic"),
)

MONTH_NAME = {number: label for number, label, _short in MONTHS}
MONTH_SHORT = {number: short for number, _label, short in MONTHS}

PERIOD_MODES = {
    "historico",
    "personalizado",
    "trimestral",
    "semestral",
    "anual",
    "ultimos_12_meses",
}

DEFAULT_DETAIL_LIMIT = 100
MAX_DETAIL_LIMIT = 500
MAX_MONTHS_FOR_CHART = 18

CANONICAL_TYPE_LABELS = {
    "AAI": "Artículo de alto impacto",
    "AR": "Artículo regional",
    "PON": "Ponencia",
    "CAP": "Capítulo de libro",
    "LIB": "Libro",
}

CANONICAL_TYPE_ALIASES = {
    "AAI": "AAI",
    "ALTO_IMPACTO": "AAI",
    "ARTICULO_ALTO_IMPACTO": "AAI",
    "AR": "AR",
    "REGIONAL": "AR",
    "ARTICULO_REGIONAL": "AR",
    "PON": "PON",
    "PONENCIA": "PON",
    "CAP": "CAP",
    "CAPITULO": "CAP",
    "CAPITULO_LIBRO": "CAP",
    "LIB": "LIB",
    "LIBRO": "LIB",
}

TITLE_FILL = "1F4FD7"
TITLE_TEXT = "FFFFFF"
SECTION_FILL = "F5F7FB"
SECTION_TEXT = "172033"
HEADER_FILL = "F5F7FB"
HEADER_TEXT = "172033"
ALT_FILL = "F5F7FA"
BORDER_COLOR = "D9E0E8"

THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)


# ============================================================
# VALIDACIÓN Y PERÍODO
# ============================================================


def _positive_int(value, field, *, maximum=None):
    if value in (None, ""):
        return None

    try:
        parsed = int(str(value).strip())
    except (TypeError, ValueError, OverflowError) as exc:
        raise ValidationError(
            {field: ["Debe indicar un identificador numérico válido."]}
        ) from exc

    if parsed < 1:
        raise ValidationError({field: ["El valor debe ser mayor o igual a 1."]})

    if maximum is not None and parsed > maximum:
        raise ValidationError({field: [f"El valor no puede ser mayor que {maximum}."]})

    return parsed


def _year(value, field="anio"):
    if value in (None, ""):
        return None

    try:
        parsed = int(str(value).strip())
    except (TypeError, ValueError, OverflowError) as exc:
        raise ValidationError({field: ["El año debe ser numérico."]}) from exc

    if not 1900 <= parsed <= 2100:
        raise ValidationError({field: ["El año debe estar entre 1900 y 2100."]})

    return parsed


def _parse_month(value, field):
    text = str(value or "").strip()

    if not text:
        return None

    parts = text.split("-")

    if len(parts) != 2:
        raise ValidationError({field: ["Utilice el formato YYYY-MM."]})

    try:
        year = int(parts[0])
        month = int(parts[1])
    except (TypeError, ValueError) as exc:
        raise ValidationError({field: ["Utilice el formato YYYY-MM."]}) from exc

    if not 1900 <= year <= 2100 or not 1 <= month <= 12:
        raise ValidationError({field: ["El mes indicado no es válido."]})

    return year, month


def _month_key(year, month):
    return f"{int(year):04d}-{int(month):02d}"


def _month_ordinal(year, month):
    return int(year) * 12 + (int(month) - 1)


def _ordinal_to_month(ordinal):
    year = ordinal // 12
    month = (ordinal % 12) + 1
    return year, month


def _month_label(value):
    parsed = _parse_month(value, "mes") if value else None

    if not parsed:
        return ""

    year, month = parsed
    return f"{MONTH_NAME[month]} {year}"


@dataclass(frozen=True)
class PeriodSelection:
    mode: str
    month_from: str | None = None
    month_to: str | None = None
    year: int | None = None
    quarter: int | None = None
    semester: int | None = None
    label: str = "Histórico"
    include_unknown_months: bool = True

    def as_dict(self):
        return {
            "periodo_modo": self.mode,
            "mes_desde": self.month_from,
            "mes_hasta": self.month_to,
            "anio": self.year,
            "trimestre": self.quarter,
            "semestre": self.semester,
            "label": self.label,
        }


def resolve_period(params, *, today=None):
    """Resuelve el período solicitado y devuelve el contrato canónico.

    Se aceptan períodos personalizados, trimestrales, semestrales, anuales,
    últimos 12 meses e histórico. El frontend no necesita calcular límites
    de trimestre/semestre: el backend es la fuente de verdad.
    """

    today = today or timezone.localdate()

    raw_mode = str(params.get("periodo_modo") or "").strip().lower()
    raw_from = params.get("mes_desde") or params.get("fecha_desde")
    raw_to = params.get("mes_hasta") or params.get("fecha_hasta")

    if not raw_mode:
        raw_mode = "personalizado" if (raw_from or raw_to) else "historico"

    if raw_mode not in PERIOD_MODES:
        raise ValidationError(
            {
                "periodo_modo": [
                    "Modo de período inválido. Use histórico, personalizado, "
                    "trimestral, semestral, anual o últimos 12 meses."
                ]
            }
        )

    if raw_mode == "historico":
        return PeriodSelection(mode="historico")

    if raw_mode == "ultimos_12_meses":
        end_ordinal = _month_ordinal(today.year, today.month)
        start_year, start_month = _ordinal_to_month(end_ordinal - 11)
        start = _month_key(start_year, start_month)
        end = _month_key(today.year, today.month)

        return PeriodSelection(
            mode=raw_mode,
            month_from=start,
            month_to=end,
            label=f"{_month_label(start)} — {_month_label(end)}",
            include_unknown_months=False,
        )

    selected_year = _year(params.get("anio"), "anio")

    if raw_mode == "anual":
        selected_year = selected_year or today.year
        return PeriodSelection(
            mode=raw_mode,
            month_from=_month_key(selected_year, 1),
            month_to=_month_key(selected_year, 12),
            year=selected_year,
            label=f"Año {selected_year}",
            # Una publicación con año conocido y mes desconocido sí forma parte
            # del reporte anual de ese año.
            include_unknown_months=True,
        )

    if raw_mode == "trimestral":
        selected_year = selected_year or today.year
        quarter = _positive_int(params.get("trimestre"), "trimestre", maximum=4)
        quarter = quarter or (((today.month - 1) // 3) + 1)
        start_month = ((quarter - 1) * 3) + 1
        end_month = start_month + 2
        start = _month_key(selected_year, start_month)
        end = _month_key(selected_year, end_month)

        return PeriodSelection(
            mode=raw_mode,
            month_from=start,
            month_to=end,
            year=selected_year,
            quarter=quarter,
            label=f"{quarter}.º trimestre {selected_year}",
            include_unknown_months=False,
        )

    if raw_mode == "semestral":
        selected_year = selected_year or today.year
        semester = _positive_int(params.get("semestre"), "semestre", maximum=2)
        semester = semester or (1 if today.month <= 6 else 2)
        start_month = 1 if semester == 1 else 7
        end_month = 6 if semester == 1 else 12
        start = _month_key(selected_year, start_month)
        end = _month_key(selected_year, end_month)

        return PeriodSelection(
            mode=raw_mode,
            month_from=start,
            month_to=end,
            year=selected_year,
            semester=semester,
            label=f"{semester}.º semestre {selected_year}",
            include_unknown_months=False,
        )

    # Personalizado
    parsed_from = _parse_month(raw_from, "mes_desde") if raw_from else None
    parsed_to = _parse_month(raw_to, "mes_hasta") if raw_to else None

    month_from = _month_key(*parsed_from) if parsed_from else None
    month_to = _month_key(*parsed_to) if parsed_to else None

    if month_from and month_to and month_from > month_to:
        raise ValidationError(
            {"mes_hasta": ["El mes final no puede ser anterior al mes inicial."]}
        )

    if not month_from and not month_to:
        return PeriodSelection(mode="historico")

    if month_from and month_to:
        label = f"{_month_label(month_from)} — {_month_label(month_to)}"
    elif month_from:
        label = f"Desde {_month_label(month_from)}"
    else:
        label = f"Hasta {_month_label(month_to)}"

    return PeriodSelection(
        mode="personalizado",
        month_from=month_from,
        month_to=month_to,
        label=label,
        include_unknown_months=False,
    )


def _period_q(period):
    """Construye Q para un período basado en año/mes separados."""

    if period.mode == "historico":
        return Q()

    if period.mode == "anual" and period.year:
        return Q(anio_publicacion=period.year)

    query = Q()

    if period.month_from:
        start_year, start_month = _parse_month(period.month_from, "mes_desde")
        query &= (
            Q(anio_publicacion__gt=start_year)
            | Q(
                anio_publicacion=start_year,
                mes_publicacion__gte=start_month,
            )
        )

    if period.month_to:
        end_year, end_month = _parse_month(period.month_to, "mes_hasta")
        query &= (
            Q(anio_publicacion__lt=end_year)
            | Q(
                anio_publicacion=end_year,
                mes_publicacion__lte=end_month,
            )
        )

    if not period.include_unknown_months:
        query &= Q(mes_publicacion__isnull=False)

    return query


# ============================================================
# QUERYSETS Y FILTROS
# ============================================================


def _approved_publications():
    return (
        Publicacion.objects
        .filter(estado=Publicacion.ESTADO_APROBADA)
        .select_related(
            "tipo",
            "sede",
            "carrera",
            "carrera__facultad",
            "proyecto",
            "usuario_creador",
            "articulo",
            "ponencia",
            "libro",
            "capitulo_libro",
        )
        .prefetch_related(
            "participaciones__autor",
        )
    )


def _canonical_type_code(value):
    if value in (None, ""):
        return None

    normalized = (
        str(value)
        .strip()
        .upper()
        .replace("Á", "A")
        .replace("É", "E")
        .replace("Í", "I")
        .replace("Ó", "O")
        .replace("Ú", "U")
        .replace("Ü", "U")
        .replace("Ñ", "N")
        .replace("-", "_")
        .replace(" ", "_")
    )

    return CANONICAL_TYPE_ALIASES.get(normalized)


def _apply_canonical_type_filter(qs, code):
    code = _canonical_type_code(code)

    if not code:
        return qs

    if code == "AAI":
        return qs.filter(
            Q(tipo__codigo__iexact="AAI")
            | Q(
                tipo__categoria__iexact="articulo",
                articulo__tipo_articulo__iexact="alto_impacto",
            )
        )

    if code == "AR":
        return qs.filter(
            Q(tipo__codigo__iexact="AR")
            | Q(
                tipo__categoria__iexact="articulo",
                articulo__tipo_articulo__iexact="regional",
            )
        )

    if code == "PON":
        return qs.filter(
            Q(tipo__codigo__iexact="PON")
            | Q(tipo__categoria__iexact="ponencia")
        )

    if code == "CAP":
        return qs.filter(
            Q(tipo__codigo__iexact="CAP")
            | Q(tipo__categoria__iexact="capitulo")
        )

    if code == "LIB":
        return qs.filter(
            Q(tipo__codigo__iexact="LIB")
            | Q(tipo__categoria__iexact="libro")
        )

    return qs


def _parse_dimension_filters(params, *, allow_teacher):
    detail_limit = _positive_int(params.get("detalle_limite"), "detalle_limite")
    detail_limit = min(detail_limit or DEFAULT_DETAIL_LIMIT, MAX_DETAIL_LIMIT)

    raw_type_code = params.get("tipo_codigo")
    type_code = _canonical_type_code(raw_type_code)

    if raw_type_code not in (None, "") and not type_code:
        raise ValidationError(
            {
                "tipo_codigo": [
                    "Tipo de publicación no válido para el reporte institucional."
                ]
            }
        )

    type_id = None
    if not type_code:
        type_id = _positive_int(
            params.get("tipo") or params.get("tipo_id"),
            "tipo",
        )

    filters = {
        "sede_id": _positive_int(params.get("sede") or params.get("sede_id"), "sede"),
        "facultad_id": _positive_int(
            params.get("facultad") or params.get("facultad_id"),
            "facultad",
        ),
        "carrera_id": _positive_int(
            params.get("carrera") or params.get("carrera_id"),
            "carrera",
        ),
        "tipo_id": type_id,
        "tipo_codigo": type_code,
        "proyecto_id": _positive_int(
            params.get("proyecto") or params.get("proyecto_id"),
            "proyecto",
        ),
        "docente_id": None,
        "detalle_limite": detail_limit,
    }

    if allow_teacher:
        filters["docente_id"] = _positive_int(
            params.get("docente")
            or params.get("docente_id")
            or params.get("autor")
            or params.get("autor_id"),
            "docente",
        )

    return filters


def _apply_dimensions(qs, filters, *, include_teacher=True):
    if filters.get("sede_id"):
        qs = qs.filter(sede_id=filters["sede_id"])

    if filters.get("facultad_id"):
        qs = qs.filter(carrera__facultad_id=filters["facultad_id"])

    if filters.get("carrera_id"):
        qs = qs.filter(carrera_id=filters["carrera_id"])

    if filters.get("tipo_codigo"):
        qs = _apply_canonical_type_filter(
            qs,
            filters["tipo_codigo"],
        )
    elif filters.get("tipo_id"):
        qs = qs.filter(tipo_id=filters["tipo_id"])

    if filters.get("proyecto_id"):
        qs = qs.filter(proyecto_id=filters["proyecto_id"])

    if include_teacher and filters.get("docente_id"):
        qs = qs.filter(
            participaciones__autor_id=filters["docente_id"],
            participaciones__autor__es_externo=False,
        )

    return qs.distinct()


def _teacher_publications(user):
    """Devuelve producción atribuible al usuario autenticado.

    La autoría vinculada es la fuente principal. ``usuario_creador`` se usa
    como respaldo para registros históricos que todavía no tengan enlazado
    el modelo Autor; nunca se acepta un ID de usuario suministrado por Vue.
    """

    return (
        _approved_publications()
        .filter(
            Q(participaciones__autor__usuario=user)
            | Q(usuario_creador=user)
        )
        .distinct()
    )


def _apply_period(qs, period):
    return qs.filter(_period_q(period)).distinct()


# ============================================================
# AGREGACIONES
# ============================================================


def _summary(qs, *, teacher_id=None):
    values = qs.aggregate(
        total=Count("id", distinct=True),
        con_pdf=Count(
            "id",
            filter=Q(archivo_pdf__isnull=False) & ~Q(archivo_pdf=""),
            distinct=True,
        ),
        con_proyecto=Count(
            "id",
            filter=Q(proyecto__isnull=False),
            distinct=True,
        ),
        sin_mes=Count(
            "id",
            filter=Q(mes_publicacion__isnull=True),
            distinct=True,
        ),
        total_docentes=Count(
            "participaciones__autor",
            filter=Q(participaciones__autor__es_externo=False),
            distinct=True,
        ),
        total_proyectos=Count("proyecto", distinct=True),
        total_tipos=Count("tipo", distinct=True),
    )

    total = int(values.get("total") or 0)
    with_pdf = int(values.get("con_pdf") or 0)
    with_project = int(values.get("con_proyecto") or 0)
    total_teachers = int(values.get("total_docentes") or 0)

    # Si el administrador filtra por un docente concreto, el indicador debe
    # representar ese docente y no todos sus colaboradores internos.
    if teacher_id:
        total_teachers = 1 if total else 0

    return {
        "total_publicaciones": total,
        "con_pdf": with_pdf,
        "sin_pdf": max(total - with_pdf, 0),
        "cobertura_pdf": round((with_pdf / total) * 100, 2) if total else 0.0,
        "con_proyecto": with_project,
        "sin_proyecto": max(total - with_project, 0),
        "vinculacion_proyectos": (
            round((with_project / total) * 100, 2) if total else 0.0
        ),
        "sin_mes": int(values.get("sin_mes") or 0),
        "total_docentes": total_teachers,
        "total_proyectos": int(values.get("total_proyectos") or 0),
        "total_tipos": int(values.get("total_tipos") or 0),
    }


def _group(qs, *, fields, mapping, order_label):
    values = qs.values(*fields).annotate(total=Count("id", distinct=True))
    values = values.order_by("-total", order_label)

    output = []

    for row in values:
        item = {target: row.get(source) for target, source in mapping.items()}
        item["total"] = int(row.get("total") or 0)
        output.append(item)

    return output


def _by_year(qs):
    """Serie anual cronológica, completando años sin producción con cero."""

    rows = (
        qs.exclude(anio_publicacion__isnull=True)
        .values("anio_publicacion")
        .annotate(total=Count("id", distinct=True))
        .order_by("anio_publicacion")
    )

    counts = {
        int(row["anio_publicacion"]): int(row.get("total") or 0)
        for row in rows
        if row.get("anio_publicacion") is not None
    }

    if not counts:
        return []

    first_year = min(counts)
    last_year = max(counts)

    return [
        {
            "anio": year,
            "total": counts.get(year, 0),
        }
        for year in range(first_year, last_year + 1)
    ]


def _by_month(qs):
    rows = (
        qs.exclude(mes_publicacion__isnull=True)
        .values("anio_publicacion", "mes_publicacion")
        .annotate(total=Count("id", distinct=True))
        .order_by("anio_publicacion", "mes_publicacion")
    )

    return [
        {
            "anio": row["anio_publicacion"],
            "mes": row["mes_publicacion"],
            "mes_label": MONTH_NAME.get(row["mes_publicacion"], str(row["mes_publicacion"])),
            "periodo": _month_key(row["anio_publicacion"], row["mes_publicacion"]),
            "total": int(row["total"] or 0),
        }
        for row in rows
    ]


def _month_ordinal_from_key(value):
    parsed = _parse_month(value, "mes") if value else None
    if not parsed:
        return None
    return _month_ordinal(*parsed)


def _dense_month_series(items, *, start=None, end=None):
    """Completa meses intermedios con cero para evitar líneas engañosas."""

    rows = list(items or [])
    counts = {}

    for item in rows:
        year = item.get("anio")
        month = item.get("mes")
        if year in (None, "") or month in (None, ""):
            continue
        key = _month_key(int(year), int(month))
        counts[key] = int(item.get("total") or 0)

    if not counts and not (start and end):
        return []

    start_ordinal = _month_ordinal_from_key(start)
    end_ordinal = _month_ordinal_from_key(end)

    if start_ordinal is None:
        start_ordinal = min(_month_ordinal_from_key(key) for key in counts)
    if end_ordinal is None:
        end_ordinal = max(_month_ordinal_from_key(key) for key in counts)

    if start_ordinal is None or end_ordinal is None:
        return rows

    if start_ordinal > end_ordinal:
        start_ordinal, end_ordinal = end_ordinal, start_ordinal

    output = []
    for ordinal in range(start_ordinal, end_ordinal + 1):
        year, month = _ordinal_to_month(ordinal)
        key = _month_key(year, month)
        output.append(
            {
                "anio": year,
                "mes": month,
                "mes_label": MONTH_NAME.get(month, str(month)),
                "periodo": key,
                "total": counts.get(key, 0),
            }
        )

    return output


def _with_month_chart_labels(items):
    """Añade etiquetas compactas y legibles para los gráficos mensuales."""

    rows = [dict(item) for item in (items or [])]
    years = {
        int(item.get("anio"))
        for item in rows
        if item.get("anio") not in (None, "")
    }
    single_year = len(years) == 1

    for item in rows:
        year = item.get("anio")
        month = item.get("mes")
        short = MONTH_SHORT.get(month, str(month or ""))

        if single_year:
            item["chart_label"] = short
        elif year not in (None, ""):
            item["chart_label"] = f"{short} {str(year)[-2:]}"
        else:
            item["chart_label"] = item.get("periodo") or short

    return rows


def _month_visual_items(payload):
    """Serie mensual compacta, continua y siempre apta para gráfico.

    Reglas visuales:
    - histórico: muestra enero-diciembre del último año con publicaciones;
    - períodos de hasta MAX_MONTHS_FOR_CHART: muestra todo el rango;
    - períodos más amplios: muestra los últimos 12 meses del rango.

    De esta forma nunca se unen meses separados por varios años y el gráfico
    mensual continúa disponible incluso cuando el reporte es histórico.
    """

    filters = payload.get("filtros_aplicados", {}) or {}
    items = (payload.get("distribuciones", {}) or {}).get("por_mes", []) or []
    mode = str(filters.get("periodo_modo") or "historico").strip().lower()

    if not items:
        return []

    if mode == "historico":
        years = sorted(
            {
                int(item.get("anio"))
                for item in items
                if item.get("anio") not in (None, "")
            }
        )
        if not years:
            return []

        year = years[-1]
        return _with_month_chart_labels(
            _dense_month_series(
                items,
                start=_month_key(year, 1),
                end=_month_key(year, 12),
            )
        )

    start = filters.get("mes_desde")
    end = filters.get("mes_hasta")

    if mode == "anual" and filters.get("anio"):
        year = int(filters["anio"])
        start = _month_key(year, 1)
        end = _month_key(year, 12)

    series = _dense_month_series(items, start=start, end=end)
    if len(series) <= MAX_MONTHS_FOR_CHART:
        return _with_month_chart_labels(series)

    # Si el rango es muy amplio, conservar una lectura mensual útil sin
    # ocultar el gráfico: se muestran los últimos 12 meses del período.
    return _with_month_chart_labels(series[-12:])


def _by_site(qs):
    return _group(
        qs,
        fields=("sede_id", "sede__nombre"),
        mapping={"sede_id": "sede_id", "sede": "sede__nombre"},
        order_label="sede__nombre",
    )


def _by_faculty(qs):
    return _group(
        qs,
        fields=("carrera__facultad_id", "carrera__facultad__nombre"),
        mapping={
            "facultad_id": "carrera__facultad_id",
            "facultad": "carrera__facultad__nombre",
        },
        order_label="carrera__facultad__nombre",
    )


def _by_career(qs):
    return _group(
        qs,
        fields=("carrera_id", "carrera__nombre"),
        mapping={"carrera_id": "carrera_id", "carrera": "carrera__nombre"},
        order_label="carrera__nombre",
    )


def _by_type(qs):
    return _group(
        qs,
        fields=("tipo_id", "tipo__codigo", "tipo__nombre", "tipo__categoria"),
        mapping={
            "tipo_id": "tipo_id",
            "codigo": "tipo__codigo",
            "tipo": "tipo__nombre",
            "categoria": "tipo__categoria",
        },
        order_label="tipo__nombre",
    )


def _by_project(qs):
    rows = _group(
        qs.filter(proyecto__isnull=False),
        fields=("proyecto_id", "proyecto__nombre"),
        mapping={"proyecto_id": "proyecto_id", "proyecto": "proyecto__nombre"},
        order_label="proyecto__nombre",
    )

    return rows


def _by_teacher(qs, *, teacher_id=None):
    teacher_filters = {
        "participaciones__autor__isnull": False,
        "participaciones__autor__es_externo": False,
    }

    if teacher_id:
        teacher_filters["participaciones__autor_id"] = teacher_id

    rows = (
        qs.filter(**teacher_filters)
        .values(
            "participaciones__autor_id",
            "participaciones__autor__nombres",
            "participaciones__autor__apellidos",
            "participaciones__autor__correo",
            "participaciones__autor__es_externo",
        )
        .annotate(total=Count("id", distinct=True))
        .order_by(
            "-total",
            "participaciones__autor__apellidos",
            "participaciones__autor__nombres",
            "participaciones__autor_id",
        )
    )

    output = []

    for row in rows:
        names = str(row.get("participaciones__autor__nombres") or "").strip()
        surnames = str(row.get("participaciones__autor__apellidos") or "").strip()
        output.append(
            {
                "docente_id": row.get("participaciones__autor_id"),
                "docente": f"{names} {surnames}".strip() or "Autor sin nombre",
                "correo": row.get("participaciones__autor__correo"),
                "es_externo": bool(row.get("participaciones__autor__es_externo")),
                "total": int(row.get("total") or 0),
            }
        )

    return output


# ============================================================
# CATÁLOGOS DEPENDIENTES Y DISPONIBILIDAD TEMPORAL
# ============================================================


def _selector_rows(qs, id_field, label_field, *, extra_fields=()):
    fields = (id_field, label_field, *extra_fields)
    qs = qs.exclude(**{f"{id_field}__isnull": True})
    rows = (
        qs.values(*fields)
        .annotate(total=Count("id", distinct=True))
        .order_by(label_field, id_field)
    )

    output = []

    for row in rows:
        item = {
            "id": row.get(id_field),
            "label": row.get(label_field) or "Sin clasificar",
            "total": int(row.get("total") or 0),
        }

        for field in extra_fields:
            item[field.split("__")[-1]] = row.get(field)

        output.append(item)

    return output


def _period_catalog(qs):
    aggregate = qs.aggregate(
        min_year=Min("anio_publicacion"),
        max_year=Max("anio_publicacion"),
    )

    current_year = timezone.localdate().year
    data_min = aggregate.get("min_year")
    data_max = aggregate.get("max_year")

    year_min = int(data_min or current_year)
    year_max = max(int(data_max or current_year), current_year)

    years = [
        {"value": year, "label": str(year)}
        for year in range(year_max, year_min - 1, -1)
    ]

    month_rows = (
        qs.exclude(mes_publicacion__isnull=True)
        .values("anio_publicacion", "mes_publicacion")
        .annotate(total=Count("id", distinct=True))
        .order_by("anio_publicacion", "mes_publicacion")
    )

    month_counts = {}
    last_data = None

    for row in month_rows:
        year = int(row["anio_publicacion"])
        month = int(row["mes_publicacion"])
        month_counts.setdefault(str(year), {})[str(month)] = int(row["total"] or 0)
        current = _month_key(year, month)
        if last_data is None or current > last_data:
            last_data = current

    current_month = _month_key(timezone.localdate().year, timezone.localdate().month)

    return {
        "anio_min": year_min,
        "anio_max": year_max,
        "mes_min": _month_key(year_min, 1),
        "mes_max": _month_key(year_max, 12),
        "mes_actual": current_month,
        "ultimo_mes_con_datos": last_data,
        "anios": years,
        "meses": [
            {"value": number, "label": label, "short_label": short}
            for number, label, short in MONTHS
        ],
        "meses_con_datos": month_counts,
        "modos": [
            {"value": "historico", "label": "Histórico"},
            {"value": "personalizado", "label": "Personalizado"},
            {"value": "trimestral", "label": "Trimestral"},
            {"value": "semestral", "label": "Semestral"},
            {"value": "anual", "label": "Anual"},
            {"value": "ultimos_12_meses", "label": "Últimos 12 meses"},
        ],
    }


def _institutional_catalogs(filters):
    base = _approved_publications()

    sites = _selector_rows(base, "sede_id", "sede__nombre")

    faculty_qs = base
    if filters.get("sede_id"):
        faculty_qs = faculty_qs.filter(sede_id=filters["sede_id"])

    faculties = _selector_rows(
        faculty_qs,
        "carrera__facultad_id",
        "carrera__facultad__nombre",
    )

    career_qs = faculty_qs
    if filters.get("facultad_id"):
        career_qs = career_qs.filter(carrera__facultad_id=filters["facultad_id"])

    careers = _selector_rows(career_qs, "carrera_id", "carrera__nombre")

    dependent_qs = career_qs
    if filters.get("carrera_id"):
        dependent_qs = dependent_qs.filter(carrera_id=filters["carrera_id"])

    types = _selector_rows(dependent_qs, "tipo_id", "tipo__nombre")
    projects = _selector_rows(
        dependent_qs.filter(proyecto__isnull=False),
        "proyecto_id",
        "proyecto__nombre",
    )

    teachers = _by_teacher(dependent_qs)
    teachers = [
        {
            "id": item["docente_id"],
            "label": item["docente"],
            "correo": item["correo"],
            "es_externo": item["es_externo"],
            "total": item["total"],
        }
        for item in teachers
    ]

    # El catálogo temporal sí responde al resto de dimensiones, pero jamás al
    # propio período. Así el selector no se autolimita al cambiar el rango.
    period_qs = _apply_dimensions(base, filters, include_teacher=True)

    return {
        "sedes": sites,
        "facultades": faculties,
        "carreras": careers,
        "docentes": teachers,
        "tipos": types,
        "proyectos": projects,
        "periodo": _period_catalog(period_qs),
    }


def _teacher_catalogs(base_qs, filters):
    dimension_qs = _apply_dimensions(base_qs, filters, include_teacher=False)

    return {
        "tipos": _selector_rows(base_qs, "tipo_id", "tipo__nombre"),
        "proyectos": _selector_rows(
            base_qs.filter(proyecto__isnull=False),
            "proyecto_id",
            "proyecto__nombre",
        ),
        "periodo": _period_catalog(dimension_qs),
    }


# ============================================================
# DETALLE
# ============================================================


def _publication_title(publication):
    category = str(getattr(publication.tipo, "categoria", "") or "").lower()

    if category == "articulo":
        obj = getattr(publication, "articulo", None)
        return str(getattr(obj, "nombre_articulo", "") or "").strip()

    if category == "ponencia":
        obj = getattr(publication, "ponencia", None)
        return str(getattr(obj, "nombre_ponencia", "") or "").strip()

    if category == "libro":
        obj = getattr(publication, "libro", None)
        return str(getattr(obj, "nombre_libro", "") or "").strip()

    if category == "capitulo":
        obj = getattr(publication, "capitulo_libro", None)
        return str(getattr(obj, "nombre_capitulo", "") or "").strip()

    return ""


def _authors_text(publication):
    names = []

    participations = sorted(
        list(publication.participaciones.all()),
        key=lambda item: (item.orden or 0, item.pk or 0),
    )

    for participation in participations:
        author = participation.autor
        name = f"{author.nombres} {author.apellidos}".strip()
        if name:
            names.append(name)

    return " | ".join(names)


def _publication_period_label(publication):
    year = publication.anio_publicacion
    month = publication.mes_publicacion

    if month:
        return f"{MONTH_NAME.get(month, month)} {year}"

    return str(year or "—")


def _detail(qs, *, limit=None):
    ordered = qs.order_by(
        "-anio_publicacion",
        "-mes_publicacion",
        "tipo__nombre",
        "numero",
        "id",
    )

    rows = ordered if limit is None else ordered[:limit]

    output = []

    for publication in rows:
        output.append(
            {
                "publicacion_id": publication.pk,
                "numero": publication.numero,
                "titulo": _publication_title(publication) or f"Publicación #{publication.pk}",
                "tipo_id": publication.tipo_id,
                "tipo": publication.tipo.nombre if publication.tipo_id else None,
                "anio_publicacion": publication.anio_publicacion,
                "mes_publicacion": publication.mes_publicacion,
                "periodo": _publication_period_label(publication),
                "sede_id": publication.sede_id,
                "sede": publication.sede.nombre if publication.sede_id else None,
                "facultad_id": publication.carrera.facultad_id if publication.carrera_id else None,
                "facultad": publication.carrera.facultad.nombre if publication.carrera_id else None,
                "carrera_id": publication.carrera_id,
                "carrera": publication.carrera.nombre if publication.carrera_id else None,
                "proyecto_id": publication.proyecto_id,
                "proyecto": publication.proyecto.nombre if publication.proyecto_id else None,
                "autores": _authors_text(publication),
                "con_pdf": bool(publication.archivo_pdf),
            }
        )

    return output


# ============================================================
# PAYLOADS
# ============================================================


def build_institutional_production_report(params, *, full_detail=False):
    filters = _parse_dimension_filters(params, allow_teacher=True)
    period = resolve_period(params)

    base = _approved_publications()
    filtered = _apply_dimensions(base, filters, include_teacher=True)
    filtered = _apply_period(filtered, period)

    total = filtered.count()
    detail_limit = None if full_detail else filters["detalle_limite"]
    detail = _detail(filtered, limit=detail_limit)

    return {
        "ok": True,
        "alcance": "institucional",
        "criterio_estado": {
            "value": Publicacion.ESTADO_APROBADA,
            "label": "Solo publicaciones aprobadas",
        },
        "filtros_aplicados": {
            **filters,
            **period.as_dict(),
        },
        "filtros_disponibles": _institutional_catalogs(filters),
        "resumen": _summary(filtered, teacher_id=filters.get("docente_id")),
        "distribuciones": {
            "por_anio": _by_year(filtered),
            "por_mes": _by_month(filtered),
            "por_sede": _by_site(filtered),
            "por_facultad": _by_faculty(filtered),
            "por_carrera": _by_career(filtered),
            "por_docente": _by_teacher(
                filtered,
                teacher_id=filters.get("docente_id"),
            ),
            "por_tipo": _by_type(filtered),
            "por_proyecto": _by_project(filtered),
        },
        "detalle": {
            "total": total,
            "limite": total if full_detail else filters["detalle_limite"],
            "truncado": False if full_detail else total > len(detail),
            "items": detail,
        },
    }


def build_teacher_production_report(user, params, *, full_detail=False):
    filters = _parse_dimension_filters(params, allow_teacher=False)

    # El docente solo puede filtrar su producción por tipo/proyecto/período.
    filters["sede_id"] = None
    filters["facultad_id"] = None
    filters["carrera_id"] = None
    filters["tipo_codigo"] = None

    period = resolve_period(params)
    base = _teacher_publications(user)
    filtered = _apply_dimensions(base, filters, include_teacher=False)
    filtered = _apply_period(filtered, period)

    total = filtered.count()
    detail_limit = None if full_detail else filters["detalle_limite"]
    detail = _detail(filtered, limit=detail_limit)

    author = getattr(user, "autor", None)

    return {
        "ok": True,
        "alcance": "personal",
        "usuario": {
            "id": getattr(user, "pk", None),
            "nombre": getattr(user, "get_full_name", lambda: "")() or str(user),
            "email": getattr(user, "email", None),
            "autor_id": getattr(author, "pk", None),
        },
        "criterio_estado": {
            "value": Publicacion.ESTADO_APROBADA,
            "label": "Solo publicaciones aprobadas",
        },
        "filtros_aplicados": {
            "tipo_id": filters["tipo_id"],
            "proyecto_id": filters["proyecto_id"],
            "detalle_limite": filters["detalle_limite"],
            **period.as_dict(),
        },
        "filtros_disponibles": _teacher_catalogs(base, filters),
        "resumen": _summary(filtered),
        "distribuciones": {
            "por_anio": _by_year(filtered),
            "por_mes": _by_month(filtered),
            "por_tipo": _by_type(filtered),
            "por_proyecto": _by_project(filtered),
        },
        "detalle": {
            "total": total,
            "limite": total if full_detail else filters["detalle_limite"],
            "truncado": False if full_detail else total > len(detail),
            "items": detail,
        },
    }


# ============================================================
# EXCEL — DISEÑO ALINEADO CON EL PDF
# ============================================================


REPORT_PRIMARY = "1F4FD7"
REPORT_PRIMARY_DARK = "153A9B"
REPORT_TEXT = "172033"
REPORT_MUTED = "667085"
REPORT_LINE = "D9E0E8"
REPORT_SOFT = "F5F7FB"
REPORT_SOFT_BLUE = "EEF3FA"
REPORT_WHITE = "FFFFFF"
REPORT_SUCCESS = "17803D"
REPORT_COLUMNS = 12
REPORT_TABLE_HEADER_ROW = 6

REPORT_BORDER = Border(
    left=Side(style="thin", color=REPORT_LINE),
    right=Side(style="thin", color=REPORT_LINE),
    top=Side(style="thin", color=REPORT_LINE),
    bottom=Side(style="thin", color=REPORT_LINE),
)


def _safe(value):
    if value is None or value == "":
        return "—"
    if isinstance(value, bool):
        return "Sí" if value else "No"
    return value


def _set_report_page(ws, *, print_end_row=None, print_end_col=REPORT_COLUMNS):
    """Configura cada hoja como parte de un único informe visual."""

    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.freeze_panes = None

    ws.sheet_properties.tabColor = REPORT_PRIMARY

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.55
    ws.page_margins.bottom = 0.55
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.25

    ws.oddFooter.left.text = "SGPC ULEAM · Producción científica"
    ws.oddFooter.left.size = 8
    ws.oddFooter.left.color = REPORT_MUTED
    ws.oddFooter.right.text = "Página &P de &N"
    ws.oddFooter.right.size = 8
    ws.oddFooter.right.color = REPORT_MUTED

    if print_end_row:
        end_letter = get_column_letter(print_end_col)
        ws.print_area = f"A1:{end_letter}{int(print_end_row)}"


def _sheet_title(ws, title, subtitle=None, *, columns=REPORT_COLUMNS):
    """Cabecera inspirada en el PDF: limpia, amplia y sin banda técnica."""

    _set_report_page(ws)

    end_letter = get_column_letter(columns)
    ws.merge_cells(f"A1:{end_letter}1")
    ws.merge_cells(f"A2:{end_letter}2")

    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(
        bold=True,
        color=REPORT_TEXT,
        size=18,
    )
    title_cell.alignment = Alignment(
        vertical="center",
        horizontal="left",
    )
    ws.row_dimensions[1].height = 30

    subtitle_cell = ws["A2"]
    subtitle_cell.value = subtitle or "Reporte de producción científica aprobada"
    subtitle_cell.font = Font(
        color=REPORT_MUTED,
        size=9,
        italic=False,
    )
    subtitle_cell.alignment = Alignment(
        vertical="center",
        horizontal="left",
        wrap_text=True,
    )
    subtitle_cell.border = Border(
        bottom=Side(style="thin", color=REPORT_LINE)
    )
    ws.row_dimensions[2].height = 21


def _section_band(ws, row, title, *, start_col=1, end_col=REPORT_COLUMNS):
    start_letter = get_column_letter(start_col)
    end_letter = get_column_letter(end_col)
    ws.merge_cells(f"{start_letter}{row}:{end_letter}{row}")
    cell = ws.cell(row=row, column=start_col, value=title)
    cell.fill = PatternFill("solid", fgColor=REPORT_SOFT)
    cell.font = Font(bold=True, color=REPORT_TEXT, size=10.5)
    cell.alignment = Alignment(vertical="center", horizontal="left")
    cell.border = Border(
        top=Side(style="thin", color=REPORT_LINE),
        bottom=Side(style="thin", color=REPORT_LINE),
    )
    ws.row_dimensions[row].height = 22
    return row + 1


def _metadata_grid(ws, *, start_row, items, columns=REPORT_COLUMNS, per_row=3):
    """Replica el bloque de contexto del PDF mediante fichas suaves."""

    if not items:
        return start_row

    block_width = max(1, columns // per_row)
    groups = [items[index:index + per_row] for index in range(0, len(items), per_row)]

    row = start_row
    for group in groups:
        for index, (label, value) in enumerate(group):
            start_col = 1 + index * block_width
            end_col = min(columns, start_col + block_width - 1)

            start_letter = get_column_letter(start_col)
            end_letter = get_column_letter(end_col)

            ws.merge_cells(
                f"{start_letter}{row}:{end_letter}{row}"
            )
            ws.merge_cells(
                f"{start_letter}{row + 1}:{end_letter}{row + 1}"
            )

            label_cell = ws.cell(row=row, column=start_col, value=label)
            value_cell = ws.cell(row=row + 1, column=start_col, value=_safe(value))

            label_cell.fill = PatternFill("solid", fgColor=REPORT_SOFT)
            value_cell.fill = PatternFill("solid", fgColor=REPORT_SOFT)
            label_cell.font = Font(bold=True, color=REPORT_MUTED, size=7.5)
            value_cell.font = Font(color=REPORT_TEXT, size=9)
            label_cell.alignment = Alignment(vertical="bottom", wrap_text=True)
            value_cell.alignment = Alignment(vertical="top", wrap_text=True)

            for cells_row in ws.iter_rows(
                min_row=row,
                max_row=row + 1,
                min_col=start_col,
                max_col=end_col,
            ):
                for cell in cells_row:
                    cell.border = REPORT_BORDER

        # Celdas vacías de la última fila conservan el mismo panel visual.
        if len(group) < per_row:
            for index in range(len(group), per_row):
                start_col = 1 + index * block_width
                end_col = min(columns, start_col + block_width - 1)
                for cells_row in ws.iter_rows(
                    min_row=row,
                    max_row=row + 1,
                    min_col=start_col,
                    max_col=end_col,
                ):
                    for cell in cells_row:
                        cell.fill = PatternFill("solid", fgColor=REPORT_SOFT)
                        cell.border = REPORT_BORDER

        ws.row_dimensions[row].height = 16
        ws.row_dimensions[row + 1].height = 24
        row += 2

    return row


def _metric_cards(ws, *, start_row, metrics, columns=REPORT_COLUMNS, cards_per_row=4):
    """Indicadores como tarjetas, igual que el bloque de métricas del PDF."""

    if not metrics:
        return start_row

    card_width = max(1, columns // cards_per_row)
    row = start_row

    for group_start in range(0, len(metrics), cards_per_row):
        group = metrics[group_start:group_start + cards_per_row]

        for index, metric in enumerate(group):
            label, value, note, primary = metric
            start_col = 1 + index * card_width
            end_col = min(columns, start_col + card_width - 1)
            start_letter = get_column_letter(start_col)
            end_letter = get_column_letter(end_col)

            for offset in range(3):
                ws.merge_cells(
                    f"{start_letter}{row + offset}:{end_letter}{row + offset}"
                )

            label_cell = ws.cell(row=row, column=start_col, value=label)
            value_cell = ws.cell(row=row + 1, column=start_col, value=_safe(value))
            note_cell = ws.cell(row=row + 2, column=start_col, value=note)

            for cells_row in ws.iter_rows(
                min_row=row,
                max_row=row + 2,
                min_col=start_col,
                max_col=end_col,
            ):
                for cell in cells_row:
                    cell.fill = PatternFill("solid", fgColor=REPORT_WHITE)
                    cell.border = REPORT_BORDER

            label_cell.font = Font(bold=True, color=REPORT_MUTED, size=7.5)
            value_cell.font = Font(
                bold=True,
                color=REPORT_PRIMARY if primary else REPORT_TEXT,
                size=16,
            )
            note_cell.font = Font(color=REPORT_MUTED, size=7.2)

            label_cell.alignment = Alignment(horizontal="left", vertical="bottom")
            value_cell.alignment = Alignment(horizontal="left", vertical="center")
            note_cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True,
            )

        # Mantiene una cuadrícula completa aunque falten tarjetas.
        if len(group) < cards_per_row:
            for index in range(len(group), cards_per_row):
                start_col = 1 + index * card_width
                end_col = min(columns, start_col + card_width - 1)
                for cells_row in ws.iter_rows(
                    min_row=row,
                    max_row=row + 2,
                    min_col=start_col,
                    max_col=end_col,
                ):
                    for cell in cells_row:
                        cell.fill = PatternFill("solid", fgColor=REPORT_WHITE)
                        cell.border = REPORT_BORDER

        ws.row_dimensions[row].height = 18
        ws.row_dimensions[row + 1].height = 30
        ws.row_dimensions[row + 2].height = 21
        row += 4

    return row


def _write_table(ws, *, row, headers, rows, start_col=1, filters=False):
    """Tabla visualmente igual a las tablas del PDF, pero editable en Excel."""

    for column_offset, header in enumerate(headers):
        column = start_col + column_offset
        cell = ws.cell(row=row, column=column, value=header)
        cell.fill = PatternFill("solid", fgColor=REPORT_SOFT)
        cell.font = Font(bold=True, color=REPORT_TEXT, size=8)
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )
        cell.border = REPORT_BORDER

    ws.row_dimensions[row].height = 22
    current = row + 1

    for row_index, values in enumerate(rows):
        for column_offset, value in enumerate(values):
            column = start_col + column_offset
            cell = ws.cell(row=current, column=column, value=_safe(value))
            cell.fill = PatternFill(
                "solid",
                fgColor=REPORT_SOFT if row_index % 2 else REPORT_WHITE,
            )
            cell.font = Font(color=REPORT_TEXT, size=8)
            cell.border = REPORT_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        current += 1

    if not rows:
        cell = ws.cell(row=current, column=start_col, value="No hay información para esta selección.")
        cell.font = Font(color=REPORT_MUTED, italic=True, size=8)
        cell.fill = PatternFill("solid", fgColor=REPORT_WHITE)
        cell.border = REPORT_BORDER
        cell.alignment = Alignment(vertical="center")
        if len(headers) > 1:
            ws.merge_cells(
                start_row=current,
                start_column=start_col,
                end_row=current,
                end_column=start_col + len(headers) - 1,
            )
        current += 1

    if filters and current > row + 2:
        end_letter = get_column_letter(start_col + len(headers) - 1)
        start_letter = get_column_letter(start_col)
        ws.auto_filter.ref = f"{start_letter}{row}:{end_letter}{current - 1}"

    return current


def _autofit(ws, *, maximum=52, minimum=10):
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = minimum
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(maximum, len(value) + 2))
        ws.column_dimensions[letter].width = width


def _panel_cells(ws, cell_range):
    """Fondo/borde detrás de un gráfico para simular su panel del PDF."""

    for row in ws[cell_range]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=REPORT_WHITE)
            cell.border = REPORT_BORDER


def _catalog_label_for_export(payload, key, selected_id, default):
    if not selected_id:
        return default

    selected = str(selected_id)
    catalogs = payload.get("filtros_disponibles", {}) or {}

    for item in catalogs.get(key, []) or []:
        if str(item.get("id")) == selected:
            return str(item.get("label") or default).strip() or default

    return default


def _personal_export_context(payload):
    filters = payload.get("filtros_aplicados", {}) or {}
    return {
        "periodo": str(filters.get("label") or "Histórico"),
        "tipo": _catalog_label_for_export(
            payload,
            "tipos",
            filters.get("tipo_id"),
            "Todos los tipos",
        ),
        "proyecto": _catalog_label_for_export(
            payload,
            "proyectos",
            filters.get("proyecto_id"),
            "Todos los proyectos",
        ),
    }


def _institutional_export_context(payload):
    filters = payload.get("filtros_aplicados", {}) or {}
    return {
        "periodo": str(filters.get("label") or "Histórico"),
        "sede": _catalog_label_for_export(payload, "sedes", filters.get("sede_id"), "Todas las sedes"),
        "facultad": _catalog_label_for_export(payload, "facultades", filters.get("facultad_id"), "Todas las facultades"),
        "carrera": _catalog_label_for_export(payload, "carreras", filters.get("carrera_id"), "Todas las carreras"),
        "docente": _catalog_label_for_export(payload, "docentes", filters.get("docente_id"), "Todos los docentes"),
        "tipo": (
            CANONICAL_TYPE_LABELS.get(filters.get("tipo_codigo"))
            or _catalog_label_for_export(
                payload,
                "tipos",
                filters.get("tipo_id"),
                "Todos los tipos",
            )
        ),
        "proyecto": _catalog_label_for_export(payload, "proyectos", filters.get("proyecto_id"), "Todos los proyectos"),
    }


def _summary_sheet(workbook, payload, *, title):
    """Resumen institucional con jerarquía visual y área de gráficos adaptable."""

    ws = workbook.active
    ws.title = "Resumen"
    context = _institutional_export_context(payload)
    summary = payload.get("resumen", {}) or {}
    total = int(summary.get("total_publicaciones") or 0)
    with_pdf = int(summary.get("con_pdf") or 0)
    with_project = int(summary.get("con_proyecto") or 0)

    _sheet_title(
        ws,
        "Informe de producción científica",
        "Publicaciones aprobadas según las opciones seleccionadas.",
        columns=REPORT_COLUMNS,
    )

    metadata = [
        ("Sede", context["sede"]),
        ("Facultad", context["facultad"]),
        ("Carrera", context["carrera"]),
        ("Docente", context["docente"]),
        ("Tipo de publicación", context["tipo"]),
        ("Proyecto", context["proyecto"]),
        ("Período", context["periodo"]),
        ("Generado", timezone.localtime().strftime("%d/%m/%Y %H:%M")),
    ]
    row = _metadata_grid(ws, start_row=4, items=metadata)
    row += 1
    row = _section_band(ws, row, "Resumen")

    metrics = [
        ("Publicaciones", total, "Incluidas en el período", True),
        ("Docentes/Autores", int(summary.get("total_docentes") or 0), "Con producción aprobada", False),
        ("Proyectos", int(summary.get("total_proyectos") or 0), "Relacionados con publicaciones", False),
        ("Tipos presentes", int(summary.get("total_tipos") or 0), "Categorías de publicación", False),
        ("Cobertura documental", f"{summary.get('cobertura_pdf', 0)}%", f"{with_pdf} de {total} con PDF", False),
        ("Vinculación a proyectos", f"{summary.get('vinculacion_proyectos', 0)}%", f"{with_project} de {total} vinculadas", False),
        ("Con PDF", with_pdf, "Publicaciones con documento", False),
        ("Sin mes informado", int(summary.get("sin_mes") or 0), "Publicaciones sin mes registrado", False),
    ]
    row = _metric_cards(ws, start_row=row, metrics=metrics)
    row = _section_band(ws, row, "Visualización")

    for letter in "ABCDEFGHIJKL":
        ws.column_dimensions[letter].width = 11.5

    # Los paneles de gráficos se construyen después, cuando ya conocemos
    # cuántas categorías y qué longitud tienen sus etiquetas.
    _set_report_page(ws, print_end_row=row + 2)


def _personal_summary_sheet(workbook, payload):
    """Resumen personal con área de gráficos adaptable a los datos."""

    ws = workbook.active
    ws.title = "Resumen"

    summary = payload.get("resumen", {}) or {}
    user = payload.get("usuario", {}) or {}
    context = _personal_export_context(payload)
    total = int(summary.get("total_publicaciones") or 0)
    with_pdf = int(summary.get("con_pdf") or 0)
    with_project = int(summary.get("con_proyecto") or 0)

    _sheet_title(
        ws,
        "Mi producción científica",
        f"Resumen de publicaciones aprobadas de {user.get('nombre') or 'usuario'}",
        columns=REPORT_COLUMNS,
    )

    metadata = [
        ("Docente", user.get("nombre") or "—"),
        ("Período", context["periodo"]),
        ("Tipo de publicación", context["tipo"]),
        ("Proyecto", context["proyecto"]),
        ("Generado", timezone.localtime().strftime("%d/%m/%Y %H:%M")),
    ]
    row = _metadata_grid(ws, start_row=4, items=metadata)
    row += 1
    row = _section_band(ws, row, "Resumen")

    metrics = [
        ("Publicaciones aprobadas", total, "Incluidas en el período", True),
        ("Cobertura documental", f"{summary.get('cobertura_pdf', 0)}%", f"{with_pdf} de {total} con PDF", False),
        ("Vinculación a proyectos", f"{summary.get('vinculacion_proyectos', 0)}%", f"{with_project} de {total} vinculadas", False),
        ("Tipos presentes", int(summary.get("total_tipos") or 0), "Categorías de publicación", False),
    ]
    row = _metric_cards(ws, start_row=row, metrics=metrics)
    row = _section_band(ws, row, "Distribución de la producción")

    for letter in "ABCDEFGHIJKL":
        ws.column_dimensions[letter].width = 11.5

    # No se dibujan paneles fijos aquí. Se generan después con el tamaño
    # real de cada gráfico para evitar compresión, solapamiento y espacios
    # vacíos innecesarios.
    _set_report_page(ws, print_end_row=row + 2)


def _distribution_sheet(workbook, *, name, title, items, columns, subtitle=None):
    """Hoja de distribución con tabla + gráfico dentro del mismo marco visual."""

    ws = workbook.create_sheet(name)
    _sheet_title(
        ws,
        "Informe de producción científica",
        subtitle or title,
        columns=REPORT_COLUMNS,
    )
    _section_band(ws, 4, title)

    headers = [header for header, _key in columns]
    rows = [[item.get(key) for _header, key in columns] for item in items]
    end_row = _write_table(ws, row=REPORT_TABLE_HEADER_ROW, headers=headers, rows=rows)

    # La tabla ocupa el bloque izquierdo; el gráfico, el bloque derecho.
    table_cols = max(2, len(headers))
    for column in range(1, table_cols + 1):
        letter = get_column_letter(column)
        ws.column_dimensions[letter].width = 26 if column == 1 else 18

    for letter in "DEFGHIJKL":
        ws.column_dimensions[letter].width = 10.5

    panel_end_row = max(30, end_row + 1)
    _panel_cells(ws, f"D6:L{panel_end_row}")
    _set_report_page(ws, print_end_row=panel_end_row)
    return ws


def _personal_distribution_sheet(workbook, *, name, title, items, label_key):
    return _distribution_sheet(
        workbook,
        name=name,
        title=title,
        items=items,
        columns=((title.replace("Producción por ", "").capitalize(), label_key), ("Publicaciones", "total")),
        subtitle=f"Detalle de {title.lower()} para el período seleccionado.",
    )


def _detail_sheet(workbook, payload):
    """Detalle institucional sin identificadores internos visibles."""

    ws = workbook.create_sheet("Detalle")
    context = _institutional_export_context(payload)
    _sheet_title(
        ws,
        "Informe de producción científica",
        "Detalle de publicaciones aprobadas",
        columns=REPORT_COLUMNS,
    )

    metadata = [
        ("Sede", context["sede"]),
        ("Facultad", context["facultad"]),
        ("Carrera", context["carrera"]),
        ("Tipo", context["tipo"]),
        ("Proyecto", context["proyecto"]),
        ("Período", context["periodo"]),
    ]
    row = _metadata_grid(ws, start_row=4, items=metadata)
    row += 1
    row = _section_band(ws, row, "Publicaciones del período")

    items = payload.get("detalle", {}).get("items", []) or []
    rows = [
        [
            item.get("titulo"),
            item.get("tipo"),
            item.get("periodo"),
            item.get("sede"),
            item.get("facultad"),
            item.get("carrera"),
            item.get("proyecto") or "Sin proyecto",
            item.get("autores"),
            "Sí" if item.get("con_pdf") else "No",
        ]
        for item in items
    ]

    end_row = _write_table(
        ws,
        row=row,
        headers=[
            "Título",
            "Tipo",
            "Período",
            "Sede",
            "Facultad",
            "Carrera",
            "Proyecto",
            "Autores",
            "PDF",
        ],
        rows=rows,
        filters=True,
    )

    widths = {
        "A": 48,
        "B": 26,
        "C": 18,
        "D": 24,
        "E": 34,
        "F": 28,
        "G": 42,
        "H": 52,
        "I": 10,
    }
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    _set_report_page(ws, print_end_row=max(end_row, row + 6), print_end_col=9)


def _personal_detail_sheet(workbook, payload):
    ws = workbook.create_sheet("Publicaciones")
    context = _personal_export_context(payload)
    user = payload.get("usuario", {}) or {}

    _sheet_title(
        ws,
        "Mi producción científica",
        "Publicaciones aprobadas del período",
        columns=REPORT_COLUMNS,
    )

    metadata = [
        ("Docente", user.get("nombre") or "—"),
        ("Período", context["periodo"]),
        ("Tipo de publicación", context["tipo"]),
        ("Proyecto", context["proyecto"]),
        ("Generado", timezone.localtime().strftime("%d/%m/%Y %H:%M")),
    ]
    row = _metadata_grid(ws, start_row=4, items=metadata)
    row += 1
    row = _section_band(ws, row, "Publicaciones del período")

    items = payload.get("detalle", {}).get("items", []) or []
    rows = [
        [
            item.get("titulo"),
            item.get("tipo"),
            item.get("periodo"),
            item.get("proyecto") or "Sin proyecto",
            item.get("autores"),
            "Sí" if item.get("con_pdf") else "No",
        ]
        for item in items
    ]

    end_row = _write_table(
        ws,
        row=row,
        headers=["Título", "Tipo", "Período", "Proyecto", "Autores", "PDF"],
        rows=rows,
        filters=True,
    )

    widths = {
        "A": 52,
        "B": 28,
        "C": 18,
        "D": 44,
        "E": 58,
        "F": 10,
    }
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    _set_report_page(ws, print_end_row=max(end_row, row + 6), print_end_col=6)


def _chart_data_end_row(
    ws,
    label_col,
    value_col,
    *,
    header_row=REPORT_TABLE_HEADER_ROW,
    max_row=None,
):
    """
    Devuelve la última fila con datos reales de una tabla usada por un gráfico.

    No usa ``ws.max_row`` como límite lógico porque las celdas formateadas del
    panel del gráfico también aumentan ``max_row`` aunque estén vacías. Ese era
    el origen de categorías/leyendas vacías en los gráficos.
    """

    physical_last = min(
        int(max_row or ws.max_row or header_row),
        int(ws.max_row or header_row),
    )
    last_data_row = header_row

    for row in range(header_row + 1, physical_last + 1):
        label = ws.cell(row=row, column=label_col).value
        value = ws.cell(row=row, column=value_col).value

        if label in (None, "") or value in (None, ""):
            continue

        try:
            float(value)
        except (TypeError, ValueError, OverflowError):
            continue

        last_data_row = row

    return last_data_row


def _chart_row_count(
    ws,
    label_col,
    value_col,
    *,
    header_row=REPORT_TABLE_HEADER_ROW,
    max_row=None,
):
    end_row = _chart_data_end_row(
        ws,
        label_col,
        value_col,
        header_row=header_row,
        max_row=max_row,
    )
    return max(0, end_row - header_row)


def _chart_values(
    ws,
    value_col,
    *,
    header_row=REPORT_TABLE_HEADER_ROW,
    end_row=None,
):
    last = int(end_row or header_row)
    values = []

    for row in range(header_row + 1, last + 1):
        try:
            values.append(
                max(
                    0,
                    int(
                        ws.cell(
                            row=row,
                            column=value_col,
                        ).value
                        or 0
                    ),
                )
            )
        except (TypeError, ValueError, OverflowError):
            values.append(0)

    return values


def _chart_labels(
    ws,
    label_col,
    *,
    header_row=REPORT_TABLE_HEADER_ROW,
    end_row=None,
):
    last = int(end_row or header_row)
    return [
        str(
            ws.cell(
                row=row,
                column=label_col,
            ).value
            or ""
        ).strip()
        for row in range(header_row + 1, last + 1)
    ]


def _integer_axis(axis, maximum, *, headroom=False):
    """Configura una escala entera legible y, si aplica, deja aire superior."""

    maximum = max(1, int(maximum or 0))

    if headroom:
        if maximum <= 1:
            axis_max = 2
        else:
            axis_max = maximum + max(
                1,
                int(round(maximum * 0.15)),
            )
    else:
        axis_max = maximum

    axis.scaling.min = 0
    axis.scaling.max = axis_max
    axis.majorUnit = (
        1
        if axis_max <= 10
        else max(1, int(round(axis_max / 5)))
    )
    axis.numFmt = "0"


def _line_chart_size(row_count, width=None, height=None):
    """Tamaño de línea proporcional a la cantidad de categorías."""

    if width is None:
        width = min(
            19.0,
            max(
                13.5,
                10.2 + (max(1, int(row_count)) * 0.78),
            ),
        )

    if height is None:
        height = 7.3 if row_count <= 12 else 8.1

    return float(width), float(height)


def _doughnut_chart_size(row_count, width=None, height=None):
    """La dona crece ligeramente cuando la leyenda necesita más espacio."""

    if width is None:
        width = min(
            16.0,
            max(
                11.5,
                10.5 + (max(1, int(row_count)) * 0.55),
            ),
        )

    if height is None:
        height = min(
            9.2,
            max(
                7.2,
                6.4 + (max(1, int(row_count)) * 0.42),
            ),
        )

    return float(width), float(height)


def _horizontal_chart_size(row_count, max_label_length, width=None, height=None):
    """Ajusta barras a la cantidad de filas y a la longitud de sus etiquetas."""

    if width is None:
        width = min(
            19.0,
            max(
                14.5,
                13.2 + min(5.8, max(0, int(max_label_length)) * 0.16),
            ),
        )

    if height is None:
        height = min(
            12.0,
            max(
                7.2,
                4.8 + (max(1, int(row_count)) * 0.72),
            ),
        )

    return float(width), float(height)


def _chart_row_span(height):
    """Convierte aproximadamente centímetros de gráfico en filas de Excel."""

    return max(
        14,
        int(round(float(height) * 2.05)),
    )


def _add_line_chart(
    source_ws,
    target_ws,
    *,
    label_col,
    value_col,
    title,
    anchor,
    width=None,
    height=None,
    header_row=REPORT_TABLE_HEADER_ROW,
):
    end_row = _chart_data_end_row(
        source_ws,
        label_col,
        value_col,
        header_row=header_row,
    )
    row_count = max(0, end_row - header_row)

    if row_count < 1:
        return False

    values = _chart_values(
        source_ws,
        value_col,
        header_row=header_row,
        end_row=end_row,
    )
    maximum = max(values or [0])
    width, height = _line_chart_size(
        row_count,
        width=width,
        height=height,
    )

    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.height = height
    chart.width = width
    chart.legend = None
    chart.y_axis.title = "Publicaciones"
    chart.x_axis.title = ""
    _integer_axis(
        chart.y_axis,
        maximum,
        headroom=True,
    )

    chart.add_data(
        Reference(
            source_ws,
            min_col=value_col,
            min_row=header_row,
            max_row=end_row,
        ),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(
            source_ws,
            min_col=label_col,
            min_row=header_row + 1,
            max_row=end_row,
        )
    )

    if chart.series:
        series = chart.series[0]
        series.graphicalProperties.line.solidFill = REPORT_PRIMARY
        series.graphicalProperties.line.width = 24000
        series.marker.symbol = "circle"
        series.marker.size = 6
        series.marker.graphicalProperties.solidFill = REPORT_WHITE
        series.marker.graphicalProperties.line.solidFill = REPORT_PRIMARY

    # Etiquetas numéricas solamente cuando siguen siendo legibles.
    if row_count <= 10:
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True

    target_ws.add_chart(chart, anchor)
    return True


def _add_doughnut_chart(
    source_ws,
    target_ws,
    *,
    label_col,
    value_col,
    title,
    anchor,
    width=None,
    height=None,
    header_row=REPORT_TABLE_HEADER_ROW,
):
    end_row = _chart_data_end_row(
        source_ws,
        label_col,
        value_col,
        header_row=header_row,
    )
    row_count = max(0, end_row - header_row)

    if row_count < 1:
        return False

    width, height = _doughnut_chart_size(
        row_count,
        width=width,
        height=height,
    )

    chart = DoughnutChart()
    chart.title = title
    chart.style = 10
    chart.height = height
    chart.width = width
    chart.holeSize = 62
    chart.firstSliceAng = 270
    chart.legend.position = "b"

    chart.add_data(
        Reference(
            source_ws,
            min_col=value_col,
            min_row=header_row,
            max_row=end_row,
        ),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(
            source_ws,
            min_col=label_col,
            min_row=header_row + 1,
            max_row=end_row,
        )
    )

    chart.dataLabels = DataLabelList()

    # Con muchas categorías, los porcentajes encima de la dona dejan de ser
    # legibles. La leyenda permanece y evita el amontonamiento.
    chart.dataLabels.showPercent = row_count <= 6
    chart.dataLabels.showLeaderLines = row_count <= 6

    target_ws.add_chart(chart, anchor)
    return True


def _add_horizontal_chart(
    source_ws,
    target_ws,
    *,
    label_col,
    value_col,
    title,
    anchor,
    limit=10,
    width=None,
    height=None,
    header_row=REPORT_TABLE_HEADER_ROW,
):
    data_end_row = _chart_data_end_row(
        source_ws,
        label_col,
        value_col,
        header_row=header_row,
    )

    if data_end_row <= header_row:
        return False

    max_row = min(
        data_end_row,
        header_row + max(1, int(limit)),
    )
    row_count = max(0, max_row - header_row)

    values = _chart_values(
        source_ws,
        value_col,
        header_row=header_row,
        end_row=max_row,
    )
    labels = _chart_labels(
        source_ws,
        label_col,
        header_row=header_row,
        end_row=max_row,
    )

    maximum = max(values or [0])
    max_label_length = max(
        [len(label) for label in labels]
        or [0]
    )
    width, height = _horizontal_chart_size(
        row_count,
        max_label_length,
        width=width,
        height=height,
    )

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.grouping = "clustered"
    chart.overlap = 0
    chart.title = title
    chart.height = height
    chart.width = width
    chart.legend = None

    # En openpyxl el eje de categorías sigue siendo x_axis aunque la barra sea
    # horizontal; el eje numérico continúa siendo y_axis.
    chart.x_axis.title = ""
    chart.y_axis.title = "Publicaciones"
    _integer_axis(
        chart.y_axis,
        maximum,
        headroom=False,
    )

    chart.add_data(
        Reference(
            source_ws,
            min_col=value_col,
            min_row=header_row,
            max_row=max_row,
        ),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(
            source_ws,
            min_col=label_col,
            min_row=header_row + 1,
            max_row=max_row,
        )
    )

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    if chart.series:
        chart.series[0].graphicalProperties.solidFill = REPORT_PRIMARY
        chart.series[0].graphicalProperties.line.solidFill = REPORT_PRIMARY

    target_ws.add_chart(chart, anchor)
    return True


def _summary_chart_panel(
    ws,
    *,
    start_row,
    height,
    start_col=1,
    end_col=REPORT_COLUMNS,
):
    """Pinta el panel detrás de un gráfico y devuelve su última fila."""

    end_row = start_row + _chart_row_span(height)
    start_letter = get_column_letter(start_col)
    end_letter = get_column_letter(end_col)

    _panel_cells(
        ws,
        f"{start_letter}{start_row}:{end_letter}{end_row}",
    )

    return end_row


def _add_production_workbook_charts(workbook, payload, *, personal=False):
    """
    Inserta gráficos con un layout adaptable.

    Series temporales y barras con etiquetas largas usan todo el ancho.
    Dona y proyectos solo comparten fila cuando la información cabe sin
    comprimir leyendas ni categorías.
    """

    summary_ws = workbook["Resumen"]
    chart_row = 15 if personal else 21
    cursor = chart_row
    summary_end_row = chart_row

    year_ws = workbook["Por año"] if "Por año" in workbook.sheetnames else None
    month_ws = workbook["Por mes"] if "Por mes" in workbook.sheetnames else None
    type_ws = workbook["Por tipo"] if "Por tipo" in workbook.sheetnames else None
    project_ws = workbook["Por proyecto"] if "Por proyecto" in workbook.sheetnames else None

    # --------------------------------------------------------
    # 1. EVOLUCIÓN ANUAL — ancho completo
    # --------------------------------------------------------
    if year_ws is not None:
        year_count = _chart_row_count(
            year_ws,
            1,
            2,
        )

        if year_count:
            year_width, year_height = _line_chart_size(
                year_count,
                width=25.2,
                height=7.4 if year_count <= 12 else 8.2,
            )

            panel_end = _summary_chart_panel(
                summary_ws,
                start_row=cursor,
                height=year_height,
            )

            _add_line_chart(
                year_ws,
                summary_ws,
                label_col=1,
                value_col=2,
                title="Producción por año",
                anchor=f"A{cursor}",
                width=year_width,
                height=year_height,
            )

            summary_end_row = max(summary_end_row, panel_end)
            cursor = panel_end + 2

        _add_line_chart(
            year_ws,
            year_ws,
            label_col=1,
            value_col=2,
            title="Producción por año",
            anchor="D6",
        )

    # --------------------------------------------------------
    # 2. TIPOS + PROYECTOS — lado a lado solo si caben
    # --------------------------------------------------------
    type_count = 0
    project_count = 0
    project_max_label = 0

    if type_ws is not None:
        type_count = _chart_row_count(
            type_ws,
            1,
            2,
        )

    if project_ws is not None:
        project_end = _chart_data_end_row(
            project_ws,
            1,
            2,
        )
        project_count = max(0, min(project_end, REPORT_TABLE_HEADER_ROW + 8) - REPORT_TABLE_HEADER_ROW)
        if project_count:
            project_labels = _chart_labels(
                project_ws,
                1,
                end_row=min(
                    project_end,
                    REPORT_TABLE_HEADER_ROW + 8,
                ),
            )
            project_max_label = max(
                [len(label) for label in project_labels]
                or [0]
            )

    can_share_row = (
        type_count > 0
        and project_count > 0
        and type_count <= 6
        and project_count <= 4
        and project_max_label <= 28
    )

    if can_share_row:
        _type_w, type_height = _doughnut_chart_size(
            type_count,
            width=12.2,
        )
        _project_w, project_height = _horizontal_chart_size(
            project_count,
            project_max_label,
            width=12.8,
        )
        shared_height = max(type_height, project_height)

        panel_end_left = _summary_chart_panel(
            summary_ws,
            start_row=cursor,
            height=shared_height,
            start_col=1,
            end_col=6,
        )
        panel_end_right = _summary_chart_panel(
            summary_ws,
            start_row=cursor,
            height=shared_height,
            start_col=7,
            end_col=12,
        )

        _add_doughnut_chart(
            type_ws,
            summary_ws,
            label_col=1,
            value_col=2,
            title="Tipos de publicación",
            anchor=f"A{cursor}",
            width=12.2,
            height=shared_height,
        )
        _add_horizontal_chart(
            project_ws,
            summary_ws,
            label_col=1,
            value_col=2,
            title="Proyectos",
            anchor=f"G{cursor}",
            limit=8,
            width=12.8,
            height=shared_height,
        )

        summary_end_row = max(
            summary_end_row,
            panel_end_left,
            panel_end_right,
        )
        cursor = summary_end_row + 2

    else:
        if type_count > 0:
            _type_w, type_height = _doughnut_chart_size(
                type_count,
                width=17.5,
            )

            panel_end = _summary_chart_panel(
                summary_ws,
                start_row=cursor,
                height=type_height,
            )

            _add_doughnut_chart(
                type_ws,
                summary_ws,
                label_col=1,
                value_col=2,
                title="Tipos de publicación",
                anchor=f"C{cursor}",
                width=17.5,
                height=type_height,
            )

            summary_end_row = max(summary_end_row, panel_end)
            cursor = panel_end + 2

        if project_count > 0:
            _project_w, project_height = _horizontal_chart_size(
                project_count,
                project_max_label,
                width=25.2,
            )

            panel_end = _summary_chart_panel(
                summary_ws,
                start_row=cursor,
                height=project_height,
            )

            _add_horizontal_chart(
                project_ws,
                summary_ws,
                label_col=1,
                value_col=2,
                title="Proyectos",
                anchor=f"A{cursor}",
                limit=8,
                width=25.2,
                height=project_height,
            )

            summary_end_row = max(summary_end_row, panel_end)
            cursor = panel_end + 2

    if type_ws is not None:
        _add_doughnut_chart(
            type_ws,
            type_ws,
            label_col=1,
            value_col=2,
            title="Tipos de publicación",
            anchor="D6",
        )

    if project_ws is not None:
        _add_horizontal_chart(
            project_ws,
            project_ws,
            label_col=1,
            value_col=2,
            title="Producción por proyecto",
            anchor="D6",
            limit=10,
        )

    # --------------------------------------------------------
    # 3. EVOLUCIÓN MENSUAL — ancho completo
    # --------------------------------------------------------
    if month_ws is not None:
        month_visual = _month_visual_items(payload)
        month_title = "Producción por mes"

        if month_visual:
            month_years = {
                int(item.get("anio"))
                for item in month_visual
                if item.get("anio") not in (None, "")
            }
            if len(month_years) == 1:
                month_title = f"Producción por mes · {next(iter(month_years))}"

        month_count = _chart_row_count(
            month_ws,
            1,
            2,
        )

        if month_count:
            month_width, month_height = _line_chart_size(
                month_count,
                width=25.2,
                height=7.5 if month_count <= 12 else 8.3,
            )

            panel_end = _summary_chart_panel(
                summary_ws,
                start_row=cursor,
                height=month_height,
            )

            _add_line_chart(
                month_ws,
                summary_ws,
                label_col=1,
                value_col=2,
                title=month_title,
                anchor=f"A{cursor}",
                width=month_width,
                height=month_height,
            )

            summary_end_row = max(summary_end_row, panel_end)
            cursor = panel_end + 2

        _add_line_chart(
            month_ws,
            month_ws,
            label_col=1,
            value_col=2,
            title=month_title,
            anchor="D6",
        )

    # --------------------------------------------------------
    # 4. DIMENSIONES INSTITUCIONALES
    # --------------------------------------------------------
    if not personal:
        dimension_specs = (
            ("Por sede", 1, 2, "Producción por sede"),
            ("Por facultad", 1, 2, "Producción por facultad"),
            ("Por carrera", 1, 2, "Producción por carrera"),
            ("Por docente", 1, 3, "Producción por docente"),
        )

        for sheet_name, label_col, value_col, title in dimension_specs:
            if sheet_name not in workbook.sheetnames:
                continue

            worksheet = workbook[sheet_name]

            _add_horizontal_chart(
                worksheet,
                worksheet,
                label_col=label_col,
                value_col=value_col,
                title=title,
                anchor="E6" if sheet_name == "Por docente" else "D6",
                limit=10,
            )

    # El área de impresión debe acompañar el layout real, no una cuadrícula fija.
    _set_report_page(
        summary_ws,
        print_end_row=max(summary_end_row, chart_row + 2),
        print_end_col=REPORT_COLUMNS,
    )


def _build_personal_workbook(payload):
    workbook = Workbook()
    _personal_summary_sheet(workbook, payload)

    distributions = payload.get("distribuciones", {}) or {}

    _personal_distribution_sheet(
        workbook,
        name="Por año",
        title="Producción por año",
        items=distributions.get("por_anio", []),
        label_key="anio",
    )
    _personal_distribution_sheet(
        workbook,
        name="Por tipo",
        title="Producción por tipo",
        items=distributions.get("por_tipo", []),
        label_key="tipo",
    )

    month_items = _month_visual_items(payload)
    _personal_distribution_sheet(
        workbook,
        name="Por mes",
        title="Producción por mes",
        items=month_items if month_items else distributions.get("por_mes", []),
        label_key="chart_label",
    )
    _personal_distribution_sheet(
        workbook,
        name="Por proyecto",
        title="Producción por proyecto",
        items=distributions.get("por_proyecto", []),
        label_key="proyecto",
    )
    _personal_detail_sheet(workbook, payload)
    _add_production_workbook_charts(workbook, payload, personal=True)
    workbook.active = 0
    return workbook


def _build_workbook(payload, *, personal=False):
    workbook = Workbook()
    _summary_sheet(
        workbook,
        payload,
        title="Informe de producción científica",
    )

    distributions = payload.get("distribuciones", {}) or {}
    context = _institutional_export_context(payload)
    common_subtitle = f"Período: {context['periodo']} · Solo publicaciones aprobadas"

    _distribution_sheet(
        workbook,
        name="Por año",
        title="Producción por año",
        items=distributions.get("por_anio", []),
        columns=(("Año", "anio"), ("Publicaciones", "total")),
        subtitle=common_subtitle,
    )

    month_items = _month_visual_items(payload)
    _distribution_sheet(
        workbook,
        name="Por mes",
        title="Producción por mes",
        items=month_items if month_items else distributions.get("por_mes", []),
        columns=(("Mes", "chart_label"), ("Publicaciones", "total")),
        subtitle=common_subtitle,
    )

    _distribution_sheet(
        workbook,
        name="Por tipo",
        title="Tipos de publicación",
        items=distributions.get("por_tipo", []),
        columns=(("Tipo", "tipo"), ("Publicaciones", "total")),
        subtitle=common_subtitle,
    )

    _distribution_sheet(
        workbook,
        name="Por proyecto",
        title="Producción por proyecto",
        items=distributions.get("por_proyecto", []),
        columns=(("Proyecto", "proyecto"), ("Publicaciones", "total")),
        subtitle=common_subtitle,
    )

    if not personal:
        _distribution_sheet(
            workbook,
            name="Por sede",
            title="Producción por sede",
            items=distributions.get("por_sede", []),
            columns=(("Sede", "sede"), ("Publicaciones", "total")),
            subtitle=common_subtitle,
        )
        _distribution_sheet(
            workbook,
            name="Por facultad",
            title="Producción por facultad",
            items=distributions.get("por_facultad", []),
            columns=(("Facultad", "facultad"), ("Publicaciones", "total")),
            subtitle=common_subtitle,
        )
        _distribution_sheet(
            workbook,
            name="Por carrera",
            title="Producción por carrera",
            items=distributions.get("por_carrera", []),
            columns=(("Carrera", "carrera"), ("Publicaciones", "total")),
            subtitle=common_subtitle,
        )
        _distribution_sheet(
            workbook,
            name="Por docente",
            title="Producción por docente/autor",
            items=distributions.get("por_docente", []),
            columns=(("Docente", "docente"), ("Correo", "correo"), ("Publicaciones", "total")),
            subtitle=common_subtitle,
        )

    _detail_sheet(workbook, payload)
    _add_production_workbook_charts(workbook, payload, personal=False)
    workbook.active = 0
    return workbook


def _workbook_bytes(workbook):
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _filename(prefix):
    stamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{stamp}.xlsx"


def build_institutional_production_report_file(params):
    payload = build_institutional_production_report(params, full_detail=True)
    workbook = _build_workbook(payload, personal=False)
    return _workbook_bytes(workbook), _filename("reporte_produccion_institucional")


def build_teacher_production_report_file(user, params):
    payload = build_teacher_production_report(user, params, full_detail=True)
    workbook = _build_personal_workbook(payload)
    return _workbook_bytes(workbook), _filename("mi_produccion_cientifica")
