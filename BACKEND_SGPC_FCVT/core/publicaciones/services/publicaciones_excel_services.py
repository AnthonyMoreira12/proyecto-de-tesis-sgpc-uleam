"""Servicio de exportación Excel de publicaciones.

Genera una hoja diferente según el tipo:

- Artículos de alto impacto
- Artículos regionales
- Ponencias
- Libros
- Capítulos de libro

Regla institucional:

    Publicacion -> Carrera -> Facultad

Los filtros, la búsqueda y el ordenamiento se delegan al
servicio centralizado de listados. De esta manera, la interfaz
y el archivo Excel utilizan exactamente el mismo queryset.
"""

from collections import OrderedDict
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import (
    get_column_letter,
)

from core.publicaciones.services.publicaciones_listado_services import (
    build_publicaciones_queryset,
    extract_publicaciones_filters,
)


TIPO_ORDER = [
    "alto_impacto",
    "regional",
    "ponencia",
    "libro",
    "capitulo",
]


TIPO_SHEET_NAMES = {
    "alto_impacto": (
        "Matriz de artículos de alto impacto"
    ),
    "regional": (
        "Matriz de artículos regionales"
    ),
    "ponencia": (
        "Matriz de ponencias"
    ),
    "libro": (
        "Matriz de libros"
    ),
    "capitulo": (
        "Matriz de capítulos de libro"
    ),
}


NOTE_MAP = {
    "alto_impacto": (
        "Artículos de alto impacto: datos institucionales, "
        "revista, impacto, cuartil, autores y archivos."
    ),
    "regional": (
        "Artículos regionales: datos institucionales, "
        "indexación, revista, autores y archivos."
    ),
    "ponencia": (
        "Ponencias: evento, ubicación, presentación, "
        "arbitraje, autores y archivos."
    ),
    "libro": (
        "Libros: información editorial, arbitraje, "
        "autores y archivos."
    ),
    "capitulo": (
        "Capítulos de libro: libro contenedor, arbitraje, "
        "autores y archivos."
    ),
}


TITLE_BG = "E5E7EB"
TITLE_TEXT = "0F172A"
NAVY_2 = "254F87"
HEADER_BG = "9BBADD"
HEADER_TEXT = "11253D"
BODY_BG = "F2F2F2"
BODY_BG_ALT = "ECECEC"
NUM_BG = "DCE9F7"
BORDER = "6F89A6"


def _normalize_text(value):
    return str(
        value or ""
    ).strip()


def _normalize_lower(value):
    value = _normalize_text(
        value
    )

    return (
        value.lower()
        if value
        else ""
    )


def _safe_related(
    instance,
    attr_name,
):
    if instance is None:
        return None

    try:
        return getattr(
            instance,
            attr_name,
        )
    except Exception:
        return None


def _facultad_from_publicacion(
    publicacion,
):
    carrera = _safe_related(
        publicacion,
        "carrera",
    )

    return _safe_related(
        carrera,
        "facultad",
    )


def _tipo_bucket(
    publicacion,
):
    articulo = _safe_related(
        publicacion,
        "articulo",
    )

    if articulo:
        tipo_articulo = (
            _normalize_lower(
                articulo.tipo_articulo
            )
        )

        if (
            tipo_articulo
            == "alto_impacto"
        ):
            return "alto_impacto"

        if tipo_articulo == "regional":
            return "regional"

    if _safe_related(
        publicacion,
        "ponencia",
    ):
        return "ponencia"

    if _safe_related(
        publicacion,
        "libro",
    ):
        return "libro"

    if _safe_related(
        publicacion,
        "capitulo_libro",
    ):
        return "capitulo"

    return None


def _build_queryset(
    filters=None,
):
    """
    Construye el queryset utilizado por el reporte Excel.

    La exportación reutiliza la misma normalización, búsqueda,
    clasificación, filtros relacionales y ordenamiento de los
    listados institucionales.
    """

    normalized_filters = (
        extract_publicaciones_filters(
            filters or {}
        )
    )

    return build_publicaciones_queryset(
        filters=normalized_filters,
        user=None,
        solo_mias=False,
    )


def count_publicaciones_excel(
    filters=None,
):
    """
    Devuelve la cantidad exacta de publicaciones que serían
    incluidas en el reporte con los filtros recibidos.

    La función usa el mismo queryset que genera el archivo, por
    lo que puede emplearse en la vista previa del frontend.
    """

    return _build_queryset(
        filters or {}
    ).count()


def _get_participaciones(
    publicacion,
):
    return (
        getattr(
            publicacion,
            "participaciones_ordenadas",
            None,
        )
        or []
    )


def _get_archivos(
    publicacion,
):
    prefetched = getattr(
        publicacion,
        "_prefetched_objects_cache",
        {},
    )

    if "archivos" in prefetched:
        return sorted(
            prefetched["archivos"],
            key=lambda item: (
                getattr(
                    item,
                    "orden",
                    0,
                ),
                getattr(
                    item,
                    "id",
                    0,
                ),
            ),
        )

    try:
        return list(
            publicacion.archivos
            .all()
            .order_by(
                "orden",
                "id",
            )
        )

    except Exception:
        return []


def _autores_text(
    publicacion,
):
    """
    Devuelve todos los autores en el orden bibliográfico
    registrado para la publicación.

    No existe clasificación entre autor principal y coautor.
    """

    participaciones = (
        _get_participaciones(
            publicacion
        )
    )

    autores = []

    for participacion in participaciones:
        autor = _safe_related(
            participacion,
            "autor",
        )

        if autor is None:
            continue

        nombre = (
            f"{_normalize_text(autor.nombres)} "
            f"{_normalize_text(autor.apellidos)}"
        ).strip()

        if not nombre:
            nombre = (
                _normalize_text(
                    autor.correo
                )
                or _normalize_text(
                    autor.identificacion
                )
            )

        if nombre:
            autores.append(nombre)

    return (
        " | ".join(autores)
        or "—"
    )


def _adjuntos_text(
    publicacion,
):
    archivos = _get_archivos(
        publicacion
    )

    nombres = [
        _normalize_text(
            archivo.nombre
        )
        for archivo in archivos
        if _normalize_text(
            archivo.nombre
        )
    ]

    if nombres:
        return " | ".join(
            nombres
        )

    return "—"


def _principal_pdf_text(
    publicacion,
):
    archivo = getattr(
        publicacion,
        "archivo_pdf",
        None,
    )

    name = _normalize_text(
        getattr(
            archivo,
            "name",
            None,
        )
    )

    return (
        name
        or "—"
    )


def _pdf_text(
    publicacion,
):
    principal_pdf = (
        _principal_pdf_text(
            publicacion
        )
    )

    adjuntos = (
        _adjuntos_text(
            publicacion
        )
    )

    values = []

    if principal_pdf != "—":
        values.append(
            principal_pdf
        )

    if adjuntos != "—":
        values.append(
            adjuntos
        )

    return (
        " | ".join(values)
        or "—"
    )


def _has_any_pdf(
    publicacion,
):
    archivo = getattr(
        publicacion,
        "archivo_pdf",
        None,
    )

    if (
        archivo
        and getattr(
            archivo,
            "name",
            None,
        )
    ):
        return True

    return bool(
        _get_archivos(
            publicacion
        )
    )


def _display_choice(
    instance,
    method_name,
    fallback="—",
):
    try:
        method = getattr(
            instance,
            method_name,
            None,
        )

        if callable(method):
            value = method()

            return (
                _normalize_text(
                    value
                )
                or fallback
            )

    except Exception:
        pass

    return fallback


def _display_cuartil(
    articulo,
):
    value = _normalize_lower(
        articulo.cuartil
    )

    if value == "sin_cuartil":
        return "Sin cuartil"

    if not value:
        return "—"

    return _display_choice(
        articulo,
        "get_cuartil_display",
        value.upper(),
    )


def _display_factor_impacto(
    articulo,
):
    value = _normalize_lower(
        articulo.factor_impacto
    )

    if not value:
        return "—"

    return _display_choice(
        articulo,
        "get_factor_impacto_display",
        value.upper(),
    )


def _display_base_indexada(
    articulo,
):
    value = _normalize_lower(
        articulo.base_datos_indexada
    )

    if not value:
        return "—"

    return _display_choice(
        articulo,
        "get_base_datos_indexada_display",
        value,
    )


def _origen_publicacion_data(
    publicacion,
):
    """
    Devuelve la etiqueta del origen y su detalle.

    Reglas:
    - TIC: el detalle corresponde al grado o programa.
    - Otro: el detalle corresponde al origen escrito
      manualmente por el usuario.
    - Demás opciones: el detalle no aplica.
    """

    origen_tipo = _normalize_lower(
        getattr(
            publicacion,
            "origen_tipo",
            "",
        )
    )

    origen_display = _display_choice(
        publicacion,
        "get_origen_tipo_display",
        fallback=(
            origen_tipo
            or "—"
        ),
    )

    origen_detalle = "—"

    if origen_tipo in {
        "tic",
        "otro",
    }:
        origen_detalle = (
            _normalize_text(
                getattr(
                    publicacion,
                    "origen_grado",
                    None,
                )
            )
            or "—"
        )

    return (
        origen_display,
        origen_detalle,
    )


def _base_common(
    publicacion,
):
    facultad = (
        _facultad_from_publicacion(
            publicacion
        )
    )

    carrera = _safe_related(
        publicacion,
        "carrera",
    )

    proyecto = _safe_related(
        publicacion,
        "proyecto",
    )

    area = _safe_related(
        publicacion,
        "area",
    )

    subarea = _safe_related(
        publicacion,
        "subarea",
    )

    (
        origen_publicacion,
        detalle_origen,
    ) = _origen_publicacion_data(
        publicacion
    )

    return OrderedDict(
        {
            "N°": (
                publicacion.numero
                or publicacion.id
            ),
            "Facultad": (
                _normalize_text(
                    getattr(
                        facultad,
                        "nombre",
                        "",
                    )
                )
                or "—"
            ),
            "Carrera": (
                _normalize_text(
                    getattr(
                        carrera,
                        "nombre",
                        "",
                    )
                )
                or "—"
            ),
            "Proyecto": (
                _normalize_text(
                    getattr(
                        proyecto,
                        "nombre",
                        "",
                    )
                )
                or "—"
            ),
            "Área": (
                _normalize_text(
                    getattr(
                        area,
                        "nombre",
                        "",
                    )
                )
                or "—"
            ),
            "Subárea": (
                _normalize_text(
                    getattr(
                        subarea,
                        "nombre",
                        "",
                    )
                )
                or "—"
            ),
            "Origen de la publicación": (
                origen_publicacion
            ),
            "Detalle del origen": (
                detalle_origen
            ),
        }
    )


def _periodo_text(
    publicacion,
):
    """
    Representa el período sin inventar un día de publicación.

    Ejemplos:
        Agosto de 2026
        2026
    """

    anio = getattr(
        publicacion,
        "anio_publicacion",
        None,
    )

    mes = getattr(
        publicacion,
        "mes_publicacion",
        None,
    )

    if not anio:
        return "—"

    if not mes:
        return str(anio)

    display = getattr(
        publicacion,
        "get_mes_publicacion_display",
        None,
    )

    if callable(display):
        try:
            mes_label = _normalize_text(
                display()
            )
        except Exception:
            mes_label = ""
    else:
        mes_label = ""

    if not mes_label:
        meses = {
            1: "Enero",
            2: "Febrero",
            3: "Marzo",
            4: "Abril",
            5: "Mayo",
            6: "Junio",
            7: "Julio",
            8: "Agosto",
            9: "Septiembre",
            10: "Octubre",
            11: "Noviembre",
            12: "Diciembre",
        }
        mes_label = meses.get(
            mes,
            str(mes),
        )

    return f"{mes_label} de {anio}"


def _row_alto_impacto(
    publicacion,
):
    autores = (
        _autores_text(
            publicacion
        )
    )

    articulo = publicacion.articulo

    row = _base_common(
        publicacion
    )

    row.update(
        {
            "Título del artículo": (
                articulo.nombre_articulo
            ),
            "Periodo de publicación": (
                _periodo_text(
                    publicacion
                )
            ),
            "Nombre revista": (
                articulo.nombre_revista
            ),
            "N° revista": (
                articulo.numero_revista
                if articulo.numero_revista
                is not None
                else "—"
            ),
            "ISSN": (
                articulo.codigo_issn
            ),
            "DOI": (
                _normalize_text(
                    articulo.codigo_doi
                )
                or "—"
            ),
            "Factor impacto": (
                _display_factor_impacto(
                    articulo
                )
            ),
            "Cuartil": (
                _display_cuartil(
                    articulo
                )
            ),
            "SJR": (
                _normalize_text(
                    articulo.sjr
                )
                or "—"
            ),
            "Link revista": (
                _normalize_text(
                    articulo.link_revista
                )
                or "—"
            ),
            "Link publicación": (
                _normalize_text(
                    articulo.link_publicacion
                )
                or "—"
            ),
            "Autores": autores,
            "Archivos PDF": (
                _pdf_text(
                    publicacion
                )
            ),
        }
    )

    return row


def _row_regional(
    publicacion,
):
    autores = (
        _autores_text(
            publicacion
        )
    )

    articulo = publicacion.articulo

    row = _base_common(
        publicacion
    )

    row.update(
        {
            "Título del artículo": (
                articulo.nombre_articulo
            ),
            "Periodo de publicación": (
                _periodo_text(
                    publicacion
                )
            ),
            "Base indexada": (
                _display_base_indexada(
                    articulo
                )
            ),
            "Otra base": (
                _normalize_text(
                    articulo.base_datos_otra
                )
                if (
                    _normalize_lower(
                        articulo
                        .base_datos_indexada
                    )
                    == "otra"
                )
                else "—"
            )
            or "—",
            "Nombre revista": (
                articulo.nombre_revista
            ),
            "N° revista": (
                articulo.numero_revista
                if articulo.numero_revista
                is not None
                else "—"
            ),
            "ISSN": (
                articulo.codigo_issn
            ),
            "DOI": (
                _normalize_text(
                    articulo.codigo_doi
                )
                or "—"
            ),
            "Link revista": (
                _normalize_text(
                    articulo.link_revista
                )
                or "—"
            ),
            "Link publicación": (
                _normalize_text(
                    articulo.link_publicacion
                )
                or "—"
            ),
            "Autores": autores,
            "Archivos PDF": (
                _pdf_text(
                    publicacion
                )
            ),
        }
    )

    return row


def _row_ponencia(
    publicacion,
):
    autores = (
        _autores_text(
            publicacion
        )
    )

    ponencia = publicacion.ponencia

    pais = _safe_related(
        publicacion,
        "pais",
    )

    ciudad = _safe_related(
        publicacion,
        "ciudad",
    )

    row = _base_common(
        publicacion
    )

    row.update(
        {
            "Nombre del evento": (
                ponencia.nombre_evento
            ),
            "Nombre ponencia": (
                ponencia.nombre_ponencia
            ),
            "Periodo de presentación": (
                _periodo_text(
                    publicacion
                )
            ),
            "País": (
                _normalize_text(
                    getattr(
                        pais,
                        "nombre",
                        "",
                    )
                )
                or "—"
            ),
            "Ciudad": (
                _normalize_text(
                    getattr(
                        ciudad,
                        "nombre",
                        "",
                    )
                )
                or "—"
            ),
            "ISSN / ISBN": (
                _normalize_text(
                    ponencia
                    .codigo_issn_isbn
                )
                or "—"
            ),
            "Tipo presentación": (
                _display_choice(
                    ponencia,
                    "get_tipo_presentacion_display",
                )
            ),
            "Tipo presentación - Otro": (
                _normalize_text(
                    ponencia
                    .tipo_presentacion_otro
                )
                or "—"
            ),
            "Revisor / arbitraje": (
                _display_choice(
                    ponencia,
                    "get_revisor_par_arbitraje_display",
                )
            ),
            "Link evento": (
                _normalize_text(
                    ponencia.link_evento
                )
                or "—"
            ),
            "Autores": autores,
            "PDF disponible": (
                "Sí"
                if _has_any_pdf(
                    publicacion
                )
                else "No"
            ),
            "Archivos PDF": (
                _pdf_text(
                    publicacion
                )
            ),
        }
    )

    return row


def _row_libro(
    publicacion,
):
    autores = (
        _autores_text(
            publicacion
        )
    )

    libro = publicacion.libro

    row = _base_common(
        publicacion
    )

    row.update(
        {
            "Nombre del libro": (
                libro.nombre_libro
            ),
            "Periodo de publicación": (
                _periodo_text(
                    publicacion
                )
            ),
            "ISBN": libro.codigo_isbn,
            "Editorial / Compilador": (
                libro.editorial_compilador
            ),
            "Revisor / arbitraje": (
                _display_choice(
                    libro,
                    "get_revisor_par_arbitraje_display",
                )
            ),
            "Link libro": (
                libro.link_libro
            ),
            "Autores": autores,
            "Archivos PDF": (
                _pdf_text(
                    publicacion
                )
            ),
        }
    )

    return row


def _row_capitulo(
    publicacion,
):
    autores = (
        _autores_text(
            publicacion
        )
    )

    capitulo = (
        publicacion.capitulo_libro
    )

    row = _base_common(
        publicacion
    )

    row.update(
        {
            "Nombre del capítulo": (
                capitulo.nombre_capitulo
            ),
            "Nombre del libro": (
                capitulo.nombre_libro
            ),
            "Periodo de publicación": (
                _periodo_text(
                    publicacion
                )
            ),
            "ISBN": (
                capitulo.codigo_isbn
            ),
            "Editor / compilador": (
                capitulo.editor_compilador
            ),
            "Revisor / arbitraje": (
                _display_choice(
                    capitulo,
                    "get_revisor_par_arbitraje_display",
                )
            ),
            "Link capítulo": (
                capitulo.link_capitulo
            ),
            "Autores": autores,
            "Archivos PDF": (
                _pdf_text(
                    publicacion
                )
            ),
        }
    )

    return row


ROW_BUILDERS = {
    "alto_impacto": (
        _row_alto_impacto
    ),
    "regional": (
        _row_regional
    ),
    "ponencia": (
        _row_ponencia
    ),
    "libro": (
        _row_libro
    ),
    "capitulo": (
        _row_capitulo
    ),
}


def _apply_title_block(
    ws,
    title,
    subtitle,
    note,
    total_columns,
):
    end_col = get_column_letter(
        max(total_columns, 1)
    )

    ws.merge_cells(
        f"A1:{end_col}1"
    )

    title_cell = ws["A1"]
    title_cell.value = title

    title_cell.font = Font(
        size=24,
        bold=True,
        color=TITLE_TEXT,
    )

    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    title_cell.fill = PatternFill(
        "solid",
        fgColor=TITLE_BG,
    )

    ws.merge_cells(
        f"A2:{end_col}2"
    )

    subtitle_cell = ws["A2"]
    subtitle_cell.value = subtitle

    subtitle_cell.font = Font(
        size=16,
        bold=True,
        color=TITLE_TEXT,
    )

    subtitle_cell.alignment = (
        Alignment(
            horizontal="center",
            vertical="center",
        )
    )

    subtitle_cell.fill = (
        PatternFill(
            "solid",
            fgColor=TITLE_BG,
        )
    )

    ws.merge_cells(
        f"A3:{end_col}3"
    )

    note_cell = ws["A3"]
    note_cell.value = note

    note_cell.font = Font(
        size=11,
        color="FFFFFF",
    )

    note_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    note_cell.fill = PatternFill(
        "solid",
        fgColor=NAVY_2,
    )

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 34


def _style_sheet(
    ws,
    *,
    total_columns,
    total_rows,
):
    thin = Side(
        style="thin",
        color=BORDER,
    )

    for cell in ws[4]:
        cell.fill = PatternFill(
            "solid",
            fgColor=HEADER_BG,
        )

        cell.font = Font(
            bold=True,
            color=HEADER_TEXT,
            size=11,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        cell.border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )

    ws.row_dimensions[4].height = 36

    for row_idx in range(
        5,
        total_rows + 1,
    ):
        fill_color = (
            BODY_BG
            if row_idx % 2 == 1
            else BODY_BG_ALT
        )

        for col_idx in range(
            1,
            total_columns + 1,
        ):
            cell = ws.cell(
                row=row_idx,
                column=col_idx,
            )

            cell.border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin,
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

            if col_idx == 1:
                cell.fill = PatternFill(
                    "solid",
                    fgColor=NUM_BG,
                )

                cell.font = Font(
                    bold=True,
                    color=TITLE_TEXT,
                    size=11,
                )

            else:
                cell.fill = PatternFill(
                    "solid",
                    fgColor=fill_color,
                )

                cell.font = Font(
                    color=TITLE_TEXT,
                    size=11,
                )

        ws.row_dimensions[
            row_idx
        ].height = 34

    ws.freeze_panes = "A5"

    if total_columns:
        ws.auto_filter.ref = (
            f"A4:"
            f"{get_column_letter(total_columns)}"
            f"{total_rows}"
        )


def _autosize_columns(ws):
    for column_index in range(
        1,
        ws.max_column + 1,
    ):
        max_length = 0

        for row_index in range(
            4,
            ws.max_row + 1,
        ):
            value = ws.cell(
                row=row_index,
                column=column_index,
            ).value

            if value is None:
                continue

            text = str(value)

            max_length = max(
                max_length,
                min(
                    len(text),
                    50,
                ),
            )

        width = min(
            max(
                max_length + 3,
                12,
            ),
            45,
        )

        ws.column_dimensions[
            get_column_letter(
                column_index
            )
        ].width = width


def build_publicaciones_excel(
    filters=None,
):
    publicaciones = (
        _build_queryset(
            filters or {}
        )
    )

    grouped = {
        key: []
        for key in TIPO_ORDER
    }

    for publicacion in publicaciones:
        bucket = _tipo_bucket(
            publicacion
        )

        if bucket in grouped:
            grouped[bucket].append(
                publicacion
            )

    workbook = Workbook()

    default_sheet = (
        workbook.active
    )

    workbook.remove(
        default_sheet
    )

    included = [
        key
        for key in TIPO_ORDER
        if grouped.get(key)
    ]

    # ---------------------------------------------------------
    # Sin resultados
    # ---------------------------------------------------------

    if not included:
        ws = workbook.create_sheet(
            "Sin resultados"
        )

        ws["A1"] = (
            "Reporte de publicaciones"
        )

        ws["A2"] = (
            "No existen registros para "
            "los filtros seleccionados."
        )

        ws["A1"].font = Font(
            size=16,
            bold=True,
        )

        ws["A2"].font = Font(
            size=12
        )

        ws.column_dimensions[
            "A"
        ].width = 60

        return workbook

    # ---------------------------------------------------------
    # Hojas por tipo
    # ---------------------------------------------------------

    for bucket in included:
        ws = workbook.create_sheet(
            TIPO_SHEET_NAMES[
                bucket
            ][:31]
        )

        rows = [
            ROW_BUILDERS[
                bucket
            ](
                publicacion
            )
            for publicacion
            in grouped[bucket]
        ]

        if not rows:
            continue

        headers = list(
            rows[0].keys()
        )

        _apply_title_block(
            ws,
            TIPO_SHEET_NAMES[
                bucket
            ],
            "Dirección de Investigación",
            NOTE_MAP.get(
                bucket,
                (
                    "Reporte de "
                    "publicaciones científicas."
                ),
            ),
            len(headers),
        )

        # -----------------------------------------------------
        # Cabeceras
        # -----------------------------------------------------

        for column_index, header in enumerate(
            headers,
            start=1,
        ):
            ws.cell(
                row=4,
                column=column_index,
                value=header,
            )

        # -----------------------------------------------------
        # Datos
        # -----------------------------------------------------

        for row_index, row in enumerate(
            rows,
            start=5,
        ):
            for column_index, value in enumerate(
                row.values(),
                start=1,
            ):
                ws.cell(
                    row=row_index,
                    column=column_index,
                    value=value,
                )

        _style_sheet(
            ws,
            total_columns=len(
                headers
            ),
            total_rows=ws.max_row,
        )

        _autosize_columns(
            ws
        )

    return workbook


def workbook_to_bytes(
    workbook,
):
    output = BytesIO()

    workbook.save(
        output
    )

    output.seek(0)

    return output.getvalue()