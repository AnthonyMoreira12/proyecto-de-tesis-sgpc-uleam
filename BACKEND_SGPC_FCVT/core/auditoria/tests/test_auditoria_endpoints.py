from io import BytesIO

from django.contrib.auth import get_user_model
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework.test import APITestCase

from core.auditoria.services.auditoria_services import (
    registrar_evento_auditoria,
)


User = get_user_model()


class AuditoriaEndpointsTests(
    APITestCase
):
    def setUp(self):
        self.admin = (
            User.objects
            .create_superuser(
                email=(
                    "audit-endpoints@example.com"
                ),
                nombres="Audit",
                apellidos="Admin",
                password=(
                    "ClaveSegura123!"
                ),
            )
        )

        self.client.force_authenticate(
            self.admin
        )

        registrar_evento_auditoria(
            actor=self.admin,
            accion="actualizar",
            modulo="publicaciones",
            entidad_tipo="Publicacion",
            entidad_id=25,
            descripcion=(
                "Prueba de auditoría"
            ),
            datos_anteriores={
                "sede": None,
            },
            datos_nuevos={
                "sede": 1,
            },
        )

    def test_resumen_auditoria(
        self,
    ):
        response = self.client.get(
            reverse(
                "admin-auditoria-resumen"
            )
        )

        self.assertEqual(
            response.status_code,
            200,
        )

        self.assertGreaterEqual(
            response.data[
                "ultimas_24_horas"
            ],
            1,
        )

        self.assertGreaterEqual(
            response.data[
                "publicaciones_24h"
            ],
            1,
        )

    def test_exportar_auditoria_excel(
        self,
    ):
        response = self.client.get(
            reverse(
                "admin-auditoria-exportar"
            )
        )

        self.assertEqual(
            response.status_code,
            200,
        )

        self.assertIn(
            (
                "application/vnd."
                "openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            response[
                "Content-Type"
            ],
        )

        self.assertIn(
            ".xlsx",
            response[
                "Content-Disposition"
            ],
        )

        workbook = load_workbook(
            BytesIO(
                response.content
            ),
            read_only=True,
            data_only=True,
        )

        self.assertEqual(
            workbook.sheetnames,
            [
                "Resumen",
                "Eventos de auditoría",
                "Cambios realizados",
                "Datos técnicos",
            ],
        )

        # ====================================================
        # EVENTOS
        # ====================================================

        events_sheet = workbook[
            "Eventos de auditoría"
        ]

        event_values = [
            cell.value
            for row
            in events_sheet.iter_rows()
            for cell
            in row
            if cell.value
            is not None
        ]

        self.assertIn(
            "Prueba de auditoría",
            event_values,
        )

        self.assertIn(
            "Publicación #25",
            event_values,
        )

        # ====================================================
        # CAMBIOS
        # ====================================================

        changes_sheet = workbook[
            "Cambios realizados"
        ]

        change_values = [
            cell.value
            for row
            in changes_sheet.iter_rows()
            for cell
            in row
            if cell.value
            is not None
        ]

        self.assertIn(
            "Sede",
            change_values,
        )

        self.assertIn(
            "Sin información",
            change_values,
        )

        self.assertIn(
            "1",
            [
                str(value)
                for value
                in change_values
            ],
        )

        # ====================================================
        # DATOS TÉCNICOS
        # ====================================================

        technical_sheet = workbook[
            "Datos técnicos"
        ]

        technical_values = [
            cell.value
            for row
            in technical_sheet.iter_rows()
            for cell
            in row
            if cell.value
            is not None
        ]

        self.assertIn(
            "Datos anteriores",
            technical_values,
        )

        self.assertIn(
            "Datos nuevos",
            technical_values,
        )

        workbook.close()

    def test_exportar_auditoria_excel_respeta_filtro_modulo(
        self,
    ):
        registrar_evento_auditoria(
            actor=self.admin,
            accion="actualizar",
            modulo="usuarios",
            entidad_tipo="Usuario",
            entidad_id=self.admin.pk,
            descripcion=(
                "Evento que no debe "
                "aparecer en el Excel filtrado"
            ),
            datos_anteriores={
                "perfil_completo":
                    False,
            },
            datos_nuevos={
                "perfil_completo":
                    True,
            },
        )

        response = self.client.get(
            reverse(
                "admin-auditoria-exportar"
            ),
            {
                "modulo":
                    "publicaciones",
            },
        )

        self.assertEqual(
            response.status_code,
            200,
        )

        workbook = load_workbook(
            BytesIO(
                response.content
            ),
            read_only=True,
            data_only=True,
        )

        events_sheet = workbook[
            "Eventos de auditoría"
        ]

        values = [
            str(
                cell.value
            )
            for row
            in events_sheet.iter_rows()
            for cell
            in row
            if cell.value
            is not None
        ]

        self.assertTrue(
            any(
                "Prueba de auditoría"
                in value
                for value
                in values
            )
        )

        self.assertFalse(
            any(
                (
                    "Evento que no debe "
                    "aparecer"
                )
                in value
                for value
                in values
            )
        )

        workbook.close()