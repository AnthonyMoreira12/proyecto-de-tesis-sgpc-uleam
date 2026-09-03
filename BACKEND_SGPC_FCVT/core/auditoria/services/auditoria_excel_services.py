"""Generación del reporte Excel de auditoría del SGPC."""

from __future__ import annotations

import json
from collections import OrderedDict
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

from django.db.models import Count
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURACIÓN
# ============================================================

EXCEL_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument."
    "spreadsheetml.sheet"
)

MAX_EXCEL_TEXT = 32000


# ============================================================
# COLORES
# ============================================================

COLOR_PRIMARY = "1F4E78"
COLOR_PRIMARY_DARK = "17365D"

COLOR_TEXT = "1F2937"
COLOR_MUTED = "667085"

COLOR_BORDER = "D0D5DD"

COLOR_WHITE = "FFFFFF"
COLOR_SOFT = "F8FAFC"


THIN_BORDER = Border(
    left=Side(
        style="thin",
        color=COLOR_BORDER,
    ),
    right=Side(
        style="thin",
        color=COLOR_BORDER,
    ),
    top=Side(
        style="thin",
        color=COLOR_BORDER,
    ),
    bottom=Side(
        style="thin",
        color=COLOR_BORDER,
    ),
)


# ============================================================
# ETIQUETAS
# ============================================================

ACTION_LABELS = {
    "crear": "Creación",
    "actualizar": "Actualización",
    "eliminar": "Eliminación",
    "activar": "Activación",
    "finalizar": "Finalización",
    "cancelar": "Cancelación",
    "login": "Inicio de sesión",
    "logout": "Cierre de sesión",
    "aprobar": "Aprobación",
    "rechazar": "Rechazo",
    "observar": "Observación",
    "enviar": "Envío",
    "exportar": "Exportación",
}


MODULE_LABELS = {
    "usuarios": "Usuarios",
    "estructura_academica": "Estructura académica",
    "proyectos": "Proyectos",
    "avisos": "Avisos",
    "catalogos": "Catálogos",
    "publicaciones": "Publicaciones",
    "autores": "Autores",
    "actualizaciones": "Actualización de datos",
    "comunicaciones": "Comunicaciones",
    "perfil": "Perfil",
    "notificaciones": "Notificaciones",
}


ENTITY_LABELS = {
    "Usuario": "Usuario",
    "Sede": "Sede",
    "Facultad": "Facultad",
    "Carrera": "Carrera",
    "CarreraSede": "Relación Carrera–Sede",

    "Proyecto": "Proyecto",
    "ProyectoAutor": "Participante de proyecto",

    "Banner": "Aviso",
    "BannerConfiguracion": "Configuración de avisos",

    "AreaConocimiento": "Área de conocimiento",
    "Subarea": "Subárea de conocimiento",

    "Pais": "País",
    "Ciudad": "Ciudad",

    "TipoPublicacion": "Tipo de publicación",
    "Publicacion": "Publicación",
    "Ponencia": "Ponencia",
    "Articulo": "Artículo",
    "Libro": "Libro",
    "CapituloLibro": "Capítulo de libro",

    "Autor": "Autor",
    "PublicacionAutor": "Autor de publicación",
    "PublicacionArchivo": "Archivo de publicación",
    "PublicacionRevision": "Revisión de publicación",
    "PublicacionHistorial": "Historial de publicación",

    "ComunicacionGlobal": "Comunicación global",

    "CampaniaActualizacion": "Campaña de actualización",
    "CampaniaActualizacionUsuario": (
        "Participante de campaña"
    ),

    "SolicitudModificacionPublicacion": (
        "Solicitud de modificación"
    ),
}


FIELD_LABELS = {
    "tipo": "Tipo",
    "titulo": "Título",
    "nombre": "Nombre",
    "mensaje": "Mensaje",
    "descripcion": "Descripción",

    "estado": "Estado",
    "activo": "Activo",
    "activa": "Activa",

    "identificacion": "Cédula",

    "sede": "Sede",
    "sede_id": "Sede",

    "facultad": "Facultad",
    "facultad_id": "Facultad",

    "carrera": "Carrera",
    "carrera_id": "Carrera",

    "area": "Área UNESCO",
    "area_id": "Área UNESCO",

    "subarea": "Subárea UNESCO",
    "subarea_id": "Subárea UNESCO",

    "pais": "País",
    "ciudad": "Ciudad",

    "proyecto": "Proyecto",
    "proyecto_id": "Proyecto",

    "fecha_inicio": "Fecha de inicio",
    "fecha_fin": "Fecha de finalización",

    "fecha_fin_planificada": (
        "Fecha de finalización planificada"
    ),

    "fecha_fin_prorrogada": (
        "Fecha de finalización prorrogada"
    ),

    "perfil_completo": "Perfil completo",

    "archivo_pdf_nombre_original": (
        "Nombre original del PDF"
    ),

    "archivo_pdf_tamano_bytes": (
        "Tamaño del PDF"
    ),

    "archivo_pdf_sha256": (
        "SHA-256 del PDF"
    ),

    "campania_id": "Campaña",

    "origen": "Origen",

    "created_at": "Creado",

    "updated_at": "Última actualización",
}


FILTER_LABELS = {
    "q": "Búsqueda",
    "actor_id": "Usuario",
    "modulo": "Módulo",
    "accion": "Acción",
    "entidad_tipo": "Tipo de entidad",
    "entidad_id": "ID de entidad",
    "fecha_desde": "Desde",
    "fecha_hasta": "Hasta",
}


# ============================================================
# ETIQUETAS HUMANAS
# ============================================================

def _humanize(value: Any) -> str:
    text = str(
        value or ""
    ).replace(
        "_",
        " ",
    ).strip()

    if not text:
        return ""

    return (
        text[:1].upper()
        + text[1:]
    )


def _action_label(value: Any) -> str:
    key = str(
        value or ""
    ).strip().lower()

    return (
        ACTION_LABELS.get(key)
        or _humanize(key)
        or "Movimiento"
    )


def _module_label(value: Any) -> str:
    key = str(
        value or ""
    ).strip().lower()

    return (
        MODULE_LABELS.get(key)
        or _humanize(key)
        or "General"
    )


def _entity_label(
    entity_type: Any,
    entity_id: Any,
) -> str:
    key = str(
        entity_type or ""
    ).strip()

    label = (
        ENTITY_LABELS.get(key)
        or _humanize(key)
        or "Registro"
    )

    if entity_id not in (
        None,
        "",
    ):
        return (
            f"{label} #{entity_id}"
        )

    return label


def _field_label(path: str) -> str:
    key = str(
        path or ""
    ).split(".")[-1]

    return (
        FIELD_LABELS.get(key)
        or _humanize(key)
        or "Campo"
    )


# ============================================================
# USUARIO
# ============================================================

def _actor_name(event) -> str:
    actor = getattr(
        event,
        "actor",
        None,
    )

    if actor is None:
        return "Sistema"

    nombres = str(
        getattr(
            actor,
            "nombres",
            "",
        )
        or ""
    ).strip()

    apellidos = str(
        getattr(
            actor,
            "apellidos",
            "",
        )
        or ""
    ).strip()

    full_name = " ".join(
        item
        for item in (
            nombres,
            apellidos,
        )
        if item
    ).strip()

    if full_name:
        return full_name

    email = str(
        getattr(
            actor,
            "email",
            "",
        )
        or ""
    ).strip()

    if email:
        return email

    return (
        f"Usuario #{actor.pk}"
    )


def _actor_email(event) -> str:
    actor = getattr(
        event,
        "actor",
        None,
    )

    if actor is None:
        return ""

    return str(
        getattr(
            actor,
            "email",
            "",
        )
        or ""
    ).strip()


# ============================================================
# SEGURIDAD DE TEXTO
# ============================================================

def _safe_text(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(
        value,
        (
            dict,
            list,
            tuple,
        ),
    ):
        value = json.dumps(
            value,
            ensure_ascii=False,
            default=str,
        )

    elif isinstance(
        value,
        Decimal,
    ):
        value = str(value)

    else:
        value = str(value)

    value = (
        ILLEGAL_CHARACTERS_RE.sub(
            "",
            value,
        )
    )

    if (
        len(value)
        > MAX_EXCEL_TEXT
    ):
        value = (
            value[
                : MAX_EXCEL_TEXT - 3
            ]
            + "..."
        )

    # Evita que un texto sea interpretado
    # como fórmula de Excel.
    if value.startswith("="):
        value = "'" + value

    return value


def _json_text(value: Any) -> str:
    if value in (
        None,
        "",
        {},
        [],
    ):
        return ""

    try:
        return _safe_text(
            json.dumps(
                value,
                ensure_ascii=False,
                indent=2,
                default=str,
            )
        )
    except Exception:
        return _safe_text(value)


# ============================================================
# FECHAS
# ============================================================

def _local_datetime(
    value: Any,
) -> datetime | None:
    if not isinstance(
        value,
        datetime,
    ):
        return None

    if timezone.is_aware(
        value
    ):
        value = (
            timezone.localtime(
                value
            ).replace(
                tzinfo=None
            )
        )

    return value


# ============================================================
# TAMAÑO
# ============================================================

def _format_bytes(
    value: int,
) -> str:
    if value < 1024:
        return f"{value} B"

    size = float(value)

    for unit in (
        "KB",
        "MB",
        "GB",
        "TB",
    ):
        size /= 1024

        if size < 1024:
            return (
                f"{size:.2f} {unit}"
            )

    return f"{size:.2f} PB"


# ============================================================
# VISUALIZACIÓN DE VALORES
# ============================================================

def _display_value(
    value: Any,
    path: str = "",
) -> str:
    if value in (
        None,
        "",
    ):
        return "Sin información"

    if isinstance(
        value,
        bool,
    ):
        return (
            "Sí"
            if value
            else "No"
        )

    if isinstance(
        value,
        datetime,
    ):
        local = (
            _local_datetime(
                value
            )
        )

        return (
            local.strftime(
                "%d/%m/%Y %H:%M"
            )
            if local
            else ""
        )

    if isinstance(
        value,
        date,
    ):
        return value.strftime(
            "%d/%m/%Y"
        )

    if path.endswith(
        "tamano_bytes"
    ):
        try:
            return _format_bytes(
                int(value)
            )
        except (
            TypeError,
            ValueError,
        ):
            pass

    if isinstance(
        value,
        (
            dict,
            list,
            tuple,
        ),
    ):
        return _json_text(value)

    return _safe_text(value)


# ============================================================
# COMPARACIÓN DE SNAPSHOTS
# ============================================================

def _stable_value(
    value: Any,
) -> str:
    try:
        return json.dumps(
            value,
            ensure_ascii=False,
            sort_keys=True,
            default=str,
        )
    except Exception:
        return str(value)


def _diff_objects(
    before: dict | None,
    after: dict | None,
    parent: str = "",
) -> list[dict]:
    before = (
        before
        if isinstance(
            before,
            dict,
        )
        else {}
    )

    after = (
        after
        if isinstance(
            after,
            dict,
        )
        else {}
    )

    keys = list(
        OrderedDict.fromkeys(
            [
                *before.keys(),
                *after.keys(),
            ]
        )
    )

    changes = []

    for key in keys:
        old_value = (
            before.get(key)
        )

        new_value = (
            after.get(key)
        )

        path = (
            f"{parent}.{key}"
            if parent
            else str(key)
        )

        if (
            isinstance(
                old_value,
                dict,
            )
            and isinstance(
                new_value,
                dict,
            )
        ):
            changes.extend(
                _diff_objects(
                    old_value,
                    new_value,
                    path,
                )
            )

            continue

        if (
            _stable_value(
                old_value
            )
            !=
            _stable_value(
                new_value
            )
        ):
            changes.append(
                {
                    "path": path,
                    "before": old_value,
                    "after": new_value,
                }
            )

    return changes


# ============================================================
# ESTILO GENERAL
# ============================================================

def _style_title(
    sheet,
    title: str,
    subtitle: str,
    end_column: int,
) -> None:
    sheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=end_column,
    )

    title_cell = sheet.cell(
        row=1,
        column=1,
        value=title,
    )

    title_cell.font = Font(
        name="Aptos Display",
        size=18,
        bold=True,
        color=COLOR_WHITE,
    )

    title_cell.fill = PatternFill(
        "solid",
        fgColor=COLOR_PRIMARY_DARK,
    )

    title_cell.alignment = Alignment(
        vertical="center",
    )

    sheet.row_dimensions[
        1
    ].height = 30

    sheet.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=end_column,
    )

    subtitle_cell = sheet.cell(
        row=2,
        column=1,
        value=subtitle,
    )

    subtitle_cell.font = Font(
        name="Aptos",
        size=10,
        color=COLOR_MUTED,
        italic=True,
    )

    subtitle_cell.alignment = (
        Alignment(
            vertical="center",
        )
    )

    sheet.row_dimensions[
        2
    ].height = 22


def _style_header_row(
    sheet,
    row: int,
    start_column: int,
    end_column: int,
) -> None:
    for column in range(
        start_column,
        end_column + 1,
    ):
        cell = sheet.cell(
            row=row,
            column=column,
        )

        cell.font = Font(
            name="Aptos",
            size=10,
            bold=True,
            color=COLOR_WHITE,
        )

        cell.fill = PatternFill(
            "solid",
            fgColor=COLOR_PRIMARY,
        )

        cell.alignment = Alignment(
            vertical="center",
            horizontal="left",
            wrap_text=True,
        )

        cell.border = THIN_BORDER

    sheet.row_dimensions[
        row
    ].height = 26


def _style_data_area(
    sheet,
    start_row: int,
    end_row: int,
    start_column: int,
    end_column: int,
) -> None:
    if (
        end_row
        < start_row
    ):
        return

    for row in range(
        start_row,
        end_row + 1,
    ):
        for column in range(
            start_column,
            end_column + 1,
        ):
            cell = sheet.cell(
                row=row,
                column=column,
            )

            cell.font = Font(
                name="Aptos",
                size=9,
                color=COLOR_TEXT,
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

            cell.border = THIN_BORDER

            if row % 2 == 0:
                cell.fill = (
                    PatternFill(
                        "solid",
                        fgColor=COLOR_SOFT,
                    )
                )


def _set_widths(
    sheet,
    widths: dict[int, float],
) -> None:
    for (
        column,
        width,
    ) in widths.items():
        sheet.column_dimensions[
            get_column_letter(
                column
            )
        ].width = width


# ============================================================
# FILTROS
# ============================================================

def _filter_rows(
    params,
) -> list[tuple[str, str]]:
    if not params:
        return []

    rows = []

    for key in params.keys():
        values = [
            str(value).strip()
            for value
            in params.getlist(key)
            if str(value).strip()
        ]

        if not values:
            continue

        label = (
            FILTER_LABELS.get(key)
            or _humanize(key)
        )

        if key == "modulo":
            display = ", ".join(
                _module_label(value)
                for value in values
            )

        elif key == "accion":
            display = ", ".join(
                _action_label(value)
                for value in values
            )

        else:
            display = ", ".join(
                values
            )

        rows.append(
            (
                label,
                display,
            )
        )

    return rows


# ============================================================
# HOJA: RESUMEN
# ============================================================

def _build_summary_sheet(
    workbook,
    queryset,
    params,
) -> None:
    sheet = workbook.active

    sheet.title = "Resumen"

    _style_title(
        sheet,
        (
            "SGPC ULEAM · "
            "REPORTE DE AUDITORÍA"
        ),
        (
            "Resumen ejecutivo de movimientos "
            "registrados en el sistema"
        ),
        5,
    )

    sheet.sheet_view.showGridLines = (
        False
    )

    sheet.freeze_panes = "A4"

    _set_widths(
        sheet,
        {
            1: 34,
            2: 28,
            3: 18,
            4: 18,
            5: 18,
        },
    )

    now = (
        timezone.localtime()
        .replace(
            tzinfo=None
        )
    )

    sheet["A4"] = "Generado"

    sheet["A4"].font = Font(
        bold=True,
        color=COLOR_MUTED,
    )

    sheet["B4"] = now

    sheet["B4"].number_format = (
        "dd/mm/yyyy hh:mm"
    )

    row = 6

    sheet.cell(
        row=row,
        column=1,
        value="FILTROS APLICADOS",
    ).font = Font(
        bold=True,
        color=COLOR_PRIMARY_DARK,
    )

    row += 1

    filters = _filter_rows(
        params
    )

    if filters:
        for (
            label,
            value,
        ) in filters:
            sheet.cell(
                row=row,
                column=1,
                value=label,
            ).font = Font(
                bold=True,
                color=COLOR_MUTED,
            )

            sheet.cell(
                row=row,
                column=2,
                value=_safe_text(
                    value
                ),
            )

            row += 1

    else:
        sheet.cell(
            row=row,
            column=1,
            value="Sin filtros",
        ).font = Font(
            bold=True,
            color=COLOR_MUTED,
        )

        sheet.cell(
            row=row,
            column=2,
            value=(
                "Se incluyeron todos "
                "los registros."
            ),
        )

        row += 1

    total = queryset.count()

    users = (
        queryset
        .exclude(
            actor_id__isnull=True
        )
        .values(
            "actor_id"
        )
        .distinct()
        .count()
    )

    modules = list(
        queryset
        .values(
            "modulo"
        )
        .annotate(
            total=Count("id")
        )
        .order_by(
            "-total",
            "modulo",
        )
    )

    actions = list(
        queryset
        .values(
            "accion"
        )
        .annotate(
            total=Count("id")
        )
        .order_by(
            "-total",
            "accion",
        )
    )

    row += 1

    sheet.cell(
        row=row,
        column=1,
        value="RESUMEN GENERAL",
    ).font = Font(
        bold=True,
        color=COLOR_PRIMARY_DARK,
    )

    row += 1

    summary_values = (
        (
            "Eventos encontrados",
            total,
        ),
        (
            "Usuarios involucrados",
            users,
        ),
        (
            "Módulos con actividad",
            len(modules),
        ),
        (
            "Tipos de acción",
            len(actions),
        ),
    )

    for (
        label,
        value,
    ) in summary_values:
        sheet.cell(
            row=row,
            column=1,
            value=label,
        ).font = Font(
            bold=True,
            color=COLOR_MUTED,
        )

        sheet.cell(
            row=row,
            column=2,
            value=value,
        ).font = Font(
            bold=True,
            size=11,
            color=COLOR_PRIMARY_DARK,
        )

        row += 1

    # --------------------------------------------------------
    # ACTIVIDAD POR ACCIÓN
    # --------------------------------------------------------

    row += 1

    sheet.cell(
        row=row,
        column=1,
        value="EVENTOS POR ACCIÓN",
    ).font = Font(
        bold=True,
        color=COLOR_PRIMARY_DARK,
    )

    row += 1

    sheet.cell(
        row=row,
        column=1,
        value="Acción",
    )

    sheet.cell(
        row=row,
        column=2,
        value="Eventos",
    )

    _style_header_row(
        sheet,
        row,
        1,
        2,
    )

    row += 1

    action_start = row

    for item in actions:
        sheet.cell(
            row=row,
            column=1,
            value=_action_label(
                item["accion"]
            ),
        )

        sheet.cell(
            row=row,
            column=2,
            value=item["total"],
        )

        row += 1

    _style_data_area(
        sheet,
        action_start,
        row - 1,
        1,
        2,
    )

    # --------------------------------------------------------
    # ACTIVIDAD POR MÓDULO
    # --------------------------------------------------------

    row += 2

    sheet.cell(
        row=row,
        column=1,
        value="EVENTOS POR MÓDULO",
    ).font = Font(
        bold=True,
        color=COLOR_PRIMARY_DARK,
    )

    row += 1

    sheet.cell(
        row=row,
        column=1,
        value="Módulo",
    )

    sheet.cell(
        row=row,
        column=2,
        value="Eventos",
    )

    _style_header_row(
        sheet,
        row,
        1,
        2,
    )

    row += 1

    module_start = row

    for item in modules:
        sheet.cell(
            row=row,
            column=1,
            value=_module_label(
                item["modulo"]
            ),
        )

        sheet.cell(
            row=row,
            column=2,
            value=item["total"],
        )

        row += 1

    _style_data_area(
        sheet,
        module_start,
        row - 1,
        1,
        2,
    )


# ============================================================
# HOJAS DE DETALLE
# ============================================================

def _prepare_detail_sheets(
    workbook,
):
    header_row = 4

    # --------------------------------------------------------
    # EVENTOS
    # --------------------------------------------------------

    events = workbook.create_sheet(
        "Eventos de auditoría"
    )

    event_headers = [
        "Fecha",
        "Hora",
        "Usuario",
        "Correo",
        "Acción",
        "Módulo",
        "Registro afectado",
        "Descripción",
        "Campos modificados",
    ]

    _style_title(
        events,
        (
            "SGPC ULEAM · "
            "EVENTOS DE AUDITORÍA"
        ),
        (
            "Listado administrativo de movimientos "
            "registrados en el sistema"
        ),
        len(event_headers),
    )

    for (
        column,
        header,
    ) in enumerate(
        event_headers,
        start=1,
    ):
        events.cell(
            row=header_row,
            column=column,
            value=header,
        )

    _style_header_row(
        events,
        header_row,
        1,
        len(event_headers),
    )

    events.freeze_panes = "A5"

    events.sheet_view.showGridLines = (
        False
    )

    _set_widths(
        events,
        {
            1: 13,
            2: 10,
            3: 30,
            4: 34,
            5: 19,
            6: 24,
            7: 30,
            8: 55,
            9: 38,
        },
    )

    # --------------------------------------------------------
    # CAMBIOS
    # --------------------------------------------------------

    changes = workbook.create_sheet(
        "Cambios realizados"
    )

    change_headers = [
        "Fecha",
        "Hora",
        "Usuario",
        "Acción",
        "Módulo",
        "Registro",
        "Campo",
        "Antes",
        "Después",
    ]

    _style_title(
        changes,
        (
            "SGPC ULEAM · "
            "CAMBIOS REALIZADOS"
        ),
        (
            "Comparación campo por campo de la "
            "información registrada o modificada"
        ),
        len(change_headers),
    )

    for (
        column,
        header,
    ) in enumerate(
        change_headers,
        start=1,
    ):
        changes.cell(
            row=header_row,
            column=column,
            value=header,
        )

    _style_header_row(
        changes,
        header_row,
        1,
        len(change_headers),
    )

    changes.freeze_panes = "A5"

    changes.sheet_view.showGridLines = (
        False
    )

    _set_widths(
        changes,
        {
            1: 13,
            2: 10,
            3: 30,
            4: 19,
            5: 24,
            6: 30,
            7: 30,
            8: 42,
            9: 42,
        },
    )

    # --------------------------------------------------------
    # DATOS TÉCNICOS
    # --------------------------------------------------------

    technical = workbook.create_sheet(
        "Datos técnicos"
    )

    technical_headers = [
        "Fecha y hora",
        "Usuario",
        "Correo",
        "IP",
        "Método HTTP",
        "Ruta",
        "User-Agent",
        "Tipo interno",
        "ID interno",
        "Datos anteriores",
        "Datos nuevos",
        "Contexto",
    ]

    _style_title(
        technical,
        (
            "SGPC ULEAM · "
            "DATOS TÉCNICOS"
        ),
        (
            "Información de soporte y trazabilidad "
            "para revisión técnica"
        ),
        len(
            technical_headers
        ),
    )

    for (
        column,
        header,
    ) in enumerate(
        technical_headers,
        start=1,
    ):
        technical.cell(
            row=header_row,
            column=column,
            value=header,
        )

    _style_header_row(
        technical,
        header_row,
        1,
        len(
            technical_headers
        ),
    )

    technical.freeze_panes = (
        "A5"
    )

    technical.sheet_view.showGridLines = (
        False
    )

    _set_widths(
        technical,
        {
            1: 21,
            2: 30,
            3: 34,
            4: 18,
            5: 15,
            6: 45,
            7: 45,
            8: 32,
            9: 15,
            10: 55,
            11: 55,
            12: 45,
        },
    )

    return (
        events,
        changes,
        technical,
        header_row,
    )


# ============================================================
# LLENADO DEL REPORTE
# ============================================================

def _fill_detail_sheets(
    workbook,
    queryset,
) -> None:
    (
        events_sheet,
        changes_sheet,
        technical_sheet,
        header_row,
    ) = _prepare_detail_sheets(
        workbook
    )

    event_row = (
        header_row + 1
    )

    change_row = (
        header_row + 1
    )

    technical_row = (
        header_row + 1
    )

    qs = (
        queryset
        .select_related(
            "actor"
        )
        .order_by(
            "-created_at",
            "-id",
        )
    )

    for event in qs.iterator(
        chunk_size=500
    ):
        local_dt = (
            _local_datetime(
                event.created_at
            )
        )

        actor_name = (
            _actor_name(event)
        )

        actor_email = (
            _actor_email(event)
        )

        before = (
            event.datos_anteriores
            if isinstance(
                event.datos_anteriores,
                dict,
            )
            else {}
        )

        after = (
            event.datos_nuevos
            if isinstance(
                event.datos_nuevos,
                dict,
            )
            else {}
        )

        differences = (
            _diff_objects(
                before,
                after,
            )
        )

        changed_fields = ", ".join(
            _field_label(
                item["path"]
            )
            for item in differences
        )

        # ====================================================
        # EVENTOS
        # ====================================================

        events_sheet.cell(
            row=event_row,
            column=1,
            value=(
                local_dt.date()
                if local_dt
                else None
            ),
        )

        events_sheet.cell(
            row=event_row,
            column=1,
        ).number_format = (
            "dd/mm/yyyy"
        )

        events_sheet.cell(
            row=event_row,
            column=2,
            value=(
                local_dt.time()
                .replace(
                    microsecond=0
                )
                if local_dt
                else None
            ),
        )

        events_sheet.cell(
            row=event_row,
            column=2,
        ).number_format = (
            "hh:mm:ss"
        )

        events_sheet.cell(
            row=event_row,
            column=3,
            value=_safe_text(
                actor_name
            ),
        )

        events_sheet.cell(
            row=event_row,
            column=4,
            value=_safe_text(
                actor_email
            ),
        )

        events_sheet.cell(
            row=event_row,
            column=5,
            value=_action_label(
                event.accion
            ),
        )

        events_sheet.cell(
            row=event_row,
            column=6,
            value=_module_label(
                event.modulo
            ),
        )

        events_sheet.cell(
            row=event_row,
            column=7,
            value=_entity_label(
                event.entidad_tipo,
                event.entidad_id,
            ),
        )

        events_sheet.cell(
            row=event_row,
            column=8,
            value=_safe_text(
                event.descripcion
            ),
        )

        events_sheet.cell(
            row=event_row,
            column=9,
            value=(
                changed_fields
                or "Sin cambios de datos"
            ),
        )

        event_row += 1

        # ====================================================
        # CAMBIOS
        # ====================================================

        for difference in differences:
            changes_sheet.cell(
                row=change_row,
                column=1,
                value=(
                    local_dt.date()
                    if local_dt
                    else None
                ),
            )

            changes_sheet.cell(
                row=change_row,
                column=1,
            ).number_format = (
                "dd/mm/yyyy"
            )

            changes_sheet.cell(
                row=change_row,
                column=2,
                value=(
                    local_dt.time()
                    .replace(
                        microsecond=0
                    )
                    if local_dt
                    else None
                ),
            )

            changes_sheet.cell(
                row=change_row,
                column=2,
            ).number_format = (
                "hh:mm:ss"
            )

            changes_sheet.cell(
                row=change_row,
                column=3,
                value=_safe_text(
                    actor_name
                ),
            )

            changes_sheet.cell(
                row=change_row,
                column=4,
                value=_action_label(
                    event.accion
                ),
            )

            changes_sheet.cell(
                row=change_row,
                column=5,
                value=_module_label(
                    event.modulo
                ),
            )

            changes_sheet.cell(
                row=change_row,
                column=6,
                value=_entity_label(
                    event.entidad_tipo,
                    event.entidad_id,
                ),
            )

            changes_sheet.cell(
                row=change_row,
                column=7,
                value=_field_label(
                    difference[
                        "path"
                    ]
                ),
            )

            changes_sheet.cell(
                row=change_row,
                column=8,
                value=_display_value(
                    difference[
                        "before"
                    ],
                    difference[
                        "path"
                    ],
                ),
            )

            changes_sheet.cell(
                row=change_row,
                column=9,
                value=_display_value(
                    difference[
                        "after"
                    ],
                    difference[
                        "path"
                    ],
                ),
            )

            change_row += 1

        # ====================================================
        # DATOS TÉCNICOS
        # ====================================================

        technical_sheet.cell(
            row=technical_row,
            column=1,
            value=local_dt,
        )

        technical_sheet.cell(
            row=technical_row,
            column=1,
        ).number_format = (
            "dd/mm/yyyy hh:mm:ss"
        )

        technical_sheet.cell(
            row=technical_row,
            column=2,
            value=_safe_text(
                actor_name
            ),
        )

        technical_sheet.cell(
            row=technical_row,
            column=3,
            value=_safe_text(
                actor_email
            ),
        )

        technical_sheet.cell(
            row=technical_row,
            column=4,
            value=_safe_text(
                event.ip
            ),
        )

        technical_sheet.cell(
            row=technical_row,
            column=5,
            value=_safe_text(
                event.metodo_http
            ),
        )

        technical_sheet.cell(
            row=technical_row,
            column=6,
            value=_safe_text(
                event.ruta
            ),
        )

        technical_sheet.cell(
            row=technical_row,
            column=7,
            value=_safe_text(
                event.user_agent
            ),
        )

        technical_sheet.cell(
            row=technical_row,
            column=8,
            value=_safe_text(
                event.entidad_tipo
            ),
        )

        technical_sheet.cell(
            row=technical_row,
            column=9,
            value=_safe_text(
                event.entidad_id
            ),
        )

        technical_sheet.cell(
            row=technical_row,
            column=10,
            value=_json_text(
                event.datos_anteriores
            ),
        )

        technical_sheet.cell(
            row=technical_row,
            column=11,
            value=_json_text(
                event.datos_nuevos
            ),
        )

        technical_sheet.cell(
            row=technical_row,
            column=12,
            value=_json_text(
                event.contexto
            ),
        )

        technical_row += 1

    # ========================================================
    # ESTILO FINAL
    # ========================================================

    _style_data_area(
        events_sheet,
        header_row + 1,
        event_row - 1,
        1,
        9,
    )

    _style_data_area(
        changes_sheet,
        header_row + 1,
        change_row - 1,
        1,
        9,
    )

    _style_data_area(
        technical_sheet,
        header_row + 1,
        technical_row - 1,
        1,
        12,
    )

    if (
        event_row
        > header_row + 1
    ):
        events_sheet.auto_filter.ref = (
            f"A{header_row}:"
            f"I{event_row - 1}"
        )

    if (
        change_row
        > header_row + 1
    ):
        changes_sheet.auto_filter.ref = (
            f"A{header_row}:"
            f"I{change_row - 1}"
        )

    if (
        technical_row
        > header_row + 1
    ):
        technical_sheet.auto_filter.ref = (
            f"A{header_row}:"
            f"L{technical_row - 1}"
        )


# ============================================================
# API DEL SERVICIO
# ============================================================

def build_auditoria_excel(
    queryset,
    params=None,
) -> bytes:
    """
    Construye el libro Excel de auditoría
    respetando el queryset ya filtrado.
    """

    workbook = Workbook()

    _build_summary_sheet(
        workbook,
        queryset,
        params,
    )

    _fill_detail_sheets(
        workbook,
        queryset,
    )

    workbook.properties.title = (
        "Reporte de Auditoría SGPC ULEAM"
    )

    workbook.properties.subject = (
        "Trazabilidad de movimientos del sistema"
    )

    workbook.properties.creator = (
        "SGPC ULEAM"
    )

    buffer = BytesIO()

    workbook.save(
        buffer
    )

    return buffer.getvalue()


def auditoria_excel_filename() -> str:
    return (
        timezone.localtime()
        .strftime(
            "auditoria_sgpc_%Y%m%d_%H%M%S.xlsx"
        )
    )