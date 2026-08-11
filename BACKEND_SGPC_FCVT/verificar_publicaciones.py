from io import BytesIO

from django.contrib.auth import get_user_model
from django.db.models import Q
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework.test import APIClient

from core.models import Publicacion


failures = []


def check(condition, label, detail=""):
    if condition:
        print(f"[OK]    {label}")
        return

    message = label

    if detail:
        message = f"{label}: {detail}"

    failures.append(message)
    print(f"[ERROR] {message}")


def response_items(response):
    try:
        payload = response.json()
    except Exception:
        return []

    if isinstance(payload, list):
        return payload

    if not isinstance(payload, dict):
        return []

    for key in ("results", "publicaciones", "items"):
        value = payload.get(key)

        if isinstance(value, list):
            return value

    return []


def response_total(response):
    try:
        payload = response.json()
    except Exception:
        return None

    if isinstance(payload, list):
        return len(payload)

    if not isinstance(payload, dict):
        return None

    if payload.get("count") is not None:
        try:
            return int(payload["count"])
        except (TypeError, ValueError):
            return None

    for key in ("results", "publicaciones", "items"):
        value = payload.get(key)

        if isinstance(value, list):
            return len(value)

    return None


def excel_data_rows(content):
    workbook = load_workbook(
        BytesIO(content),
        read_only=True,
        data_only=True,
    )

    total = 0

    try:
        for worksheet in workbook.worksheets:
            if worksheet.title == "Sin resultados":
                continue

            # Filas 1-3: título y descripción.
            # Fila 4: cabecera.
            total += max(int(worksheet.max_row or 0) - 4, 0)

    finally:
        workbook.close()

    return total


print("=" * 72)
print("PRUEBA INTEGRAL DE PUBLICACIONES SGPC ULEAM")
print("=" * 72)

User = get_user_model()

user = (
    User.objects
    .filter(is_active=True)
    .order_by(
        "-is_superuser",
        "-is_staff",
        "id",
    )
    .first()
)

check(
    user is not None,
    "Existe al menos un usuario activo",
)

if user is None:
    raise SystemExit(1)

print(f"Usuario utilizado: {user}")

client = APIClient()
client.force_authenticate(user=user)

try:
    urls = {
        "listado": reverse("publicaciones-list"),
        "mias": reverse("publicaciones-mias"),
        "preview": reverse("preview-publicaciones-excel"),
        "excel": reverse("exportar-publicaciones-excel"),
    }

    check(True, "Las rutas principales pueden resolverse")

except Exception as exc:
    check(False, "Resolución de rutas", repr(exc))
    raise SystemExit(1)


sample = (
    Publicacion.objects
    .select_related(
        "carrera",
        "carrera__facultad",
        "proyecto",
    )
    .order_by("id")
    .first()
)

cases = [
    ("Sin filtros", {}),
    ("Tipo AAI", {"tipo": "AAI"}),
    ("Tipo AR", {"tipo": "AR"}),
    ("Tipo PON", {"tipo": "PON"}),
    ("Tipo CAP", {"tipo": "CAP"}),
    ("Tipo LIB", {"tipo": "LIB"}),
]

if sample is not None:
    if sample.anio_publicacion:
        cases.append(
            (
                "Año existente",
                {
                    "anio": sample.anio_publicacion,
                },
            )
        )

    if sample.carrera_id:
        cases.append(
            (
                "Carrera existente",
                {
                    "carrera": sample.carrera_id,
                },
            )
        )

    facultad_id = getattr(
        getattr(sample, "carrera", None),
        "facultad_id",
        None,
    )

    if facultad_id:
        cases.append(
            (
                "Facultad existente",
                {
                    "facultad": facultad_id,
                },
            )
        )

    if sample.proyecto_id:
        cases.append(
            (
                "Proyecto existente",
                {
                    "proyecto": sample.proyecto_id,
                },
            )
        )


print()
print("-" * 72)
print("LISTADO Y VISTA PREVIA")
print("-" * 72)

base_total = None
base_response = None

for label, params in cases:
    response = client.get(
        urls["listado"],
        params,
    )

    check(
        response.status_code == 200,
        f"{label}: listado responde 200",
        f"HTTP {response.status_code}",
    )

    if response.status_code != 200:
        try:
            print(response.json())
        except Exception:
            print(response.content[:500])

        continue

    total_listado = response_total(
        response
    )

    check(
        total_listado is not None,
        f"{label}: respuesta reconocible",
    )

    preview_response = client.get(
        urls["preview"],
        params,
    )

    check(
        preview_response.status_code == 200,
        f"{label}: preview responde 200",
        f"HTTP {preview_response.status_code}",
    )

    if preview_response.status_code != 200:
        try:
            print(preview_response.json())
        except Exception:
            print(preview_response.content[:500])

        continue

    preview_payload = preview_response.json()
    preview_total = preview_payload.get("total")

    try:
        preview_total = int(preview_total)
    except (TypeError, ValueError):
        preview_total = None

    check(
        preview_total is not None,
        f"{label}: preview contiene total",
    )

    check(
        total_listado == preview_total,
        f"{label}: listado y preview coinciden",
        (
            f"listado={total_listado}, "
            f"preview={preview_total}"
        ),
    )

    print(
        f"        Resultados encontrados: "
        f"{total_listado}"
    )

    if label == "Sin filtros":
        base_total = total_listado
        base_response = response


print()
print("-" * 72)
print("CONTRATO DEL SERIALIZER")
print("-" * 72)

if base_response is not None:
    items = response_items(
        base_response
    )

    if items:
        required_fields = {
            "id",
            "titulo",
            "autor",
            "facultad_id",
            "facultad",
            "carrera_id",
            "carrera",
            "proyecto_id",
            "proyecto",
            "tipo_publicacion_final",
            "tipo_publicacion_final_label",
            "tiene_pdf",
            "pdf_endpoint",
            "puede_editar",
        }

        received_fields = set(
            items[0].keys()
        )

        missing_fields = (
            required_fields
            - received_fields
        )

        check(
            not missing_fields,
            "El serializer contiene todos los campos nuevos",
            (
                "Faltan: "
                + ", ".join(
                    sorted(missing_fields)
                )
                if missing_fields
                else ""
            ),
        )

    else:
        print(
            "[AVISO] No existen publicaciones para "
            "comprobar los campos del serializer."
        )


print()
print("-" * 72)
print("MIS PUBLICACIONES")
print("-" * 72)

mias_response = client.get(
    urls["mias"]
)

check(
    mias_response.status_code == 200,
    "Mis publicaciones responde 200",
    f"HTTP {mias_response.status_code}",
)

if mias_response.status_code == 200:
    mias_total = response_total(
        mias_response
    )

    check(
        mias_total is not None,
        "Mis publicaciones devuelve una respuesta reconocible",
    )

    if (
        mias_total is not None
        and base_total is not None
    ):
        check(
            mias_total <= base_total,
            "Mis publicaciones es subconjunto del listado",
            (
                f"mias={mias_total}, "
                f"institucional={base_total}"
            ),
        )

    print(
        f"        Publicaciones del usuario: "
        f"{mias_total}"
    )


print()
print("-" * 72)
print("FILTROS INVÁLIDOS")
print("-" * 72)

invalid_cases = [
    (
        "Tipo inválido",
        {
            "tipo": "TIPO_INEXISTENTE",
        },
    ),
    (
        "Año inválido",
        {
            "anio": "abc",
        },
    ),
    (
        "Booleano inválido",
        {
            "solo_con_pdf": "quizas",
        },
    ),
]

for label, params in invalid_cases:
    response = client.get(
        urls["listado"],
        params,
    )

    check(
        response.status_code == 400,
        f"{label} devuelve 400",
        f"HTTP {response.status_code}",
    )


print()
print("-" * 72)
print("EXPORTACIÓN EXCEL")
print("-" * 72)

excel_cases = [
    ("Excel sin filtros", {}),
]

if sample is not None and sample.anio_publicacion:
    excel_cases.append(
        (
            "Excel por año",
            {
                "anio": sample.anio_publicacion,
            },
        )
    )

for label, params in excel_cases:
    preview_response = client.get(
        urls["preview"],
        params,
    )

    excel_response = client.get(
        urls["excel"],
        params,
    )

    check(
        preview_response.status_code == 200,
        f"{label}: preview responde 200",
        f"HTTP {preview_response.status_code}",
    )

    check(
        excel_response.status_code == 200,
        f"{label}: descarga responde 200",
        f"HTTP {excel_response.status_code}",
    )

    if (
        preview_response.status_code != 200
        or excel_response.status_code != 200
    ):
        continue

    content_type = (
        excel_response.get(
            "Content-Type",
            "",
        )
    )

    check(
        "spreadsheetml" in content_type,
        f"{label}: Content-Type de Excel correcto",
        content_type,
    )

    try:
        preview_total = int(
            preview_response.json()["total"]
        )

        workbook_rows = excel_data_rows(
            excel_response.content
        )

        check(
            workbook_rows == preview_total,
            f"{label}: filas del Excel y preview coinciden",
            (
                f"excel={workbook_rows}, "
                f"preview={preview_total}"
            ),
        )

    except Exception as exc:
        check(
            False,
            f"{label}: abrir y analizar archivo Excel",
            repr(exc),
        )


print()
print("-" * 72)
print("PDF AUTENTICADO")
print("-" * 72)

pdf_publicacion = (
    Publicacion.objects
    .filter(
        (
            Q(
                archivo_pdf__isnull=False
            )
            & ~Q(
                archivo_pdf=""
            )
        )
        |
        (
            Q(
                archivos__archivo__isnull=False
            )
            & ~Q(
                archivos__archivo=""
            )
        )
    )
    .distinct()
    .order_by("id")
    .first()
)

if pdf_publicacion is None:
    print(
        "[AVISO] No existe una publicación con PDF "
        "para probar el endpoint."
    )

else:
    pdf_url = reverse(
        "publicacion-pdf-inline",
        kwargs={
            "id": pdf_publicacion.id,
        },
    )

    pdf_response = client.get(
        pdf_url
    )

    check(
        pdf_response.status_code == 200,
        "Endpoint PDF autenticado responde 200",
        (
            f"publicacion={pdf_publicacion.id}, "
            f"HTTP {pdf_response.status_code}"
        ),
    )

    if pdf_response.status_code == 200:
        pdf_content_type = (
            pdf_response.get(
                "Content-Type",
                "",
            )
            .lower()
        )

        check(
            (
                "pdf" in pdf_content_type
                or "octet-stream"
                in pdf_content_type
            ),
            "El endpoint devuelve contenido de archivo",
            pdf_content_type,
        )

    try:
        pdf_response.close()
    except Exception:
        pass


print()
print("=" * 72)

if failures:
    print(
        f"PRUEBA FINALIZADA CON "
        f"{len(failures)} ERROR(ES)"
    )

    for index, failure in enumerate(
        failures,
        start=1,
    ):
        print(
            f"{index}. {failure}"
        )

    print("=" * 72)
    raise SystemExit(1)

print("TODAS LAS PRUEBAS PASARON CORRECTAMENTE")
print("=" * 72)
